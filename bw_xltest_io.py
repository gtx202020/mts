import openpyxl
import ast
import logging
from typing import Dict, List, Optional, Any, Tuple
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 로깅 설정
logger = logging.getLogger(__name__)


class ExcelReader:
    """Excel 입력 파일을 읽는 클래스"""
    
    def __init__(self, file_path: str):
        """초기화
        
        Args:
            file_path: 읽을 Excel 파일 경로
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def open_workbook(self) -> bool:
        """워크북 열기
        
        Returns:
            성공 여부
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            logger.info(f"Excel 파일 열기 성공: {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"Excel 파일 열기 실패: {str(e)}")
            return False
    
    def read_interface_info(self, start_col: int) -> Optional[Dict[str, Any]]:
        """인터페이스 정보 읽기
        
        Args:
            start_col: 시작 컬럼 번호 (1-based)
        
        Returns:
            인터페이스 정보 딕셔너리 또는 None
        """
        if not self.worksheet:
            return None
        
        try:
            interface_info = {
                'interface_name': self.worksheet.cell(row=1, column=start_col).value or '',
                'interface_id': self.worksheet.cell(row=2, column=start_col).value or '',
                'send': {'db_info': None, 'table_info': None, 'columns': []},
                'recv': {'db_info': None, 'table_info': None, 'columns': []}
            }
            
            # DB 정보 파싱 (3행)
            send_db_info = self.parse_db_info(
                self.worksheet.cell(row=3, column=start_col).value
            )
            recv_db_info = self.parse_db_info(
                self.worksheet.cell(row=3, column=start_col + 1).value
            )
            
            if not send_db_info or not recv_db_info:
                logger.error("DB 연결 정보 파싱 실패")
                return None
            
            interface_info['send']['db_info'] = send_db_info
            interface_info['recv']['db_info'] = recv_db_info
            
            # 테이블 정보 파싱 (4행)
            send_table_info = self.parse_table_info(
                self.worksheet.cell(row=4, column=start_col).value
            )
            recv_table_info = self.parse_table_info(
                self.worksheet.cell(row=4, column=start_col + 1).value
            )
            
            if not send_table_info or not recv_table_info:
                logger.error("테이블 정보 파싱 실패")
                return None
            
            interface_info['send']['table_info'] = send_table_info
            interface_info['recv']['table_info'] = recv_table_info
            
            # 컬럼 매핑 읽기 (5행부터)
            send_columns, recv_columns = self.read_column_mappings(start_col)
            interface_info['send']['columns'] = send_columns
            interface_info['recv']['columns'] = recv_columns
            
            return interface_info
            
        except Exception as e:
            logger.error(f"인터페이스 정보 읽기 실패: {str(e)}")
            return None
    
    def parse_db_info(self, cell_value: Any) -> Optional[Dict[str, str]]:
        """DB 연결 정보 파싱
        
        Args:
            cell_value: 셀 값
        
        Returns:
            DB 정보 딕셔너리 또는 None
        """
        if not cell_value:
            return None
        
        try:
            db_info = ast.literal_eval(str(cell_value))
            if isinstance(db_info, dict) and all(k in db_info for k in ['sid', 'username', 'password']):
                return db_info
            else:
                logger.error(f"잘못된 DB 정보 형식: {cell_value}")
                return None
        except Exception as e:
            logger.error(f"DB 정보 파싱 오류: {str(e)}")
            return None
    
    def parse_table_info(self, cell_value: Any) -> Optional[Dict[str, str]]:
        """테이블 정보 파싱
        
        Args:
            cell_value: 셀 값
        
        Returns:
            테이블 정보 딕셔너리 또는 None
        """
        if not cell_value:
            return None
        
        try:
            table_info = ast.literal_eval(str(cell_value))
            if isinstance(table_info, dict) and all(k in table_info for k in ['owner', 'table_name']):
                return table_info
            else:
                logger.error(f"잘못된 테이블 정보 형식: {cell_value}")
                return None
        except Exception as e:
            logger.error(f"테이블 정보 파싱 오류: {str(e)}")
            return None
    
    def read_column_mappings(self, start_col: int) -> Tuple[List[str], List[str]]:
        """컬럼 매핑 정보 읽기
        
        Args:
            start_col: 시작 컬럼 번호
        
        Returns:
            (송신 컬럼 리스트, 수신 컬럼 리스트)
        """
        send_columns = []
        recv_columns = []
        row = 5
        
        while row <= self.worksheet.max_row:
            send_col = self.worksheet.cell(row=row, column=start_col).value
            recv_col = self.worksheet.cell(row=row, column=start_col + 1).value
            
            # 둘 다 비어있으면 종료
            if not send_col and not recv_col:
                break
            
            send_columns.append(send_col if send_col else '')
            recv_columns.append(recv_col if recv_col else '')
            row += 1
        
        return send_columns, recv_columns
    
    def close_workbook(self):
        """워크북 닫기"""
        if self.workbook:
            self.workbook.close()
            logger.info("Excel 파일 닫기 완료")


class ExcelWriter:
    """결과를 Excel 파일로 출력하는 클래스"""
    
    def __init__(self, file_path: str):
        """초기화
        
        Args:
            file_path: 출력할 Excel 파일 경로
        """
        self.file_path = file_path
        self.workbook = None
        
        # 스타일 정의
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True, size=9)
        self.normal_font = Font(name='맑은 고딕', size=9)
        self.bold_font = Font(bold=True, size=9)
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 상태별 색상
        self.status_colors = {
            '정상': PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'),
            '경고': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            '오류': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        }
    
    def create_workbook(self):
        """워크북 생성"""
        self.workbook = openpyxl.Workbook()
        # 기본 시트 제거
        self.workbook.remove(self.workbook.active)
        logger.info("출력 워크북 생성 완료")
    
    def write_interface_result(self, interface_info: Dict[str, Any], 
                             validation_results: List[Dict[str, Any]], 
                             interface_num: int):
        """인터페이스 결과를 시트에 기록
        
        Args:
            interface_info: 인터페이스 정보
            validation_results: 검증 결과
            interface_num: 인터페이스 번호
        """
        interface_name = interface_info.get('interface_name', '').strip()
        interface_id = interface_info.get('interface_id', '').strip()
        sheet_name = f'{interface_num}_{interface_name}' if interface_name else f'Interface_{interface_num}'
        
        # 시트명 길이 제한 (31자)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        ws = self.workbook.create_sheet(sheet_name)
        
        # 인터페이스 정보 작성
        self._write_interface_info(ws, interface_info)
        
        # 컬럼 비교 결과 작성
        self._write_comparison_results(ws, validation_results, start_row=5)
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
        
        logger.info(f"시트 '{sheet_name}' 작성 완료")
    
    def _write_interface_info(self, ws, interface_info: Dict[str, Any]):
        """인터페이스 기본 정보 작성"""
        # 헤더
        ws['A1'] = '인터페이스 기본 정보'
        ws['A1'].fill = self.header_fill
        ws['A1'].font = self.header_font
        ws['A1'].alignment = self.center_alignment
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 25
        
        # 인터페이스명과 ID
        ws['A2'] = '인터페이스명'
        ws['A2'].alignment = self.center_alignment
        ws['A2'].font = self.bold_font
        ws.merge_cells('B2:D2')
        ws['B2'] = interface_info.get('interface_name', '')
        ws['B2'].alignment = self.left_alignment
        
        ws['E2'] = '인터페이스 ID'
        ws['E2'].alignment = self.center_alignment
        ws['E2'].font = self.bold_font
        ws.merge_cells('F2:H2')
        ws['F2'] = interface_info.get('interface_id', '')
        ws['F2'].alignment = self.left_alignment
        
        # 송신 테이블 정보
        ws['A3'] = '송신 테이블'
        ws['A3'].alignment = self.center_alignment
        ws['A3'].font = self.bold_font
        ws.merge_cells('B3:D3')
        send_table = interface_info['send']['table_info']
        ws['B3'] = f"{send_table['owner']}.{send_table['table_name']}"
        ws['B3'].alignment = self.left_alignment
        
        # 수신 테이블 정보
        ws['E3'] = '수신 테이블'
        ws['E3'].alignment = self.center_alignment
        ws['E3'].font = self.bold_font
        ws.merge_cells('F3:H3')
        recv_table = interface_info['recv']['table_info']
        ws['F3'] = f"{recv_table['owner']}.{recv_table['table_name']}"
        ws['F3'].alignment = self.left_alignment
        
        # 테두리 적용
        for row in range(1, 4):
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
    
    def _write_comparison_results(self, ws, results: List[Dict[str, Any]], start_row: int):
        """컬럼 비교 결과 작성"""
        # 헤더
        ws[f'A{start_row}'] = '컬럼 비교 결과'
        ws[f'A{start_row}'].fill = self.header_fill
        ws[f'A{start_row}'].font = self.header_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws.merge_cells(f'A{start_row}:J{start_row}')
        ws.row_dimensions[start_row].height = 25
        
        # 컬럼 헤더
        headers = [
            '송신 컬럼', '송신 타입', '송신 크기', '송신 Null여부',
            '수신 컬럼', '수신 타입', '수신 크기', '수신 Null여부',
            '비교 결과', '상태'
        ]
        
        header_row = start_row + 1
        for idx, header in enumerate(headers):
            col = chr(ord('A') + idx)
            ws[f'{col}{header_row}'] = header
            ws[f'{col}{header_row}'].fill = PatternFill(
                start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'
            )
            ws[f'{col}{header_row}'].font = self.bold_font
            ws[f'{col}{header_row}'].alignment = self.center_alignment
            ws[f'{col}{header_row}'].border = self.border
        ws.row_dimensions[header_row].height = 20
        
        # 데이터 행
        data_row = header_row + 1
        for result in results:
            # 송신 정보
            ws[f'A{data_row}'] = result.get('send_column', '')
            if result.get('send_info'):
                ws[f'B{data_row}'] = result['send_info'].get('type', '')
                ws[f'C{data_row}'] = result['send_info'].get('size', '')
                ws[f'D{data_row}'] = result['send_info'].get('nullable', '')
            
            # 수신 정보
            ws[f'E{data_row}'] = result.get('recv_column', '')
            if result.get('recv_info'):
                ws[f'F{data_row}'] = result['recv_info'].get('type', '')
                ws[f'G{data_row}'] = result['recv_info'].get('size', '')
                ws[f'H{data_row}'] = result['recv_info'].get('nullable', '')
            
            # 비교 결과
            messages = []
            if result.get('errors'):
                messages.extend(result['errors'])
            if result.get('warnings'):
                messages.extend(result['warnings'])
            ws[f'I{data_row}'] = '\n'.join(messages)
            
            # 상태
            status = result.get('status', '정상')
            ws[f'J{data_row}'] = status
            ws[f'J{data_row}'].fill = self.status_colors.get(status, self.status_colors['정상'])
            
            # 스타일 적용
            for col in range(ord('A'), ord('K')):
                col_letter = chr(col)
                cell = ws[f'{col_letter}{data_row}']
                cell.alignment = self.left_alignment if col_letter == 'I' else self.center_alignment
                cell.font = self.normal_font
                cell.border = self.border
            
            data_row += 1
    
    def _adjust_column_widths(self, ws):
        """열 너비 조정"""
        column_widths = {
            'A': 20,  # 송신 컬럼
            'B': 15,  # 송신 타입
            'C': 12,  # 송신 크기
            'D': 12,  # 송신 Null여부
            'E': 20,  # 수신 컬럼
            'F': 15,  # 수신 타입
            'G': 12,  # 수신 크기
            'H': 12,  # 수신 Null여부
            'I': 40,  # 비교 결과
            'J': 10   # 상태
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def save_workbook(self):
        """워크북 저장"""
        if self.workbook:
            self.workbook.save(self.file_path)
            logger.info(f"결과 파일 저장 완료: {self.file_path}")
    
    def close_workbook(self):
        """워크북 닫기"""
        if self.workbook:
            self.workbook.close()
            logger.info("출력 워크북 닫기 완료")