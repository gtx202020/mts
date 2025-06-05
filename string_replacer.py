import openpyxl
import yaml
import difflib
import os
import datetime
import pandas as pd
import shutil
import re

# 디버그 모드 설정
DEBUG_MODE = True  # 디버그 정보 출력 여부를 제어하는 플래그

def debug_print(*args, **kwargs):
    """디버그 모드일 때만 메시지를 출력하는 함수"""
    if DEBUG_MODE:
        print("[DEBUG]", *args, **kwargs)

# chardet 모듈 가져오기 시도 (설치되지 않았을 경우 대비)
try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    print("경고: chardet 모듈을 찾을 수 없습니다. 기본 인코딩(utf-8)을 사용합니다.")
    print("pip install chardet 명령어로 설치할 수 있습니다.")
    HAS_CHARDET = False

def generate_yaml_from_excel(excel_path, yaml_path):
    """엑셀 파일을 읽어 YAML 파일을 생성한다."""
    # pandas로 엑셀 파일 읽기
    df = pd.read_excel(excel_path, engine='openpyxl')
    
    # 전체 YAML 구조를 저장할 딕셔너리
    full_yaml_structure = {}
    
    # 2행씩 처리 (일반행, 매칭행)
    for i in range(0, len(df), 2):
        if i + 1 >= len(df):  # 마지막 행이 홀수인 경우 처리
            break
            
        normal_row = df.iloc[i]  # 일반행
        match_row = df.iloc[i+1]  # 매칭행
        
        print(f"\n=== {i//2 + 1}번째 행 쌍 ===")
        
        # 파일 생성 여부 확인 및 YAML 구조 생성
        yaml_structure = {
            f"{i//2 + 1}번째 행": {}
        }
        
        def modify_path(path):
            """파일 경로를 수정하는 함수 (테스트용)"""
            if isinstance(path, str) and path.startswith("C:\\BwProject"):
                return path.replace("C:\\BwProject", "C:\\TBwProject")
            return path

        def extract_filename(path):
            """파일 경로에서 파일명만 추출"""
            if not isinstance(path, str):
                return ""
            return os.path.basename(path)

        def extract_process_filename(path):
            """프로세스 파일 경로에서 파일명만 추출"""
            if not isinstance(path, str):
                return ""
            return os.path.basename(path)

        def extract_existing_namespace(source_file_path, base_name):
            """
            소스 파일에서 기존 namespace를 추출
            
            Args:
                source_file_path: 소스 파일 경로
                base_name: 스키마 파일의 기본 이름
                
            Returns:
                기존 namespace 또는 None
            """
            if not source_file_path or not os.path.exists(source_file_path):
                return None
                
            try:
                with open(source_file_path, 'r', encoding='utf-8') as f:
                    source_content = f.read()
                    
                # namespace 패턴 찾기
                namespace_pattern = f'namespace\\s*=\\s*"([^"]*{base_name}[^"]*)"'
                match = re.search(namespace_pattern, source_content)
                
                if match:
                    return match.group(1)
                
                # xmlns:pfx3 패턴도 확인
                xmlns_pattern = f'xmlns:pfx3\\s*=\\s*"([^"]*{base_name}[^"]*)"'
                match = re.search(xmlns_pattern, source_content)
                
                if match:
                    return match.group(1)
                    
            except Exception as e:
                debug_print(f"기존 namespace 추출 중 오류: {e}")
                
            return None

        def process_schema_path(schema_path, preserve_no_namespace=False):
            """
            스키마 파일 경로를 처리하여 namespace와 schemaLocation 생성
            
            Args:
                schema_path: 스키마 파일 경로
                preserve_no_namespace: no_namespace_schema 경로 보존 여부
            """
            if not isinstance(schema_path, str):
                return None, None
            
            # 1. 경로 구분자 변경
            normalized_path = schema_path.replace('\\', '/')
            
            # 2. '/SharedResources' 이후 부분 추출
            shared_idx = normalized_path.find('/SharedResources')
            if shared_idx == -1:
                return None, None
                
            # BB 부분을 포함한 경로 추출
            bb_start = normalized_path.rfind('/', 0, shared_idx)
            if bb_start == -1:
                return None, None
                
            relative_path = normalized_path[bb_start:]  # /BB/SharedResources/...
            schema_location = relative_path[relative_path.find('/SharedResources'):]  # /SharedResources/...
            
            # no_namespace_schema 처리 로직
            if preserve_no_namespace:
                # SharedResources 이전 부분은 그대로 유지하고 SharedResources 이후만 새로운 로직 적용
                namespace = f"http://www.tibco.com/ns/no_namespace_schema{schema_location}"
            else:
                # 기존 로직 그대로 적용
                namespace = f"http://www.tibco.com/schemas{relative_path}"
            
            return namespace, schema_location

        def create_schema_replacements(filename, schema_path, source_file_path=None):
            """스키마 파일 치환 목록 생성"""
            if not filename.endswith('.xsd'):
                return []
            
            base_name = os.path.splitext(filename)[0]
            
            # 소스 파일에서 기존 namespace 확인
            has_no_namespace = False
            if source_file_path:
                existing_namespace = extract_existing_namespace(source_file_path, base_name)
                if existing_namespace and 'no_namespace_schema' in existing_namespace:
                    has_no_namespace = True
                    debug_print(f"no_namespace_schema 감지됨: {existing_namespace}")
            
            # namespace 생성 (no_namespace 여부에 따라 다르게 처리)
            namespace, schema_location = process_schema_path(schema_path, preserve_no_namespace=has_no_namespace)
            if not namespace or not schema_location:
                return []
            return [{
                "설명": "스키마 namespace 치환",
                "찾기": {
                    "정규식": f'namespace\\s*=\\s*"[^"]*{base_name}[^"]*"'
                },
                "교체": {
                    "값": f'namespace="{namespace}"'
                }
            },
            {
                "설명": "스키마 schemaLocation 치환",
                "찾기": {
                    "정규식": f'schemaLocation\\s*=\\s*"[^"]*{base_name}[^"]*"'
                },
                "교체": {
                    "값": f'schemaLocation="{schema_location}"'
                }
            },
            {
                "설명": "ProcessDefinition namespace 치환",
                "찾기": {
                    "정규식": f'xmlns:pfx3\\s*=\\s*"[^"]*{base_name}[^"]*"'
                },
                "교체": {
                    "값": f'xmlns:pfx3="{namespace}"'
                }
            }]
        
        def extract_process_path(file_path):
            """프로세스 파일 경로에서 'Processes' 이후의 경로를 추출하고 디렉토리 구분자를 변경"""
            if not isinstance(file_path, str):
                return ""
            
            # 디렉토리 구분자를 '/'로 통일
            normalized_path = file_path.replace('\\', '/')
            
            # 'Processes' 위치 찾기
            processes_idx = normalized_path.find('Processes/')
            if processes_idx == -1:
                return ""
            
            # 'Processes/' 이후의 경로 추출
            relative_path = normalized_path[processes_idx + len('Processes/'):]
            
            return relative_path

        def create_process_replacements(source_path, target_path, match_row, normal_row):
            """프로세스 파일의 치환 목록 생성"""
            if not isinstance(source_path, str) or not isinstance(target_path, str):
                return []
            
            # 매칭행의 파일명으로 패턴 매칭 (찾을 패턴)
            source_filename = extract_process_filename(source_path)
            # 기본행의 경로에서 Processes 이후 경로 추출 (교체할 값)
            target_process_path = extract_process_path(target_path)
            
            if not source_filename or not target_process_path:
                return []
            
            replacements = [{
                "설명": "프로세스 이름 치환",
                "찾기": {
                    "정규식": f'<pd:name>Processes/[^<]*</pd:name>'
                },
                "교체": {
                    "값": f'<pd:name>Processes/{target_process_path}</pd:name>'
                }
            }]

            # 고정 문자열 치환 규칙 추가
            fixed_replacements = [
                {
                    "설명": "LHMES_MGR 치환",
                    "찾기": {
                        "정규식": "LHMES_MGR"
                    },
                    "교체": {
                        "값": "LYMES_MGR"
                    }
                },
                {
                    "설명": "VOMES_MGR 치환",
                    "찾기": {
                        "정규식": "VOMES_MGR"
                    },
                    "교체": {
                        "값": "LZMES_MGR"
                    }
                },
                {
                    "설명": "LH 문자열 치환",
                    "찾기": {
                        "정규식": "'LH'"
                    },
                    "교체": {
                        "값": "'LY'"
                    }
                },
                {
                    "설명": "VO 문자열 치환",
                    "찾기": {
                        "정규식": "'VO'"
                    },
                    "교체": {
                        "값": "'LZ'"
                    }
                },
                {
                    "설명": "LH 따옴표 문자열 치환",
                    "찾기": {
                        "정규식": "&quot;LH&quot;"
                    },
                    "교체": {
                        "값": "&quot;LY&quot;"
                    }
                },
                {
                    "설명": "VO 따옴표 문자열 치환",
                    "찾기": {
                        "정규식": "&quot;VO&quot;"
                    },
                    "교체": {
                        "값": "&quot;LZ&quot;"
                    }
                }
            ]
            replacements.extend(fixed_replacements)

            # IFID와 수신업무명 조합 치환 규칙 추가
            origin_ifid_with_susin = f"{match_row['Group ID']}.{match_row['Event_ID']}.{match_row['수신\n업무명']}"
            dest_ifid_with_susin = f"{normal_row['Group ID']}.{match_row['Event_ID']}.{normal_row['수신\n업무명']}"
            
            # IFID와 수신업무명 조합이 다른 경우에만 치환 규칙 추가
            if origin_ifid_with_susin != dest_ifid_with_susin:
                replacements.append({
                    "설명": "IFID와 수신업무명 조합 치환",
                    "찾기": {
                        "정규식": origin_ifid_with_susin.replace(".", "\\.")
                    },
                    "교체": {
                        "값": dest_ifid_with_susin
                    }
                })

            # IFID 치환 규칙 추가
            origin_ifid = f"{match_row['Group ID']}.{match_row['Event_ID']}"
            dest_ifid = f"{normal_row['Group ID']}.{match_row['Event_ID']}"
            
            # IFID가 다른 경우에만 치환 규칙 추가
            if origin_ifid != dest_ifid:
                replacements.append({
                    "설명": "IFID 치환",
                    "찾기": {
                        "정규식": origin_ifid.replace(".", "\\.")
                    },
                    "교체": {
                        "값": dest_ifid
                    }
                })

            # Event_ID 치환 규칙 추가
            if match_row['Event_ID'] != normal_row['Event_ID']:
                replacements.append({
                    "설명": "Event_ID 치환",
                    "찾기": {
                        "정규식": match_row['Event_ID']
                    },
                    "교체": {
                        "값": normal_row['Event_ID']
                    }
                })

            # Group ID &quot; 형식 치환 규칙 추가
            if match_row['Group ID'] != normal_row['Group ID']:
                replacements.append({
                    "설명": "Group ID &quot; 형식 치환",
                    "찾기": {
                        "정규식": f'&quot;{match_row["Group ID"]}&quot;'
                    },
                    "교체": {
                        "값": f'&quot;{normal_row["Group ID"]}&quot;'
                    }
                })

            # 송신업무명 &quot; 형식 치환 규칙 추가
            if match_row['송신\n업무명'] != normal_row['송신\n업무명']:
                replacements.append({
                    "설명": "송신업무명 &quot; 형식 치환",
                    "찾기": {
                        "정규식": f'&quot;{match_row["송신\n업무명"]}&quot;'
                    },
                    "교체": {
                        "값": f'&quot;{normal_row["송신\n업무명"]}&quot;'
                    }
                })

            # 수신업무명 &quot; 형식 치환 규칙 추가
            if match_row['수신\n업무명'] != normal_row['수신\n업무명']:
                replacements.append({
                    "설명": "수신업무명 &quot; 형식 치환",
                    "찾기": {
                        "정규식": f'&quot;{match_row["수신\n업무명"]}&quot;'
                    },
                    "교체": {
                        "값": f'&quot;{normal_row["수신\n업무명"]}&quot;'
                    }
                })

            # 송신업무명 치환 규칙 추가 (pd:from/to Check 형식)
            if match_row['송신\n업무명'] != normal_row['송신\n업무명']:
                # pd:from 태그 치환
                replacements.append({
                    "설명": "송신업무명 from 태그 치환",
                    "찾기": {
                        "정규식": f'(<pd:from>Check {match_row["송신\n업무명"]})'
                    },
                    "교체": {
                        "값": f'<pd:from>Check {normal_row["송신\n업무명"]}'
                    }
                })
                # pd:to 태그 치환
                replacements.append({
                    "설명": "송신업무명 to 태그 치환",
                    "찾기": {
                        "정규식": f'(<pd:to>Check {match_row["송신\n업무명"]})'
                    },
                    "교체": {
                        "값": f'<pd:to>Check {normal_row["송신\n업무명"]}'
                    }
                })
                # pd:activity name 태그 치환
                replacements.append({
                    "설명": "송신업무명 activity name 태그 치환",
                    "찾기": {
                        "정규식": f'(<pd:activity\\s+name="Check {match_row["송신\n업무명"]})'
                    },
                    "교체": {
                        "값": f'<pd:activity name="Check {normal_row["송신\n업무명"]}'
                    }
                })
                # sharedjdbc 치환
                replacements.append({
                    "설명": "송신업무명 sharedjdbc 치환",
                    "찾기": {
                        "정규식": f'{match_row["송신\n업무명"]}\\.sharedjdbc'
                    },
                    "교체": {
                        "값": f'{normal_row["송신\n업무명"]}.sharedjdbc'
                    }
                })

            # 수신업무명 치환 규칙 추가 (pd:from/to Check 형식)
            if match_row['수신\n업무명'] != normal_row['수신\n업무명']:
                # pd:from 태그 치환
                replacements.append({
                    "설명": "수신업무명 from 태그 치환",
                    "찾기": {
                        "정규식": f'(<pd:from>Check {match_row["수신\n업무명"]})'
                    },
                    "교체": {
                        "값": f'<pd:from>Check {normal_row["수신\n업무명"]}'
                    }
                })
                # pd:to 태그 치환
                replacements.append({
                    "설명": "수신업무명 to 태그 치환",
                    "찾기": {
                        "정규식": f'(<pd:to>Check {match_row["수신\n업무명"]})'
                    },
                    "교체": {
                        "값": f'<pd:to>Check {normal_row["수신\n업무명"]}'
                    }
                })
                # pd:activity name 태그 치환
                replacements.append({
                    "설명": "수신업무명 activity name 태그 치환",
                    "찾기": {
                        "정규식": f'(<pd:activity\\s+name="Check {match_row["수신\n업무명"]})'
                    },
                    "교체": {
                        "값": f'<pd:activity name="Check {normal_row["수신\n업무명"]}'
                    }
                })
                # sharedjdbc 치환
                replacements.append({
                    "설명": "수신업무명 sharedjdbc 치환",
                    "찾기": {
                        "정규식": f'{match_row["수신\n업무명"]}\\.sharedjdbc'
                    },
                    "교체": {
                        "값": f'{normal_row["수신\n업무명"]}.sharedjdbc'
                    }
                })
            
            return replacements

        # 1. 송신파일경로 처리
        if pd.notna(normal_row.get('송신파일생성여부')) and float(normal_row['송신파일생성여부']) == 1.0:
            yaml_structure[f"{i//2 + 1}번째 행"]["송신파일경로"] = {
                "원본파일": match_row['송신파일경로'],
                "복사파일": modify_path(normal_row['송신파일경로']),  # 경로 수정
                "치환목록": create_schema_replacements(
                    extract_filename(normal_row['송신스키마파일명']),
                    normal_row['송신스키마파일명'],
                    match_row['송신파일경로']  # 소스 파일 경로 전달
                ) + create_process_replacements(
                    match_row['송신파일경로'],    # 매칭행의 경로로 패턴 매칭
                    normal_row['송신파일경로'],    # 기본행의 경로로 교체
                    match_row,
                    normal_row
                )
            }
            print("\n[송신파일경로 생성]")
            print(f"  원본파일: {match_row['송신파일경로']}")
            print(f"  복사파일: {modify_path(normal_row['송신파일경로'])}")
        
        # 2. 수신파일경로 처리
        if pd.notna(normal_row.get('수신파일생성여부')) and float(normal_row['수신파일생성여부']) == 1.0:
            yaml_structure[f"{i//2 + 1}번째 행"]["수신파일경로"] = {
                "원본파일": match_row['수신파일경로'],
                "복사파일": modify_path(normal_row['수신파일경로']),  # 경로 수정
                "치환목록": create_schema_replacements(
                    extract_filename(normal_row['수신스키마파일명']),
                    normal_row['송신스키마파일명'],
                    match_row['수신파일경로']  # 소스 파일 경로 전달
                ) + create_process_replacements(
                    match_row['수신파일경로'],    # 매칭행의 경로로 패턴 매칭
                    normal_row['수신파일경로'],    # 기본행의 경로로 교체
                    match_row,
                    normal_row
                )
            }
            print("\n[수신파일경로 생성]")
            print(f"  원본파일: {match_row['수신파일경로']}")
            print(f"  복사파일: {modify_path(normal_row['수신파일경로'])}")
        
        # 3. 송신스키마파일명 처리
        if pd.notna(normal_row.get('송신스키마파일생성여부')) and float(normal_row['송신스키마파일생성여부']) == 1.0:
            # 스키마 파일의 base_name과 namespace 추출
            base_name = os.path.splitext(os.path.basename(normal_row['송신스키마파일명']))[0]
            
            # 소스 파일에서 기존 namespace 확인
            has_no_namespace = False
            existing_namespace = extract_existing_namespace(match_row['송신스키마파일명'], base_name)
            if existing_namespace and 'no_namespace_schema' in existing_namespace:
                has_no_namespace = True
                debug_print(f"송신스키마파일명에서 no_namespace_schema 감지됨: {existing_namespace}")
            
            # namespace 생성 (no_namespace 여부에 따라 다르게 처리)
            namespace, _ = process_schema_path(normal_row['송신스키마파일명'], preserve_no_namespace=has_no_namespace)
            
            schema_replacements = []
            
            # no_namespace가 아닌 경우에만 namespace 치환 적용
            if not has_no_namespace:
                schema_replacements.extend([
                    {
                        "설명": "xs:schema xmlns 치환",
                        "찾기": {
                            "정규식": f'xmlns\\s*=\\s*"[^"]*{base_name}[^"]*"'
                        },
                        "교체": {
                            "값": f'xmlns="{namespace}"'
                        }
                    },
                    {
                        "설명": "xs:schema targetNamespace 치환",
                        "찾기": {
                            "정규식": f'targetNamespace\\s*=\\s*"[^"]*{base_name}[^"]*"'
                        },
                        "교체": {
                            "값": f'targetNamespace="{namespace}"'
                        }
                    }
                ])
            
            yaml_structure[f"{i//2 + 1}번째 행"]["송신스키마파일명"] = {
                "원본파일": match_row['송신스키마파일명'],
                "복사파일": modify_path(normal_row['송신스키마파일명']),  # 경로 수정
                "치환목록": schema_replacements
            }
            print("\n[송신스키마파일명 생성]")
            print(f"  원본파일: {match_row['송신스키마파일명']}")
            print(f"  복사파일: {modify_path(normal_row['송신스키마파일명'])}")
        
        # 4. 수신스키마파일명 처리
        if pd.notna(normal_row.get('수신스키마파일생성여부')) and float(normal_row['수신스키마파일생성여부']) == 1.0:
            # 스키마 파일의 base_name과 namespace 추출
            base_name = os.path.splitext(os.path.basename(normal_row['수신스키마파일명']))[0]
            
            # 소스 파일에서 기존 namespace 확인
            has_no_namespace = False
            existing_namespace = extract_existing_namespace(match_row['수신스키마파일명'], base_name)
            if existing_namespace and 'no_namespace_schema' in existing_namespace:
                has_no_namespace = True
                debug_print(f"수신스키마파일명에서 no_namespace_schema 감지됨: {existing_namespace}")
            
            # namespace 생성 (no_namespace 여부에 따라 다르게 처리)
            namespace, _ = process_schema_path(normal_row['송신스키마파일명'], preserve_no_namespace=has_no_namespace)
            
            schema_replacements = []
            
            # no_namespace가 아닌 경우에만 namespace 치환 적용
            if not has_no_namespace:
                schema_replacements.extend([
                    {
                        "설명": "xs:schema xmlns 치환",
                        "찾기": {
                            "정규식": f'xmlns\\s*=\\s*"[^"]*{base_name}[^"]*"'
                        },
                        "교체": {
                            "값": f'xmlns="{namespace}"'
                        }
                    },
                    {
                        "설명": "xs:schema targetNamespace 치환",
                        "찾기": {
                            "정규식": f'targetNamespace\\s*=\\s*"[^"]*{base_name}[^"]*"'
                        },
                        "교체": {
                            "값": f'targetNamespace="{namespace}"'
                        }
                    }
                ])
            
            yaml_structure[f"{i//2 + 1}번째 행"]["수신스키마파일명"] = {
                "원본파일": match_row['수신스키마파일명'],
                "복사파일": modify_path(normal_row['수신스키마파일명']),  # 경로 수정
                "치환목록": schema_replacements
            }
            print("\n[수신스키마파일명 생성]")
            print(f"  원본파일: {match_row['수신스키마파일명']}")
            print(f"  복사파일: {modify_path(normal_row['수신스키마파일명'])}")
        
        # YAML 구조 출력
        print("\n[YAML 구조]")
        print(yaml.dump(yaml_structure, allow_unicode=True, sort_keys=False))
        print("=" * 50)
        
        # 전체 YAML 구조에 현재 구조 추가
        full_yaml_structure.update(yaml_structure)
    
    # YAML 파일 생성
    try:
        with open(yaml_path, 'w', encoding='utf-8') as yf:
            yaml.dump(full_yaml_structure, yf, allow_unicode=True, sort_keys=False)
        print(f"\nYAML 파일이 생성되었습니다: {yaml_path}")
        return len(full_yaml_structure)  # 생성된 작업 수 반환
    except Exception as e:
        print(f"\nYAML 파일 생성 중 오류 발생: {str(e)}")
        return 0

def apply_replacements(text, replacements):
    """여러 치환 규칙을 순차적으로 적용하여 새로운 텍스트 반환."""
    new_text = text
    for repl in replacements:
        # 단순 문자열 치환 (정규식의 경우 필요시 re.sub로 변경 가능)
        from_str = repl.get("from", "")
        to_str = repl.get("to", "")
        if from_str:
            new_text = new_text.replace(from_str, to_str)
    return new_text

def compute_diff(original_text, modified_text, fromfile="[Before]", tofile="[After]"):
    """두 텍스트 버전에 대한 unified diff를 생성하여 리스트로 반환."""
    original_lines = original_text.splitlines(keepends=True)
    modified_lines = modified_text.splitlines(keepends=True)
    diff_lines = difflib.unified_diff(original_lines, modified_lines,
                                    fromfile=fromfile, tofile=tofile, lineterm='')
    return list(diff_lines)

def preview_diff(yaml_path):
    """YAML 파일에 정의된 각 작업에 대해 diff 미리보기를 콘솔에 출력."""
    try:
        with open(yaml_path, 'r', encoding='utf-8') as yf:
            data = yaml.safe_load(yf)
    except FileNotFoundError:
        print(f"YAML 파일을 찾을 수 없습니다: {yaml_path}")
        return

    jobs = data.get("jobs", []) if data else []
    for job in jobs:
        source = job.get("source")
        dest = job.get("destination", "(preview)")
        replacements = job.get("replacements", [])
        if not source or not replacements:
            continue
        try:
            with open(source, 'r', encoding='utf-8') as sf:
                original_text = sf.read()
        except FileNotFoundError:
            print(f"\n[오류] 원본 파일을 찾을 수 없습니다: {source}")
            continue

        # 치환 적용 (미리보기이므로 파일 저장 안 함)
        modified_text = apply_replacements(original_text, replacements)

        # diff 계산
        diff_lines = compute_diff(original_text, modified_text, fromfile=source, tofile=f"{dest} (preview)")

        # diff 출력
        print(f"\n*** {source} vs {dest} 미리보기 diff ***")
        for line in diff_lines:
            print(line, end='')

def detect_encoding(file_path):
    """파일의 인코딩을 감지합니다."""
    if HAS_CHARDET:
        with open(file_path, 'rb') as f:
            raw = f.read()
            result = chardet.detect(raw)
            return result['encoding']
    else:
        return 'utf-8'  # 기본값으로 utf-8 사용

def copy_file_with_check(source, dest):
    """파일을 복사하되, 대상 파일이 이미 존재하면 경고를 출력합니다."""
    try:
        # 대상 디렉토리가 없으면 생성
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        
        # 대상 파일이 이미 존재하는지 확인
        if os.path.exists(dest):
            print(f"경고: 파일이 이미 존재합니다 - {dest}")
            return False
            
        # 바이너리 모드로 파일 복사
        with open(source, 'rb') as src, open(dest, 'wb') as dst:
            dst.write(src.read())
        print(f"파일 복사 완료: {source} -> {dest}")
        return True
    except Exception as e:
        print(f"파일 복사 중 오류 발생: {str(e)}")
        return False

def apply_schema_replacements(file_path, replacements):
    """파일에 치환 목록을 적용합니다."""
    try:
        debug_print(f"\n=== 파일 치환 시작: {file_path} ===")
        
        # 파일 인코딩 감지 및 읽기
        encoding = detect_encoding(file_path)
        if not encoding:
            debug_print(f"경고: 파일의 인코딩을 감지할 수 없습니다 - {file_path}")
            encoding = 'utf-8'
        debug_print(f"감지된 인코딩: {encoding}")
        
        # 파일 읽기
        with open(file_path, 'rb') as f:
            content_bytes = f.read()
            debug_print(f"파일 크기: {len(content_bytes)} bytes")
            
        # 디코딩
        try:
            content = content_bytes.decode(encoding)
            debug_print("디코딩 성공")
        except UnicodeDecodeError:
            debug_print("utf-8로 재시도")
            try:
                content = content_bytes.decode('utf-8')
                encoding = 'utf-8'
            except UnicodeDecodeError:
                debug_print("latin-1로 시도")
                content = content_bytes.decode('latin-1')
                encoding = 'latin-1'
        
        modified = False
        # 각 치환 규칙 적용
        for idx, repl in enumerate(replacements, 1):
            debug_print(f"\n--- 치환 규칙 {idx}/{len(replacements)} 적용 시도 ---")
            debug_print(f"설명: {repl.get('설명', '설명 없음')}")
            
            pattern = repl['찾기']['정규식']
            replacement = repl['교체']['값']
            
            debug_print(f"정규식 패턴: {pattern}")
            debug_print(f"교체할 값: {replacement}")
            
            # 패턴이 파일에 존재하는지 확인
            matches = list(re.finditer(pattern, content))
            if not matches:
                debug_print("패턴이 파일에서 발견되지 않음")
                continue
            debug_print(f"패턴 매칭 수: {len(matches)}")
            
            # 첫 번째 매칭 내용 출력
            if matches:
                first_match = matches[0]
                debug_print("첫 번째 매칭:")
                debug_print(f"  위치: {first_match.start()}-{first_match.end()}")
                debug_print(f"  내용: {first_match.group()}")
            
            try:
                # 정규식 치환 수행
                new_content = re.sub(pattern, replacement, content)
                if new_content == content:
                    debug_print("치환 후 변경사항 없음")
                else:
                    content = new_content
                    modified = True
                    debug_print("치환 성공")
            except Exception as e:
                debug_print(f"치환 중 오류 발생: {str(e)}")
                continue
        
        # 변경된 경우에만 파일 저장
        if modified:
            debug_print("\n파일 저장 시작")
            content_bytes = content.encode(encoding)
            with open(file_path, 'wb') as f:
                f.write(content_bytes)
            debug_print(f"파일 저장 완료 (크기: {len(content_bytes)} bytes)")
            return True
        else:
            debug_print("\n변경사항이 없어 파일을 저장하지 않음")
            return False
            
    except Exception as e:
        debug_print(f"치환 작업 중 예외 발생: {str(e)}")
        return False

def check_file_locks(file_path):
    """파일이 다른 프로세스에 의해 잠겨있는지 확인합니다."""
    try:
        # 파일을 독점 모드로 열어서 잠금 상태 확인
        with open(file_path, 'r+b'):
            return False  # 파일이 잠겨있지 않음
    except (IOError, OSError):
        return True  # 파일이 잠겨있음 또는 접근 불가

def generate_delete_batch(data, batch_path):
    """생성된 파일들을 삭제하는 배치 파일을 생성합니다."""
    with open(batch_path, 'w', encoding='utf-8') as f:
        f.write('@echo off\n')
        f.write('chcp 65001 > nul\n')  # UTF-8 코드 페이지로 설정
        f.write('cls\n')  # 화면 지우기
        f.write('setlocal EnableDelayedExpansion\n\n')  # 지연 환경 변수 확장 활성화
        
        # 삭제 대상 파일 목록을 주석으로 먼저 작성
        f.write('rem ===================================================================\n')
        f.write('rem 삭제 대상 파일 목록\n')
        f.write('rem ===================================================================\n')
        
        # 삭제 대상 파일을 저장할 리스트
        delete_files = []
        locked_files = []
        
        # 각 행 처리하여 삭제 대상 파일 수집
        for row_key, row_data in data.items():
            # 파일 타입별 처리
            file_types = ['송신파일경로', '수신파일경로', '송신스키마파일명', '수신스키마파일명']
            
            for file_type in file_types:
                if file_type in row_data:
                    file_info = row_data[file_type]
                    copy_file = file_info.get('복사파일', '')
                    
                    # 파일이 실제로 존재하는 경우에만 목록에 추가
                    if copy_file and os.path.exists(copy_file):
                        delete_files.append((file_type, copy_file))
                        f.write(f'rem [{file_type}] {copy_file}\n')
                        
                        # 파일 잠금 상태 확인
                        if check_file_locks(copy_file):
                            locked_files.append((file_type, copy_file))
                            f.write(f'rem     ^^ 경고: 이 파일은 다른 프로세스가 사용 중입니다\n')
        
        f.write(f'rem 총 삭제 대상 파일 수: {len(delete_files)}개\n')
        if locked_files:
            f.write(f'rem 잠긴 파일 수: {len(locked_files)}개 (삭제 시 주의 필요)\n')
        f.write('rem ===================================================================\n\n')
        
        # 실제 삭제 명령어 작성
        f.write('echo 파일 삭제를 시작합니다...\n')
        f.write(f'echo 총 {len(delete_files)}개의 파일을 삭제합니다.\n')
        if locked_files:
            f.write(f'echo 경고: {len(locked_files)}개의 파일이 다른 프로세스에 의해 사용 중입니다.\n')
            f.write('echo       Everything, 안티바이러스, 텍스트 에디터 등을 종료한 후 실행하세요.\n')
        f.write('echo.\n\n')
        
        # 성공/실패 카운터 초기화
        f.write('set success_count=0\n')
        f.write('set fail_count=0\n\n')
        
        # 각 파일 삭제 명령어 작성
        for file_type, copy_file in delete_files:
            f.write(f'echo [{file_type}] 삭제 시도: {copy_file}\n')
            
            # 파일 존재 여부 확인
            f.write(f'if exist "{copy_file}" (\n')
            
            # 파일 속성 제거 (읽기 전용, 숨김, 시스템 속성 제거)
            f.write(f'    attrib -r -h -s "{copy_file}" 2>nul\n')
            
            # 파일이 사용 중인지 확인하고 강제 삭제 시도
            f.write(f'    del /f /q "{copy_file}" 2>nul\n')
            
            # 삭제 성공 여부 확인
            f.write(f'    if exist "{copy_file}" (\n')
            f.write(f'        echo    [실패] 파일을 삭제할 수 없습니다. 다른 프로세스가 사용 중일 수 있습니다.\n')
            f.write(f'        set /a fail_count+=1\n')
            
            # 파일이 잠겨있는 경우 추가 시도
            f.write(f'        echo    [재시도] PowerShell로 삭제를 시도합니다...\n')
            f.write(f'        powershell -Command "try {{ Remove-Item -Path \'{copy_file}\' -Force -ErrorAction Stop; Write-Host \'    [성공] PowerShell로 삭제 완료\' }} catch {{ Write-Host \'    [실패] PowerShell 삭제도 실패: $_\' }}"\n')
            
            f.write(f'    ) else (\n')
            f.write(f'        echo    [성공] 파일 삭제 완료\n')
            f.write(f'        set /a success_count+=1\n')
            f.write(f'    )\n')
            f.write(f') else (\n')
            f.write(f'    echo    [정보] 파일이 이미 존재하지 않습니다.\n')
            f.write(f'    set /a success_count+=1\n')
            f.write(f')\n')
            f.write('echo.\n')
        
        if not delete_files:
            f.write('echo 삭제할 파일이 없습니다.\n')
        else:
            # 결과 요약 출력
            f.write('\necho ===================================================================\n')
            f.write('echo 작업 완료 요약:\n')
            f.write('echo    성공: !success_count!개\n')
            f.write('echo    실패: !fail_count!개\n')
            f.write('echo ===================================================================\n')
            
            # Windows 탐색기 새로고침을 위한 명령
            f.write('\necho.\n')
            f.write('echo Windows 탐색기를 새로고침합니다...\n')
            f.write('powershell -Command "$shell = New-Object -ComObject Shell.Application; $shell.Windows() | ForEach-Object { $_.Refresh() }"\n')
            
            # Everything 캐시 지우기 제안
            f.write('\necho.\n')
            f.write('echo 참고: Everything 검색 도구를 사용 중이라면 F5를 눌러 인덱스를 새로고침하세요.\n')
        
        f.write('\necho.\n')
        f.write('pause\n')
    
    print(f"\n삭제 배치 파일이 생성되었습니다: {batch_path}")
    print(f"삭제 대상 파일 수: {len(delete_files)}개")
    if locked_files:
        print(f"경고: {len(locked_files)}개의 파일이 다른 프로세스에 의해 사용 중입니다.")
        print("     Everything, 안티바이러스 소프트웨어, 텍스트 에디터 등을 종료 후 배치 파일을 실행하세요.")

def execute_replacements(yaml_path, log_path, summary_path):
    """YAML에 정의된 복사 및 치환 작업을 실행하고 로그를 생성합니다."""
    try:
        debug_print(f"YAML 파일 읽기 시작: {yaml_path}")
        with open(yaml_path, 'r', encoding='utf-8') as yf:
            data = yaml.safe_load(yf)
            debug_print(f"YAML 데이터 로드 완료: {len(data) if data else 0}개 항목")
    except FileNotFoundError:
        print(f"YAML 파일을 찾을 수 없습니다: {yaml_path}")
        return
    except Exception as e:
        print(f"YAML 파일 읽기 중 오류 발생: {str(e)}")
        return

    if not data:
        print("실행할 작업이 없습니다.")
        return

    # 로그 파일 초기화
    debug_print(f"로그 파일 초기화: {log_path}")
    with open(log_path, 'w', encoding='utf-8') as lf:
        lf.write(f"[{datetime.datetime.now()}] 작업 시작\n")

    summary_data = []
    total_copies = 0
    total_replacements = 0

    # 각 행 처리
    for row_key, row_data in data.items():
        debug_print(f"\n=== {row_key} 처리 시작 ===")
        debug_print(f"행 데이터: {row_data.keys()}")
        
        # 각 파일 타입 처리 (송신파일경로, 수신파일경로 등)
        for file_type, file_info in row_data.items():
            debug_print(f"\n--- {file_type} 처리 ---")
            source = file_info.get('원본파일')
            dest = file_info.get('복사파일')
            replacements = file_info.get('치환목록', [])
            
            debug_print(f"원본파일: {source}")
            debug_print(f"복사파일: {dest}")
            debug_print(f"치환규칙 수: {len(replacements)}")
            
            if not source or not dest:
                debug_print("원본 또는 대상 파일 경로가 없음, 건너뜀")
                continue
            
            if DEBUG_MODE and replacements:
                debug_print("\n치환 규칙 목록:")
                for idx, repl in enumerate(replacements, 1):
                    debug_print(f"  {idx}. {repl.get('설명', '설명 없음')}")
                    debug_print(f"     조건: {repl.get('조건', {})}")
                    debug_print(f"     찾기: {repl.get('찾기', {})}")
                    debug_print(f"     교체: {repl.get('교체', {})}")
                
            # 1. 파일 복사
            debug_print(f"\n파일 복사 시도: {source} -> {dest}")
            if copy_file_with_check(source, dest):
                total_copies += 1
                debug_print("파일 복사 성공")
                
                # 2. 치환 목록이 있는 경우 치환 수행
                if replacements:
                    debug_print(f"치환 작업 시작: {dest}")
                    if apply_schema_replacements(dest, replacements):
                        total_replacements += 1
                        debug_print("치환 작업 성공")
                    else:
                        debug_print("치환 작업 실패 또는 변경사항 없음")
                else:
                    debug_print("치환 규칙 없음, 건너뜀")
                        
            # 작업 결과 기록
            summary = f"{file_type}: {source} -> {dest}"
            if replacements:
                summary += f" (치환: {len(replacements)}개 규칙)"
            summary_data.append(summary)
            debug_print(f"작업 결과: {summary}")
            
            # 로그 기록
            with open(log_path, 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.datetime.now()}] {summary}\n")

    # 요약 파일 생성
    debug_print(f"\n요약 파일 생성: {summary_path}")
    with open(summary_path, 'w', encoding='utf-8') as sf:
        sf.write("작업 요약:\n")
        for line in summary_data:
            sf.write(line + "\n")
        sf.write(f"\n총 복사 파일 수: {total_copies}")
        sf.write(f"\n총 치환 파일 수: {total_replacements}")

    # 엑셀 로그 파일 생성
    excel_path = os.path.splitext(log_path)[0] + '.xlsx'
    generate_excel_log(data, excel_path)

    # 삭제 배치 파일 생성
    batch_path = os.path.splitext(log_path)[0] + '_delete.bat'
    generate_delete_batch(data, batch_path)

    debug_print("\n=== 전체 작업 완료 ===")
    debug_print(f"총 복사 파일 수: {total_copies}")
    debug_print(f"총 치환 파일 수: {total_replacements}")

    print(f"\n작업이 완료되었습니다.")
    print(f"총 복사 파일 수: {total_copies}")
    print(f"총 치환 파일 수: {total_replacements}")

def generate_excel_log(data, excel_path):
    """YAML 실행 결과를 엑셀 파일로 생성합니다."""
    # 엑셀 데이터를 저장할 리스트
    excel_rows = []
    
    # 각 행 처리
    for row_key, row_data in data.items():
        # 파일 타입별 처리
        file_types = ['송신파일경로', '수신파일경로', '송신스키마파일명', '수신스키마파일명']
        
        for file_type in file_types:
            if file_type in row_data:
                file_info = row_data[file_type]
                excel_rows.append({
                    '파일타입': file_type,
                    '원본파일경로': file_info.get('원본파일', ''),
                    '복사파일경로': file_info.get('복사파일', ''),
                    '생성여부': 'O' if os.path.exists(file_info.get('복사파일', '')) else 'X'
                })
    
    # DataFrame 생성
    df = pd.DataFrame(excel_rows)
    
    # 컬럼 순서 지정
    df = df[['파일타입', '원본파일경로', '복사파일경로', '생성여부']]
    
    # 엑셀 파일 저장
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    df.to_excel(writer, index=False)
    
    # 워크시트 가져오기
    worksheet = writer.sheets['Sheet1']
    
    # 폰트 크기를 10으로 설정
    from openpyxl.styles import Font, PatternFill, Alignment
    font_10 = Font(size=10)
    
    # 헤더(첫 행) 스타일 설정
    light_blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    header_font = Font(size=10, bold=True)
    
    # 모든 셀에 대해 폰트 크기 10 적용
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = font_10
    
    # 헤더 행에 스타일 적용
    for cell in worksheet[1]:
        cell.fill = light_blue_fill
        cell.font = header_font
    
    # 생성여부 컬럼 가운데 정렬
    for row in worksheet.iter_rows(min_row=2):  # 헤더 제외
        row[-1].alignment = Alignment(horizontal='center')  # 마지막 컬럼(생성여부)
    
    # 컬럼 너비 자동 조절
    for column in worksheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        
        # 각 셀의 길이를 확인하여 최대 길이 계산
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # 컬럼 너비 설정 (최대 길이 + 여유 공간)
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    writer.close()
    print(f"\n엑셀 로그 파일이 생성되었습니다: {excel_path}")

def main():
    while True:
        print("\n=== 문자열 치환 도구 ===")
        print("1. YAML 생성 (엑셀 -> YAML)")
        print("2. 미리보기 (YAML 기반 diff 출력)")
        print("3. 실행 (파일 복사 및 치환)")
        print("0. 종료")
        
        choice = input("\n원하는 작업을 선택하세요: ").strip()
        
        if choice == "1":
            excel_path = input("엑셀 파일 경로를 입력하세요: ").strip()
            yaml_path = input("생성할 YAML 파일 경로를 입력하세요: ").strip()
            try:
                count = generate_yaml_from_excel(excel_path, yaml_path)
                print(f"YAML 파일이 생성되었습니다. (총 {count}개 작업)")
            except Exception as e:
                print(f"오류 발생: {str(e)}")
        
        elif choice == "2":
            yaml_path = input("YAML 파일 경로를 입력하세요: ").strip()
            preview_diff(yaml_path)
        
        elif choice == "3":
            yaml_path = input("YAML 파일 경로를 입력하세요: ").strip()
            log_path = input("로그 파일 경로를 입력하세요: ").strip()
            summary_path = input("요약 파일 경로를 입력하세요: ").strip()
            execute_replacements(yaml_path, log_path, summary_path)
        
        elif choice == "0":
            print("프로그램을 종료합니다.")
            break
        
        else:
            print("잘못된 선택입니다. 다시 시도하세요.")

if __name__ == "__main__":
    main() 