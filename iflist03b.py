"""
파일명: iflist03b.py
버전: v8.4 (리팩토링 버전)
작성일: 2025년 1월

설명:
이 프로그램은 SQLite 데이터베이스에서 인터페이스 목록을 추출하고 가공하여 Excel 파일로 출력합니다.
설정 가능한 변환 규칙을 통해 다양한 시스템 매핑을 지원합니다.
또한 기본행과 매칭행을 비교하여 다양한 검증 규칙에 따라 오류를 확인하고 결과를 '비교로그' 컬럼에 기록합니다.

이 파일은 iflist03a.py를 리팩토링하여 재사용 가능하게 만든 버전입니다.

사용법:
1. 'iflist.sqlite' 데이터베이스 파일이 현재 디렉토리에 있어야 합니다.
2. 'iflist' 테이블에 '송신시스템', '수신시스템', 'I/F명' 컬럼이 있어야 합니다.
3. 상단의 설정 섹션에서 변환 규칙을 수정할 수 있습니다.
4. 명령행에서 다음과 같이 실행: 'python iflist03b.py'
5. 결과는 '{스크립트명}_reordered_v8.4.xlsx' 파일로 저장됩니다.

필요 라이브러리:
- sqlite3: SQLite 데이터베이스 액세스
- pandas: 데이터 처리 및 조작
- xlsxwriter: Excel 출력 및 서식 지정 (pip install xlsxwriter)
- os.path: 파일 존재 여부 확인

수정 이력:
- v8.3: iflist03a.py 원본
- v8.4: 재사용 가능하도록 리팩토링, 설정 섹션 추가
"""

import sqlite3
import pandas as pd
import sys
import os
import os.path
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ========== 설정 섹션 시작 ==========
# 이 섹션의 값들을 수정하여 다른 시스템 매핑에도 사용할 수 있습니다.

# 시스템 변환 규칙
# 키: 필터링할 패턴, 값: 변환할 대상
SYSTEM_CONVERSION_RULES = {
    'RTS_GM2': 'RTS_GM'  # RTS_GM2를 RTS_GM으로 변환
}

# 업무명 변환 규칙
BUSINESS_NAME_MAPPING = {
    'RTS_GM2': 'RTS_GM'
}

# 법인 코드 매핑
CORP_CODE_MAPPING = {
    'KR': 'KR',
    'NJ': 'CN',
    'VH': 'VN'
}

# EMS 매핑
EMS_MAPPING = {
    'MES01': 'EMS_64000',
    'default': 'EMS_63000'
}

# 환경별 디렉토리 접미사
DIR_SUFFIX = {
    'base': '_TEST_SOURCE',    # 기본행
    'match': '_PROD_SOURCE'    # 매칭행
}

# 파일 경로 기본 설정
BASE_PATH = "C:\\BwProject"
SCHEMA_PATH = "SharedResources\\Schema\\source"

# 디버그 모드 설정
# 0 또는 2: 최종 필터링된 행(연두색)만 표시
# 1: 모든 매칭 행(노란색)과 필터링된 행(연두색) 모두 표시
DEBUG_MODE = 2

# 데이터베이스 설정
DB_FILENAME = 'iflist.sqlite'
TABLE_NAME = 'iflist'
OUTPUT_VERSION = 'v8.4'

# 컬럼명 설정
COLUMN_MAPP_SEQ = '매핑SEQ'
COLUMN_SEND_SYSTEM = '송신시스템'
COLUMN_RECV_SYSTEM = '수신시스템'
COLUMN_IF_NAME = 'I/F명'
COLUMN_SEND_CORP = '송신\n법인'
COLUMN_RECV_CORP = '수신\n법인'
COLUMN_SEND_PKG = '송신패키지'
COLUMN_RECV_PKG = '수신패키지'
COLUMN_SEND_TASK = '송신\n업무명'
COLUMN_RECV_TASK = '수신\n업무명'
COLUMN_EMS_NAME = 'EMS명'
COLUMN_GROUP_ID = 'Group ID'
COLUMN_EVENT_ID = 'Event_ID'

# 업무명 키워드 (디렉토리 생성에 사용)
TASK_KEYWORDS = ["PNL", "EAS", "MOD", "MES"]

# 색상 설정
COLORS = {
    'yellow': 'FFFF00',      # 노란색
    'green': '90EE90',       # 연두색
    'orange': 'FFA500',      # 주황색
    'error': 'FFC000',       # 오류 표시용 주황색
    'white': 'FFFFFF',       # 흰색
    'gray': 'F2F2F2',        # 회색
    'blue_very_light': 'E6F2FF',  # 매우 밝은 파란색
    'blue_light': '99CCFF',        # 밝은 파란색
    'blue_medium': '3399FF',       # 중간 파란색
    'blue_dark': '0066CC'          # 진한 파란색
}

# 오류 표시를 위한 주황색 배경 정의
ORANGE_FILL = PatternFill(start_color=COLORS['error'], end_color=COLORS['error'], fill_type='solid')

# ========== 설정 섹션 끝 ==========

# 1. 출력 Excel 파일 이름 설정
try:
    script_basename = os.path.basename(sys.argv[0])
    script_name_without_ext = os.path.splitext(script_basename)[0]
    excel_filename = f"{script_name_without_ext}_reordered_{OUTPUT_VERSION}.xlsx"
except Exception:
    excel_filename = f"output_reordered_{OUTPUT_VERSION}.xlsx"
    print(f"스크립트 이름을 감지할 수 없어 기본 파일명 '{excel_filename}'을 사용합니다.")

df_complete_table = pd.DataFrame()  # 원본 전체 테이블
df_filtered = pd.DataFrame()        # 초기 필터링된 테이블

# --- 유틸리티 함수 ---
def apply_replacement_rules(text, rules=None):
    """설정된 변환 규칙에 따라 문자열 변환"""
    if rules is None:
        rules = SYSTEM_CONVERSION_RULES
    
    if not isinstance(text, str):
        return text
        
    result = text
    for pattern, replacement in rules.items():
        result = result.replace(pattern, replacement)
    return result

# --- 검증 함수 ---
def check_systems(base_value, match_value, column_name):
    """송신시스템/수신시스템 비교 로직"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (비어있는 값)"
    
    expected_value = apply_replacement_rules(base_value)
    if expected_value != match_value:
        return f"{column_name} 비교오류"
    return ""

def check_business_name(base_value, match_value, column_name):
    """업무명 비교 로직 - 설정 기반"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (비어있는 값)"
    
    # 설정된 매핑 규칙 확인
    if base_value in BUSINESS_NAME_MAPPING:
        expected = BUSINESS_NAME_MAPPING[base_value]
        if match_value != expected:
            return f"{column_name} 비교오류"
    
    return ""

def check_package(base_value, match_value, column_name):
    """패키지 비교 로직 - 설정 기반"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (비어있는 값)"
    
    # 모든 변환 규칙에 대해 확인
    for pattern, replacement in SYSTEM_CONVERSION_RULES.items():
        if pattern in base_value and replacement not in match_value:
            return f"{column_name} 비교오류"
    
    return ""

def check_same_content(base_value, match_value, column_name):
    """내용이 같아야 하는 컬럼 비교"""
    if not isinstance(base_value, str) and not isinstance(match_value, str):
        # 둘 다 문자열이 아닌 경우 (예: NaN)
        if pd.isna(base_value) and pd.isna(match_value):
            return ""
        return f"{column_name} 비교오류 (유형 불일치)"
    
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (유형 불일치)"
    
    if base_value.strip() != match_value.strip():
        return f"{column_name} 비교오류"
    return ""

def check_table_or_routing(base_value, match_value, column_name):
    """테이블명 또는 라우팅 비교 로직"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        if pd.isna(base_value) and pd.isna(match_value):
            return ""
        return f"{column_name} 비교오류 (비어있는 값)"
    
    # 변환 규칙의 패턴이 포함되어 있는지 확인
    contains_pattern = any(pattern in base_value for pattern in SYSTEM_CONVERSION_RULES.keys())
    
    if contains_pattern:
        expected_value = apply_replacement_rules(base_value)
        if expected_value != match_value:
            return f"{column_name} 비교오류"
    elif base_value.strip() != match_value.strip():
        return f"{column_name} 비교오류"
    return ""

def check_table_with_split(base_value, match_value, column_name):
    """Source Table, Destination Table용 비교 로직 (단어 분할 후 패턴 확인)"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        if pd.isna(base_value) and pd.isna(match_value):
            return ""
        return f"{column_name} 비교오류 (비어있는 값)"
    
    # 기본값으로 단순 비교 
    should_check_patterns = False
    
    # '.'과 '_'로 분할하여 단어 확인
    words = re.split('[._]', base_value)
    for word in words:
        # 변환 규칙의 패턴으로 시작하는지 확인
        if any(word.startswith(pattern) for pattern in SYSTEM_CONVERSION_RULES.keys()):
            should_check_patterns = True
            break
    
    if should_check_patterns:
        expected_value = apply_replacement_rules(base_value)
        if expected_value != match_value:
            return f"{column_name} 비교오류"
    elif base_value.strip() != match_value.strip():
        return f"{column_name} 비교오류"
    
    return ""

# --- 파일 경로 생성 함수 ---
def create_file_path(row, is_send=True, color_flag=None):
    """
    주어진 행 데이터로부터 파일 경로를 생성합니다.
    
    Args:
        row: 데이터프레임의 행
        is_send: 송신 파일 경로인지 여부 (False면 수신 파일 경로)
        color_flag: 행의 색상 정보 (None: 기본행, 'yellow'/'green': 매칭행)
        
    Returns:
        생성된 파일 경로 문자열
    """
    if DEBUG_MODE == 1:
        print(f"color_flag: {color_flag}")  # color_flag 값 출력
    try:
        # 사용할 컬럼 선택 (송신/수신에 따라)
        corp_col = COLUMN_SEND_CORP if is_send else COLUMN_RECV_CORP
        pkg_col = COLUMN_SEND_PKG if is_send else COLUMN_RECV_PKG
        task_col = COLUMN_SEND_TASK if is_send else COLUMN_RECV_TASK
        
        # 안전하게 컬럼값 가져오기 (컬럼이 없는 경우 빈 문자열 반환)
        def safe_get_value(df_row, column_name):
            try:
                val = df_row[column_name] if column_name in df_row.index else ""
                return str(val).strip() if pd.notna(val) else ""
            except:
                return ""
        
        # 필요한 값들 가져오기
        corp_val = safe_get_value(row, corp_col)
        pkg_val = safe_get_value(row, pkg_col)
        task_val = safe_get_value(row, task_col)
        ems_val = safe_get_value(row, COLUMN_EMS_NAME)
        group_id = safe_get_value(row, COLUMN_GROUP_ID)
        event_id = safe_get_value(row, COLUMN_EVENT_ID)
        recv_task = "" if is_send else safe_get_value(row, COLUMN_RECV_TASK)
        
        # 1번 디렉토리 (법인 정보에 따라)
        dir1 = CORP_CODE_MAPPING.get(corp_val, "UNK")
        
        # 접미사 결정
        if color_flag is None:  # 기본행
            dir1 += DIR_SUFFIX['base']
        else:  # 매칭행
            dir1 += DIR_SUFFIX['match']
        
        # 2번 디렉토리 (패키지의 첫 '_' 이전 부분)
        dir2 = pkg_val.split('_')[0] if '_' in pkg_val and pkg_val else pkg_val
        
        # 3번 디렉토리 (조건부)
        dir3 = ""
        if task_val and any(keyword in task_val for keyword in TASK_KEYWORDS):
            parts = task_val.split('_')
            if len(parts) > 1:
                dir3 = parts[-1]
        
        # 4번 디렉토리 (EMS명에 따라)
        dir4 = EMS_MAPPING.get(ems_val, EMS_MAPPING['default'])
        
        # 5번 디렉토리 (패키지 전체 이름)
        dir5 = pkg_val
        
        # 파일명
        if is_send:
            filename = f"{group_id}.{event_id}.process" if group_id and event_id else "unknown.process"
        else:
            filename = f"{group_id}.{event_id}.{recv_task}.process" if group_id and event_id else "unknown.process"
        
        # 경로 구성 (dir3가 없을 수 있음)
        path_parts = [BASE_PATH, dir1, dir2, "Processes"]
        if dir3:
            path_parts.append(dir3)
        path_parts.extend([dir4, dir5, filename])
        
        # 경로 조합
        return "\\".join(path_parts)
    except Exception as e:
        print(f"파일 경로 생성 중 오류 발생: {e}")
        return ""

# --- 파일 존재 여부 확인 함수 ---
def check_file_exists(file_path):
    """
    주어진 파일 경로가 실제로 존재하는지 확인
    
    Args:
        file_path: 확인할 파일 경로
        
    Returns:
        파일 존재 여부 (1: 존재, 0: 존재하지 않음)
    """
    try:
        if os.path.isfile(file_path):
            return 1
        else:
            return 0
    except Exception as e:
        print(f"파일 존재 여부 확인 중 오류: {e}")
        return 0

# --- 디렉토리 내 파일 개수 확인 함수 ---
def count_files_in_directory(file_path):
    """
    주어진 파일 경로의 디렉토리 내 파일 개수를 반환
    
    Args:
        file_path: 파일 경로 (디렉토리 추출용)
        
    Returns:
        디렉토리 내 파일 개수, 경로가 잘못되었거나 오류 시 0 반환
    """
    try:
        # 파일 경로에서 디렉토리 추출
        directory = os.path.dirname(file_path)
        if not directory or not os.path.isdir(directory):
            return 0
        
        # 디렉토리 내 파일만 카운트 (하위 디렉토리 제외)
        file_count = len([f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))])
        return file_count
    
    except Exception as e:
        print(f"디렉토리 파일 개수 확인 중 오류: {e}")
        return 0

# --- 스키마 파일 경로 생성 함수 ---
def create_schema_file_path(row, is_send=True, color_flag=None):
    """
    주어진 행 데이터로부터 스키마 파일 경로를 생성합니다.
    
    Args:
        row: 데이터프레임의 행
        is_send: 송신 스키마 파일 경로인지 여부 (False면 수신 스키마 파일 경로)
        color_flag: 행의 색상 정보 (None: 기본행, 'yellow'/'green': 매칭행)
        
    Returns:
        생성된 스키마 파일 경로 문자열
    """
    try:
        # 사용할 컬럼 선택 (송신/수신에 따라)
        corp_col = COLUMN_SEND_CORP if is_send else COLUMN_RECV_CORP
        pkg_col = COLUMN_SEND_PKG if is_send else COLUMN_RECV_PKG
        db_name_col = '송신\nDB Name'
        schema_col = '송신 \nSchema'
        
        # 안전하게 컬럼값 가져오기 (컬럼이 없는 경우 빈 문자열 반환)
        def safe_get_value(df_row, column_name):
            try:
                val = df_row[column_name] if column_name in df_row.index else ""
                return str(val).strip() if pd.notna(val) else ""
            except:
                return ""
        
        # 필요한 값들 가져오기
        corp_val = safe_get_value(row, corp_col)
        pkg_val = safe_get_value(row, pkg_col)
        db_name = safe_get_value(row, db_name_col)
        schema = safe_get_value(row, schema_col)
        source_table = safe_get_value(row, 'Source Table')
        
        # 1번 디렉토리 (법인 정보에 따라)
        dir1 = CORP_CODE_MAPPING.get(corp_val, "UNK")
        
        # 접미사 결정
        if color_flag is None:  # 기본행
            dir1 += DIR_SUFFIX['base']
        else:  # 매칭행
            dir1 += DIR_SUFFIX['match']
        
        # 2번 디렉토리 (패키지의 첫 '_' 이전 부분)
        dir2 = pkg_val.split('_')[0] if '_' in pkg_val and pkg_val else pkg_val
        
        # 3번 디렉토리 (고정값)
        dir3 = SCHEMA_PATH
        
        # 4번 디렉토리 (DB Name)
        dir4 = db_name
        
        # 5번 디렉토리 (Schema)
        dir5 = schema
        
        # 파일명 생성
        # Source Table을 '.'으로 분할해 두 번째 부분 추출 후 .xsd 확장자 추가
        file_name = ""
        if '.' in source_table:
            file_name = source_table.split('.')[1] + '.xsd'
        else:
            file_name = source_table + '.xsd'  # '.'이 없는 경우 전체 이름 사용
        
        # 경로 구성
        path_parts = [BASE_PATH, dir1, dir2, dir3, dir4, dir5, file_name]
        
        # 빈 값 제거
        path_parts = [part for part in path_parts if part]
        
        # 경로 조합
        return "\\".join(path_parts)
    
    except Exception as e:
        print(f"스키마 파일 경로 생성 오류 ({('송신' if is_send else '수신')}): {e}")
        return "경로 생성 오류"


# --- DB에서 전체 데이터 로드 및 df_filtered 생성 ---
try:
    conn = sqlite3.connect(DB_FILENAME)
    cursor = conn.cursor()

    # 1. DB에서 전체 데이터 로드
    cursor.execute(f'SELECT * FROM "{TABLE_NAME}"')
    all_rows_from_db = cursor.fetchall()

    if not all_rows_from_db:
        print(f"원본 테이블 '{TABLE_NAME}'에 데이터가 없습니다. 처리를 중단합니다.")
    else:
        column_names_from_db = [description[0] for description in cursor.description]
        df_complete_table = pd.DataFrame(all_rows_from_db, columns=column_names_from_db)
        print(f"원본 전체 테이블에 총 {len(df_complete_table)}개의 행이 로드되었습니다.")

        # 2. df_filtered 생성: 매핑SEQ='BBB' AND (송신시스템 또는 수신시스템에 변환 규칙 패턴 포함 조건)
        # 필터 패턴 목록
        filter_patterns = list(SYSTEM_CONVERSION_RULES.keys())
        
        # 매핑SEQ 조건
        cond_mapp_seq = df_complete_table[COLUMN_MAPP_SEQ] == 'BBB'
        
        # 필터 조건 동적 생성
        cond_b_contains = pd.Series([False] * len(df_complete_table))
        cond_c_contains = pd.Series([False] * len(df_complete_table))
        
        for pattern in filter_patterns:
            cond_b_contains |= df_complete_table[COLUMN_SEND_SYSTEM].astype(str).str.contains(pattern, na=False)
            cond_c_contains |= df_complete_table[COLUMN_RECV_SYSTEM].astype(str).str.contains(pattern, na=False)
        
        # AND 조건으로 결합: 매핑SEQ='BBB' AND (송신/수신시스템에 패턴 포함)
        df_filtered = df_complete_table[cond_mapp_seq & (cond_b_contains | cond_c_contains)].copy()
        
        if df_filtered.empty:
            patterns_str = '/'.join(filter_patterns)
            print(f"초기 필터링 조건(매핑SEQ='BBB' AND '{patterns_str}' 포함)에 맞는 데이터가 없습니다. 후속 처리를 진행할 수 없습니다.")
        else:
            print(f"초기 필터링 후 df_filtered에 {len(df_filtered)}개의 행이 남았습니다. (매핑SEQ='BBB' AND 시스템에 '{'/'.join(filter_patterns)}' 포함)")

except sqlite3.Error as e:
    print(f"SQLite 오류 발생: {e}")
except FileNotFoundError:
    print(f"오류: 데이터베이스 파일 '{DB_FILENAME}'을 찾을 수 없습니다.")
except Exception as e:
    print(f"데이터 조회 또는 초기 필터링 중 예상치 못한 오류 발생: {e}")
finally:
    if 'conn' in locals() and conn:
        conn.close()

# --- 데이터 준비 완료 ---

output_rows_info = []

if not df_filtered.empty and not df_complete_table.empty:
    # 필수 컬럼 존재 여부 확인
    required_cols = [COLUMN_IF_NAME, COLUMN_SEND_SYSTEM, COLUMN_RECV_SYSTEM]
    if not all(col in df_filtered.columns for col in required_cols) or \
       not all(col in df_complete_table.columns for col in required_cols):
        missing_cols_filtered = [col for col in required_cols if col not in df_filtered.columns]
        missing_cols_complete = [col for col in required_cols if col not in df_complete_table.columns]
        if missing_cols_filtered: print(f"오류: 필수 컬럼 {missing_cols_filtered}이(가) df_filtered에 없습니다.")
        if missing_cols_complete: print(f"오류: 필수 컬럼 {missing_cols_complete}이(가) df_complete_table에 없습니다.")
        print("컬럼명을 확인하세요. 처리를 중단합니다.")
        df_filtered = pd.DataFrame()

if not df_filtered.empty and not df_complete_table.empty:
    print("조건에 따라 행 재정렬 및 삽입 작업을 시작합니다 (비교 대상: 원본 전체 테이블)...")
    for idx_filtered, current_row in df_filtered.iterrows():
        output_rows_info.append({'data_row': current_row.copy(), 'color_flag': None})

        current_d_val = str(current_row[COLUMN_IF_NAME]) if pd.notna(current_row[COLUMN_IF_NAME]) else ""
        current_d_val_stripped = current_d_val.strip()
        
        current_b_val = str(current_row[COLUMN_SEND_SYSTEM]) if pd.notna(current_row[COLUMN_SEND_SYSTEM]) else ""
        current_c_val = str(current_row[COLUMN_RECV_SYSTEM]) if pd.notna(current_row[COLUMN_RECV_SYSTEM]) else ""

        # 매칭되는 행들을 저장할 리스트
        matching_rows = []

        # 비교 대상은 원본 전체 테이블
        for idx_complete, target_row in df_complete_table.iterrows():
            # 같은 행인지 확인 후 건너뛰기
            if current_row.name == target_row.name:
                continue

            target_d_val = str(target_row[COLUMN_IF_NAME]) if pd.notna(target_row[COLUMN_IF_NAME]) else ""
            target_d_val_stripped = target_d_val.strip()

            # 1차 필터링: I/F명 비교
            cond1_match = (current_d_val_stripped == target_d_val_stripped)
            
            if not cond1_match:
                continue

            target_b_val = str(target_row[COLUMN_SEND_SYSTEM]) if pd.notna(target_row[COLUMN_SEND_SYSTEM]) else ""
            target_c_val = str(target_row[COLUMN_RECV_SYSTEM]) if pd.notna(target_row[COLUMN_RECV_SYSTEM]) else ""

            # 2차 필터링: 송신시스템 조건 (동적 처리)
            cond2_b_match_for_target = False
            for pattern, replacement in SYSTEM_CONVERSION_RULES.items():
                if pattern in current_b_val:
                    transformed_b = current_b_val.replace(pattern, replacement)
                    if target_b_val == transformed_b:
                        cond2_b_match_for_target = True
                        break

            # 3차 필터링: 수신시스템 조건 (동적 처리)
            cond3_c_match_for_target = False
            for pattern, replacement in SYSTEM_CONVERSION_RULES.items():
                if pattern in current_c_val:
                    transformed_c = current_c_val.replace(pattern, replacement)
                    if target_c_val == transformed_c:
                        cond3_c_match_for_target = True
                        break
            
            if cond2_b_match_for_target or cond3_c_match_for_target:
                # 매칭 정보와 함께 행 저장
                matching_rows.append({
                    'row': target_row.copy(), 
                    'b_match': cond2_b_match_for_target,
                    'c_match': cond3_c_match_for_target,
                    'same_b_val': target_b_val == current_b_val,
                    'same_c_val': target_c_val == current_c_val
                })
        
        # 매칭된 행이 있을 경우
        if matching_rows:
            # 매칭된 행이 1개일 경우 그냥 연두색으로 표시
            if len(matching_rows) == 1:
                output_rows_info.append({'data_row': matching_rows[0]['row'], 'color_flag': 'green'})
            else:
                # 매칭된 행이 2개 이상인 경우
                # 디버깅을 위해 모든 매칭 행을 노란색으로 먼저 추가
                if DEBUG_MODE == 1:  # 디버그 모드가 1일 때만
                    for row in matching_rows:
                        output_rows_info.append({'data_row': row['row'], 'color_flag': 'yellow'})
                
                # 우선순위별 필터링
                filtered_row = None
                
                # 케이스 1: 송신시스템과 수신시스템 모두 매칭되는 행
                case1_rows = [row for row in matching_rows if row['b_match'] and row['c_match']]
                if case1_rows:
                    filtered_row = case1_rows[0]['row']
                    print(f"  - 케이스1 적용: 송신/수신시스템 모두 매칭되는 행 선택 (총 {len(case1_rows)}개 중 1개)")
                else:
                    # 케이스 2: 송신시스템 값이 같은 행 선택
                    case2_rows = [row for row in matching_rows if row['same_b_val']]
                    if case2_rows:
                        filtered_row = case2_rows[0]['row']
                        print(f"  - 케이스2 적용: 송신시스템 값이 같은 행 선택 (총 {len(case2_rows)}개 중 1개)")
                    else:
                        # 케이스 2-1: 수신시스템 값이 같은 행 선택
                        case2_1_rows = [row for row in matching_rows if row['same_c_val']]
                        if case2_1_rows:
                            filtered_row = case2_1_rows[0]['row']
                            print(f"  - 케이스2-1 적용: 수신시스템 값이 같은 행 선택 (총 {len(case2_1_rows)}개 중 1개)")
                        else:
                            print(f"  - 케이스 미적용: 모든 매칭 행 {len(matching_rows)}개 처리")
                
                # 우선순위 필터링된 행을 연두색으로 추가
                if filtered_row is not None:
                    output_rows_info.append({'data_row': filtered_row.copy(), 'color_flag': 'green'})
    
    print("행 재정렬 및 삽입 작업 완료.")

# 최종 DataFrame 생성
if output_rows_info:
    # 각 행의 data_row에 color_flag를 컬럼으로 추가
    for item in output_rows_info:
        item['data_row']['color_flag'] = item['color_flag']
    final_df_data = [item['data_row'] for item in output_rows_info]
    # DataFrame 생성 시 컬럼 순서 유지
    cols_for_final_df = column_names_from_db if 'column_names_from_db' in locals() and column_names_from_db else (df_filtered.columns if not df_filtered.empty else None)
    # color_flag 컬럼을 명시적으로 추가
    if cols_for_final_df is not None:
        cols_for_final_df = list(cols_for_final_df) + ['color_flag'] if 'color_flag' not in cols_for_final_df else list(cols_for_final_df)
        df_excel_output = pd.DataFrame(final_df_data, columns=cols_for_final_df).reset_index(drop=True)
    else:
        df_excel_output = pd.DataFrame(final_df_data).reset_index(drop=True)

    # 송신/수신 파일 경로 컬럼 추가
    df_excel_output['송신파일경로'] = df_excel_output.apply(lambda row: create_file_path(row, is_send=True, color_flag=row.get('color_flag')), axis=1)
    df_excel_output['수신파일경로'] = df_excel_output.apply(lambda row: create_file_path(row, is_send=False, color_flag=row.get('color_flag')), axis=1)
    
    # 송신/수신 파일 존재 여부 확인 및 컬럼 추가
    df_excel_output['송신파일존재'] = df_excel_output['송신파일경로'].apply(check_file_exists)
    df_excel_output['수신파일존재'] = df_excel_output['수신파일경로'].apply(check_file_exists)

    # 송신파일생성여부 컬럼 추가
    df_excel_output['송신파일생성여부'] = df_excel_output.apply(lambda row: '' if row.get('color_flag') is not None else ('1' if row.get('개발구분') == '신규' else ''), axis=1)

    # 수신파일생성여부 컬럼 추가
    df_excel_output['수신파일생성여부'] = df_excel_output.apply(lambda row: '' if row.get('color_flag') is not None else '1', axis=1)

    # 송신/수신 디렉토리 파일 개수 계산 함수
    def calc_dir_file_count(row, is_send=True):
        column_name = '송신파일존재' if is_send else '수신파일존재'
        file_path_column = '송신파일경로' if is_send else '수신파일경로'
        
        # 파일이 존재하는 경우에만 디렉토리 파일 개수 계산
        if row[column_name] == 1:
            return count_files_in_directory(row[file_path_column])
        else:
            return 0
    
    # 송신/수신 디렉토리 파일 개수 컬럼 추가
    df_excel_output['송신DF'] = df_excel_output.apply(lambda row: calc_dir_file_count(row, is_send=True), axis=1)
    df_excel_output['수신DF'] = df_excel_output.apply(lambda row: calc_dir_file_count(row, is_send=False), axis=1)

    # 송신/수신 스키마 파일 경로 추가
    df_excel_output['송신스키마파일명'] = df_excel_output.apply(lambda row: create_schema_file_path(row, is_send=True, color_flag=row.get('color_flag')), axis=1)
    df_excel_output['수신스키마파일명'] = df_excel_output.apply(lambda row: create_schema_file_path(row, is_send=False, color_flag=row.get('color_flag')), axis=1)

    # 스키마 파일 존재 여부 확인 및 컬럼 추가
    df_excel_output['송신스키마파일존재'] = df_excel_output['송신스키마파일명'].apply(check_file_exists)
    df_excel_output['수신스키마파일존재'] = df_excel_output['수신스키마파일명'].apply(check_file_exists)

    # 송신스키마파일생성여부 컬럼 추가
    df_excel_output['송신스키마파일생성여부'] = df_excel_output.apply(lambda row: '' if row.get('color_flag') is not None else ('1' if row.get('개발구분') == '신규' else ''), axis=1)

    # 수신스키마파일생성여부 컬럼 추가
    df_excel_output['수신스키마파일생성여부'] = df_excel_output.apply(lambda row: '' if row.get('color_flag') is not None else '1', axis=1)

    # 색상 플래그에 따라 행 인덱스 분리
    yellow_row_indices = [idx for idx, item in enumerate(output_rows_info) if item['color_flag'] == 'yellow']
    green_row_indices = [idx for idx, item in enumerate(output_rows_info) if item['color_flag'] == 'green']
    
    # --- 기본행-매칭행 비교 검증 추가 ---
    # '비교로그' 컬럼 추가
    df_excel_output['비교로그'] = ''
    
    # 기본행과 매칭행 비교 (녹색 행만 대상으로)
    for i in range(0, len(df_excel_output)):
        # 녹색 행이 아니면 건너뛰기
        if i not in green_row_indices:
            continue
            
        # 이전 행이 기본행
        if i == 0:
            print(f"경고: 첫 번째 행이 매칭행입니다. 건너뜁니다.")
            continue
            
        base_row = df_excel_output.iloc[i-1]
        match_row = df_excel_output.iloc[i]
        
        comparison_log = []
        column_names = df_excel_output.columns.tolist()
        
        # 1. 송신시스템 비교
        if COLUMN_SEND_SYSTEM in column_names:
            result = check_systems(base_row[COLUMN_SEND_SYSTEM], match_row[COLUMN_SEND_SYSTEM], '1.송신시스템')
            if result:
                comparison_log.append(result)
        
        # 2. 수신시스템 비교
        if COLUMN_RECV_SYSTEM in column_names:
            result = check_systems(base_row[COLUMN_RECV_SYSTEM], match_row[COLUMN_RECV_SYSTEM], '2.수신시스템')
            if result:
                comparison_log.append(result)
        
        # 3. I/F명 비교
        if COLUMN_IF_NAME in column_names:
            result = check_same_content(base_row[COLUMN_IF_NAME], match_row[COLUMN_IF_NAME], '3.I/F명')
            if result:
                comparison_log.append(result)
        
        # 4. Event_ID 비교
        if COLUMN_EVENT_ID in column_names:
            result = check_table_with_split(base_row[COLUMN_EVENT_ID], match_row[COLUMN_EVENT_ID], '4.Event_ID')
            if result:
                comparison_log.append(result)
        
        # 5. 수신업무명 비교
        if COLUMN_RECV_TASK in column_names:
            result = check_business_name(base_row[COLUMN_RECV_TASK], match_row[COLUMN_RECV_TASK], '5.수신업무명')
            if result:
                comparison_log.append(result)
        
        # 6. 송신업무명 비교
        if COLUMN_SEND_TASK in column_names:
            result = check_business_name(base_row[COLUMN_SEND_TASK], match_row[COLUMN_SEND_TASK], '6.송신업무명')
            if result:
                comparison_log.append(result)
        
        # 7. 송신패키지 비교
        if COLUMN_SEND_PKG in column_names:
            result = check_package(base_row[COLUMN_SEND_PKG], match_row[COLUMN_SEND_PKG], '7.송신패키지')
            if result:
                comparison_log.append(result)
        
        # 8. 수신패키지 비교
        if COLUMN_RECV_PKG in column_names:
            result = check_package(base_row[COLUMN_RECV_PKG], match_row[COLUMN_RECV_PKG], '8.수신패키지')
            if result:
                comparison_log.append(result)
        
        # 9. EMS명 비교
        if COLUMN_EMS_NAME in column_names:
            result = check_same_content(base_row[COLUMN_EMS_NAME], match_row[COLUMN_EMS_NAME], '9.EMS명')
            if result:
                comparison_log.append(result)
        
        # 10. Source Table 비교
        if 'Source Table' in column_names:
            result = check_table_with_split(base_row['Source Table'], match_row['Source Table'], '10.Source Table')
            if result:
                comparison_log.append(result)
        
        # 11. Destination Table 비교
        if 'Destination Table' in column_names:
            result = check_table_with_split(base_row['Destination Table'], match_row['Destination Table'], '11.Destination Table')
            if result:
                comparison_log.append(result)
        
        # 12. Routing 비교
        if 'Routing' in column_names:
            result = check_table_or_routing(base_row['Routing'], match_row['Routing'], '12.Routing')
            if result:
                comparison_log.append(result)
        
        # 13. 스케쥴 비교
        schedule_col = [col for col in column_names if '스케쥴' in col]
        if schedule_col:
            schedule_col = schedule_col[0]
            result = check_same_content(base_row[schedule_col], match_row[schedule_col], '13.스케쥴')
            if result:
                comparison_log.append(result)
        
        # 14. 주기구분 비교
        if '주기구분' in column_names:
            result = check_same_content(base_row['주기구분'], match_row['주기구분'], '14.주기구분')
            if result:
                comparison_log.append(result)
        
        # 15. 주기 비교
        if '주기' in column_names:
            result = check_same_content(base_row['주기'], match_row['주기'], '15.주기')
            if result:
                comparison_log.append(result)
        
        # 비교로그 업데이트
        log_value = ', '.join(comparison_log) if comparison_log else 'OK'
        df_excel_output.at[i-1, '비교로그'] = log_value
        df_excel_output.at[i, '비교로그'] = log_value
    
else:
    df_excel_output = pd.DataFrame()
    yellow_row_indices = []
    green_row_indices = []

# --- DataFrame을 Excel 파일로 저장하고 스타일 적용 ---
if not df_excel_output.empty:
    try:
        # "Unnamed: XX" 형식의 컬럼 중 XX가 10 이상인 컬럼 제외하기
        cols_to_keep = [col for col in df_excel_output.columns
                        if not ((isinstance(col, str) and 
                               col.startswith('Unnamed:') and 
                               len(col.split(':')) > 1 and 
                               col.split(':')[1].strip().isdigit() and 
                               int(col.split(':')[1].strip()) >= 10) or
                               col.startswith('XXXXX'))]
        
        # 필터링된 컬럼만 유지
        df_excel_output = df_excel_output[cols_to_keep]
        
        # 송신/수신 파일 경로 생성 여부 확인 메시지
        print("\n송신 및 수신 파일 경로를 계산했습니다.")
        if DEBUG_MODE == 1:
            # 디버그 모드일 때만 첫 5개 행의 결과 출력
            print("샘플 파일 경로 (처음 5개 행):")
            for idx in range(min(5, len(df_excel_output))):
                print(f"행 {idx+1} - 송신: {df_excel_output.iloc[idx]['송신파일경로']}")
                print(f"행 {idx+1} - 수신: {df_excel_output.iloc[idx]['수신파일경로']}")
        
        print("파일 존재 여부를 확인합니다...")
        send_exists_count = df_excel_output['송신파일존재'].sum()
        recv_exists_count = df_excel_output['수신파일존재'].sum()
        send_schema_exists_count = df_excel_output['송신스키마파일존재'].sum()
        recv_schema_exists_count = df_excel_output['수신스키마파일존재'].sum()
        print(f"송신 파일 존재: {send_exists_count}/{len(df_excel_output)}개")
        print(f"수신 파일 존재: {recv_exists_count}/{len(df_excel_output)}개")
        print(f"송신 스키마 파일 존재: {send_schema_exists_count}/{len(df_excel_output)}개")
        print(f"수신 스키마 파일 존재: {recv_schema_exists_count}/{len(df_excel_output)}개")
        
        print("\n디렉토리 파일 개수를 계산합니다...")
        send_df_total = df_excel_output['송신DF'].sum()
        recv_df_total = df_excel_output['수신DF'].sum()
        send_df_avg = df_excel_output.loc[df_excel_output['송신파일존재'] == 1, '송신DF'].mean() if send_exists_count > 0 else 0
        recv_df_avg = df_excel_output.loc[df_excel_output['수신파일존재'] == 1, '수신DF'].mean() if recv_exists_count > 0 else 0
        
        print(f"송신 디렉토리 총 파일 수: {send_df_total}개")
        print(f"수신 디렉토리 총 파일 수: {recv_df_total}개")
        print(f"송신 디렉토리당 평균 파일 수: {send_df_avg:.1f}개")
        print(f"수신 디렉토리당 평균 파일 수: {recv_df_avg:.1f}개")
        
        with pd.ExcelWriter(excel_filename, engine='xlsxwriter') as writer:
            df_excel_output.to_excel(writer, sheet_name='ProcessedData', index=False)

            workbook = writer.book
            worksheet = writer.sheets['ProcessedData']
            
            # 색상 형식 정의
            yellow_format = workbook.add_format({'bg_color': '#' + COLORS['yellow']})
            green_format = workbook.add_format({'bg_color': '#' + COLORS['green']})
            exist_format = workbook.add_format({'bg_color': '#' + COLORS['green']})
            not_exist_format = workbook.add_format({'bg_color': '#' + COLORS['orange']})
            error_format = workbook.add_format({'bg_color': '#' + COLORS['error']})
            white_format = workbook.add_format({'bg_color': '#' + COLORS['white']})
            
            # 디렉토리 파일 개수에 따른 색상 형식 정의
            df_color_none = workbook.add_format({'bg_color': '#' + COLORS['gray']})
            df_color_very_low = workbook.add_format({'bg_color': '#' + COLORS['blue_very_light']})
            df_color_low = workbook.add_format({'bg_color': '#' + COLORS['blue_light']})
            df_color_medium = workbook.add_format({'bg_color': '#' + COLORS['blue_medium']})
            df_color_high = workbook.add_format({'bg_color': '#' + COLORS['blue_dark']})

            # 노란색 행 적용
            if yellow_row_indices:
                for zero_based_row_idx in yellow_row_indices:
                    worksheet.set_row(zero_based_row_idx + 1, None, yellow_format)
            
            # 연두색 행 적용
            if green_row_indices:
                for zero_based_row_idx in green_row_indices:
                    worksheet.set_row(zero_based_row_idx + 1, None, green_format)
            
            # 송신/수신 파일 존재 여부에 따른 색상 적용
            send_file_exist_col = df_excel_output.columns.get_loc('송신파일존재')
            recv_file_exist_col = df_excel_output.columns.get_loc('수신파일존재')
            send_df_col = df_excel_output.columns.get_loc('송신DF')
            recv_df_col = df_excel_output.columns.get_loc('수신DF')
            
            # 스키마 파일 존재 여부 열 인덱스 찾기
            send_schema_exist_col = df_excel_output.columns.get_loc('송신스키마파일존재')
            recv_schema_exist_col = df_excel_output.columns.get_loc('수신스키마파일존재')
            
            # 비교로그 열 인덱스 찾기
            log_col_idx = df_excel_output.columns.get_loc('비교로그') if '비교로그' in df_excel_output.columns else -1
            
            for row_idx in range(len(df_excel_output)):
                send_exists = df_excel_output.iloc[row_idx]['송신파일존재']
                recv_exists = df_excel_output.iloc[row_idx]['수신파일존재']
                send_df_count = df_excel_output.iloc[row_idx]['송신DF']
                recv_df_count = df_excel_output.iloc[row_idx]['수신DF']
                
                # 스키마 파일 존재 여부
                send_schema_exists = df_excel_output.iloc[row_idx]['송신스키마파일존재']
                recv_schema_exists = df_excel_output.iloc[row_idx]['수신스키마파일존재']
                
                # 송신 파일 존재 여부에 따른 색상 적용
                if send_exists == 1:
                    worksheet.write(row_idx + 1, send_file_exist_col, 1, exist_format)
                else:
                    worksheet.write(row_idx + 1, send_file_exist_col, 0, not_exist_format)
                
                # 수신 파일 존재 여부에 따른 색상 적용
                if recv_exists == 1:
                    worksheet.write(row_idx + 1, recv_file_exist_col, 1, exist_format)
                else:
                    worksheet.write(row_idx + 1, recv_file_exist_col, 0, not_exist_format)
                
                # 송신 스키마 파일 존재 여부에 따른 색상 적용
                if send_schema_exists == 1:
                    worksheet.write(row_idx + 1, send_schema_exist_col, 1, exist_format)
                else:
                    worksheet.write(row_idx + 1, send_schema_exist_col, 0, not_exist_format)
                
                # 수신 스키마 파일 존재 여부에 따른 색상 적용
                if recv_schema_exists == 1:
                    worksheet.write(row_idx + 1, recv_schema_exist_col, 1, exist_format)
                else:
                    worksheet.write(row_idx + 1, recv_schema_exist_col, 0, not_exist_format)
                
                # 송신 디렉토리 파일 개수에 따른 색상 적용
                if send_df_count == 0:
                    worksheet.write(row_idx + 1, send_df_col, send_df_count, df_color_none)
                elif send_df_count <= 3:
                    worksheet.write(row_idx + 1, send_df_col, send_df_count, df_color_very_low)
                elif send_df_count <= 10:
                    worksheet.write(row_idx + 1, send_df_col, send_df_count, df_color_low)
                elif send_df_count <= 20:
                    worksheet.write(row_idx + 1, send_df_col, send_df_count, df_color_medium)
                else:
                    worksheet.write(row_idx + 1, send_df_col, send_df_count, df_color_high)
                
                # 수신 디렉토리 파일 개수에 따른 색상 적용
                if recv_df_count == 0:
                    worksheet.write(row_idx + 1, recv_df_col, recv_df_count, df_color_none)
                elif recv_df_count <= 3:
                    worksheet.write(row_idx + 1, recv_df_col, recv_df_count, df_color_very_low)
                elif recv_df_count <= 10:
                    worksheet.write(row_idx + 1, recv_df_col, recv_df_count, df_color_low)
                elif recv_df_count <= 20:
                    worksheet.write(row_idx + 1, recv_df_col, recv_df_count, df_color_medium)
                else:
                    worksheet.write(row_idx + 1, recv_df_col, recv_df_count, df_color_high)
                
                # 비교로그 오류 표시
                if log_col_idx >= 0:
                    log_value = df_excel_output.iloc[row_idx]['비교로그']
                    if log_value and log_value != 'OK':
                        worksheet.write(row_idx + 1, log_col_idx, log_value, error_format)

            # 컬럼 너비 자동 조절
            for i, col_name_str in enumerate(df_excel_output.columns.astype(str)):
                data_max_len_series = df_excel_output[col_name_str].astype(str).map(len)
                data_max_len = data_max_len_series.max() if not data_max_len_series.empty else 0
                header_len = len(col_name_str)
                if pd.isna(data_max_len): data_max_len = 0
                column_width = max(int(data_max_len), header_len) + 2
                worksheet.set_column(i, i, column_width)

            # 송신파일생성여부/수신파일생성여부 셀 색상 처리
            send_gen_col = df_excel_output.columns.get_loc('송신파일생성여부')
            recv_gen_col = df_excel_output.columns.get_loc('수신파일생성여부')
            send_exist_col = df_excel_output.columns.get_loc('송신파일존재')
            recv_exist_col = df_excel_output.columns.get_loc('수신파일존재')
            
            for row_idx in range(len(df_excel_output)):
                color_flag = df_excel_output.iloc[row_idx].get('color_flag')
                if color_flag is not None:
                    # 매칭행인 경우 흰색으로 칠하기
                    worksheet.write(row_idx + 1, send_gen_col, df_excel_output.iloc[row_idx]['송신파일생성여부'], white_format)
                    worksheet.write(row_idx + 1, recv_gen_col, df_excel_output.iloc[row_idx]['수신파일생성여부'], white_format)
                    continue
                send_gen_val = df_excel_output.iloc[row_idx]['송신파일생성여부']
                recv_gen_val = df_excel_output.iloc[row_idx]['수신파일생성여부']
                send_exist_val = df_excel_output.iloc[row_idx]['송신파일존재']
                recv_exist_val = df_excel_output.iloc[row_idx]['수신파일존재']
                if send_gen_val == '1':
                    if send_exist_val == 1:
                        worksheet.write(row_idx + 1, send_gen_col, send_gen_val, not_exist_format)
                    else:
                        worksheet.write(row_idx + 1, send_gen_col, send_gen_val, exist_format)
                if recv_gen_val == '1':
                    if recv_exist_val == 1:
                        worksheet.write(row_idx + 1, recv_gen_col, recv_gen_val, not_exist_format)
                    else:
                        worksheet.write(row_idx + 1, recv_gen_col, recv_gen_val, exist_format)

            # 송신스키마파일생성여부/수신스키마파일생성여부 셀 색상 처리
            send_schema_gen_col = df_excel_output.columns.get_loc('송신스키마파일생성여부')
            recv_schema_gen_col = df_excel_output.columns.get_loc('수신스키마파일생성여부')
            send_schema_exist_col = df_excel_output.columns.get_loc('송신스키마파일존재')
            recv_schema_exist_col = df_excel_output.columns.get_loc('수신스키마파일존재')
            
            for row_idx in range(len(df_excel_output)):
                color_flag = df_excel_output.iloc[row_idx].get('color_flag')
                if color_flag is not None:
                    # 매칭행인 경우 흰색으로 칠하기
                    worksheet.write(row_idx + 1, send_schema_gen_col, df_excel_output.iloc[row_idx]['송신스키마파일생성여부'], white_format)
                    worksheet.write(row_idx + 1, recv_schema_gen_col, df_excel_output.iloc[row_idx]['수신스키마파일생성여부'], white_format)
                    continue
                send_schema_gen_val = df_excel_output.iloc[row_idx]['송신스키마파일생성여부']
                recv_schema_gen_val = df_excel_output.iloc[row_idx]['수신스키마파일생성여부']
                send_schema_exist_val = df_excel_output.iloc[row_idx]['송신스키마파일존재']
                recv_schema_exist_val = df_excel_output.iloc[row_idx]['수신스키마파일존재']
                if send_schema_gen_val == '1':
                    if send_schema_exist_val == 1:
                        worksheet.write(row_idx + 1, send_schema_gen_col, send_schema_gen_val, not_exist_format)
                    else:
                        worksheet.write(row_idx + 1, send_schema_gen_col, send_schema_gen_val, exist_format)
                if recv_schema_gen_val == '1':
                    if recv_schema_exist_val == 1:
                        worksheet.write(row_idx + 1, recv_schema_gen_col, recv_schema_gen_val, not_exist_format)
                    else:
                        worksheet.write(row_idx + 1, recv_schema_gen_col, recv_schema_gen_val, exist_format)

        print(f"\n결과가 '{excel_filename}' 파일로 저장되었습니다.")
        if DEBUG_MODE == 1:
            print("매칭된 모든 행은 노란색으로, 우선순위로 필터링된 행은 연두색으로 표시됩니다.")
        else:
            print("우선순위로 필터링된 행은 연두색으로 표시됩니다.")
        print("파일 존재 여부 컬럼: 존재하면 1(연두색), 존재하지 않으면 0(주황색)으로 표시됩니다.")
        print("'송신스키마파일존재'/'수신스키마파일존재' 컬럼: 존재하면 1(연두색), 존재하지 않으면 0(주황색)으로 표시됩니다.")
        print("'송신스키마파일명'/'수신스키마파일명' 컬럼: 스키마 파일 경로를 나타냅니다.")
        print("'송신DF'와 '수신DF' 컬럼은 각 파일이 위치한 디렉토리의 파일 개수를 나타냅니다.")
        print("  - 파일 개수가 0개: 회색")
        print("  - 파일 개수가 1-3개: 매우 밝은 파란색")
        print("  - 파일 개수가 4-10개: 밝은 파란색")
        print("  - 파일 개수가 11-20개: 중간 파란색")
        print("  - 파일 개수가 21개 이상: 진한 파란색")
        print("'비교로그' 컬럼: 오류가 있으면 주황색으로 표시됩니다.")
        
        # 설정된 변환 규칙 출력
        print("\n현재 설정된 변환 규칙:")
        for pattern, replacement in SYSTEM_CONVERSION_RULES.items():
            print(f"  {pattern} → {replacement}")

    except ImportError:
        print("Excel 파일 저장을 위해 'xlsxwriter' 라이브러리가 필요합니다. 'pip install xlsxwriter' 명령어로 설치해주세요.")
    except Exception as e_excel:
        print(f"Excel 파일 저장 중 오류 발생: {e_excel}")

elif not df_complete_table.empty and df_filtered.empty:
    print("초기 필터링된 데이터(df_filtered)가 없어 Excel 파일을 생성하지 않았습니다.")
elif df_complete_table.empty:
     print("원본 데이터(df_complete_table)가 없어 Excel 파일을 생성하지 않았습니다.")
else:
    print("조건에 맞는 데이터가 없어 최종적으로 Excel 파일에 저장할 내용이 없습니다.")