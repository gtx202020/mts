#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
iflist04.py - 엑셀 파일 검증 프로그램

이 파일은 iflist03.py의 작업을 이어서 수행하는 내용으로,
iflist03.py의 출력결과물인 엑셀파일을 기반으로 작업합니다.
기본행과 매칭행을 비교하여 다양한 검증 규칙에 따라 오류를 확인하고
결과를 '비교로그' 컬럼에 기록합니다.

개발 버전: v1.0
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
import re

# 오류 표시를 위한 주황색 배경 정의
ORANGE_FILL = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

def replace_ly_lz(text):
    """문자열에서 'LY'를 'LH'로, 'LZ'를 'VO'로 교체"""
    if not isinstance(text, str):
        return text
    result = text.replace('LY', 'LH').replace('LZ', 'VO')
    return result

def check_systems(base_value, match_value, column_name):
    """송신시스템/수신시스템 비교 로직"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (비어있는 값)"
    
    expected_value = replace_ly_lz(base_value)
    if expected_value != match_value:
        return f"{column_name} 비교오류"
    return ""

def check_business_name(base_value, match_value, column_name):
    """업무명 비교 로직"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (비어있는 값)"
    
    if base_value == 'PNL_LY' and match_value != 'MES_LH':
        return f"{column_name} 비교오류"
    elif base_value == 'MOD_LZ' and match_value != 'MES_VO':
        return f"{column_name} 비교오류"
    return ""

def check_package(base_value, match_value, column_name):
    """패키지 비교 로직"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (비어있는 값)"
    
    if 'LY' in base_value and 'LH' not in match_value:
        return f"{column_name} 비교오류"
    elif 'LZ' in base_value and 'VO' not in match_value:
        return f"{column_name} 비교오류"
    return ""

def check_same_content(base_value, match_value, column_name):
    """내용이 같아야 하는 컬럼 비교"""
    if not isinstance(base_value, str) and not isinstance(match_value, str):
        # 둘 다 문자열이 아닌 경우 (예: NaN)
        if pd.isna(base_value) and pd.isna(match_value):
            return ""
        return f"{column_name} 비교오류 (유형 불일치)"
    
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        return f"{column_name} 비교오류 (유형 불일치)"
    
    if base_value.strip() != match_value.strip():
        return f"{column_name} 비교오류"
    return ""

def check_table_or_routing(base_value, match_value, column_name):
    """테이블명 또는 라우팅 비교 로직"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        if pd.isna(base_value) and pd.isna(match_value):
            return ""
        return f"{column_name} 비교오류 (비어있는 값)"
    
    if 'LY' in base_value or 'LZ' in base_value:
        expected_value = replace_ly_lz(base_value)
        if expected_value != match_value:
            return f"{column_name} 비교오류"
    elif base_value.strip() != match_value.strip():
        return f"{column_name} 비교오류"
    return ""

def check_table_with_split(base_value, match_value, column_name):
    """Source Table, Destination Table용 비교 로직 (단어 분할 후 LY/LZ 확인)"""
    if not isinstance(base_value, str) or not isinstance(match_value, str):
        if pd.isna(base_value) and pd.isna(match_value):
            return ""
        return f"{column_name} 비교오류 (비어있는 값)"
    
    # 기본값으로 단순 비교 
    should_check_ly_lz = False
    
    # '.'과 '_'로 분할하여 단어 확인
    words = re.split('[._]', base_value)
    for word in words:
        if word.startswith('LY') or word.startswith('LZ'):
            should_check_ly_lz = True
            break
    
    if should_check_ly_lz:
        expected_value = replace_ly_lz(base_value)
        if expected_value != match_value:
            return f"{column_name} 비교오류"
    elif base_value.strip() != match_value.strip():
        return f"{column_name} 비교오류"
    
    return ""

def validate_excel_file(input_file):
    """엑셀 파일을 읽고 검증을 수행하는 함수"""
    try:
        # 엑셀 파일 로드
        print(f"파일 '{input_file}'을 로드 중...")
        df = pd.read_excel(input_file, engine='openpyxl')
        print(f"파일 로드 완료. 총 {len(df)} 행을 분석합니다.")
        
        # 열 이름 확인
        column_names = df.columns.tolist()
        print(f"찾은 컬럼: {column_names}")
        
        # '비교로그' 컬럼 추가
        df['비교로그'] = ''
        
        # 기본행과 매칭행 비교
        for i in range(0, len(df), 2):
            if i+1 >= len(df):
                print(f"경고: 행 {i}의 매칭행이 없습니다. 건너뜁니다.")
                continue
                
            base_row = df.iloc[i]
            match_row = df.iloc[i+1]
            
            comparison_log = []
            
            # 1. 송신시스템 비교
            if '송신시스템' in column_names:
                result = check_systems(base_row['송신시스템'], match_row['송신시스템'], '1.송신시스템')
                if result:
                    comparison_log.append(result)
            
            # 2. 수신시스템 비교
            if '수신시스템' in column_names:
                result = check_systems(base_row['수신시스템'], match_row['수신시스템'], '2.수신시스템')
                if result:
                    comparison_log.append(result)
            
            # 3. I/F명 비교
            if 'I/F명' in column_names:
                result = check_same_content(base_row['I/F명'], match_row['I/F명'], '3.I/F명')
                if result:
                    comparison_log.append(result)
            
            # 4. Event_ID 비교
            if 'Event_ID' in column_names:
                result = check_table_with_split(base_row['Event_ID'], match_row['Event_ID'], '4.Event_ID')
                if result:
                    comparison_log.append(result)
            
            # 5. 수신업무명 비교
            if '수신업무명' in column_names:
                result = check_business_name(base_row['수신업무명'], match_row['수신업무명'], '5.수신업무명')
                if result:
                    comparison_log.append(result)
            
            # 6. 송신업무명 비교
            if '송신업무명' in column_names:
                result = check_business_name(base_row['송신업무명'], match_row['송신업무명'], '6.송신업무명')
                if result:
                    comparison_log.append(result)
            
            # 7. 송신패키지 비교
            if '송신패키지' in column_names:
                result = check_package(base_row['송신패키지'], match_row['송신패키지'], '7.송신패키지')
                if result:
                    comparison_log.append(result)
            
            # 8. 수신패키지 비교
            if '수신패키지' in column_names:
                result = check_package(base_row['수신패키지'], match_row['수신패키지'], '8.수신패키지')
                if result:
                    comparison_log.append(result)
            
            # 9. EMS명 비교
            if 'EMS명' in column_names:
                result = check_same_content(base_row['EMS명'], match_row['EMS명'], '9.EMS명')
                if result:
                    comparison_log.append(result)
            
            # 10. Source Table 비교
            if 'Source Table' in column_names:
                result = check_table_with_split(base_row['Source Table'], match_row['Source Table'], '10.Source Table')
                if result:
                    comparison_log.append(result)
            
            # 11. Destination Table 비교
            if 'Destination Table' in column_names:
                result = check_table_with_split(base_row['Destination Table'], match_row['Destination Table'], '11.Destination Table')
                if result:
                    comparison_log.append(result)
            
            # 12. Routing 비교
            if 'Routing' in column_names:
                result = check_table_or_routing(base_row['Routing'], match_row['Routing'], '12.Routing')
                if result:
                    comparison_log.append(result)
            
            # 13. 스케쥴 비교
            schedule_col = [col for col in column_names if '스케쥴' in col]
            if schedule_col:
                schedule_col = schedule_col[0]
                result = check_same_content(base_row[schedule_col], match_row[schedule_col], '13.스케쥴')
                if result:
                    comparison_log.append(result)
            
            # 14. 주기구분 비교
            if '주기구분' in column_names:
                result = check_same_content(base_row['주기구분'], match_row['주기구분'], '14.주기구분')
                if result:
                    comparison_log.append(result)
            
            # 15. 주기 비교
            if '주기' in column_names:
                result = check_same_content(base_row['주기'], match_row['주기'], '15.주기')
                if result:
                    comparison_log.append(result)
            
            # 비교로그 업데이트
            log_value = ', '.join(comparison_log) if comparison_log else 'OK'
            df.at[i, '비교로그'] = log_value
            df.at[i+1, '비교로그'] = log_value
        
        # 결과 저장
        output_file = input_file.replace('.xlsx', '_검증결과.xlsx')
        df.to_excel(output_file, index=False)
        print(f"검증 결과가 '{output_file}'에 저장되었습니다.")
        
        # 주황색 배경 적용
        apply_formatting(output_file)
        print("형식 적용 완료.")
        
        return output_file
    
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return None

def apply_formatting(file_path):
    """검증 결과 파일에 주황색 배경 적용"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 헤더 행 찾기
        header_row = 1
        
        # 비교로그 열 인덱스 찾기
        log_col_idx = None
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value == '비교로그':
                log_col_idx = col_idx
                break
        
        if not log_col_idx:
            print("'비교로그' 열을 찾을 수 없습니다.")
            return
        
        # 오류가 있는 셀에 주황색 배경 적용
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=log_col_idx)
            if cell.value and cell.value != 'OK':
                cell.fill = ORANGE_FILL
        
        wb.save(file_path)
        print(f"색상 형식이 '{file_path}'에 적용되었습니다.")
    
    except Exception as e:
        print(f"형식 적용 중 오류 발생: {str(e)}")

def find_latest_excel_file(directory='.'):
    """가장 최근의 iflist03 엑셀 파일 찾기"""
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx') and 'iflist03' in f]
    if not files:
        return None
    
    files.sort(key=lambda x: os.path.getmtime(os.path.join(directory, x)), reverse=True)
    return os.path.join(directory, files[0])

def main():
    """메인 함수"""
    print("iflist04.py - 엑셀 파일 검증 프로그램 실행")
    
    # 명령행 인수로 파일 경로를 받거나 자동으로 찾기
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = find_latest_excel_file()
        if not input_file:
            print("iflist03 엑셀 파일을 찾을 수 없습니다. 파일 경로를 인수로 지정해주세요.")
            print("사용법: python iflist04.py <엑셀파일경로>")
            return
    
    if not os.path.exists(input_file):
        print(f"오류: 파일 '{input_file}'이 존재하지 않습니다.")
        return
    
    print(f"입력 파일: {input_file}")
    validate_excel_file(input_file)
    print("검증 프로세스 완료.")

if __name__ == "__main__":
    main() 