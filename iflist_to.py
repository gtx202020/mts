import pandas as pd
import openpyxl
import yaml
import os
import datetime
import shutil
from openpyxl.styles import Font, PatternFill, Alignment

# 디버그 모드 설정
DEBUG_MODE = True

def debug_print(*args, **kwargs):
    """디버그 모드일 때만 메시지를 출력하는 함수"""
    if DEBUG_MODE:
        print("[DEBUG]", *args, **kwargs)

def process_file_path(file_path, check_flag):
    """
    파일 경로를 TEST에서 PROD로 변환하고 관련 정보를 수집
    
    Args:
        file_path: 원본 파일 경로
        check_flag: 생성여부 플래그 (1인 경우만 처리)
    
    Returns:
        tuple: (PROD 경로, 파일 존재 여부, 디렉토리 파일 개수)
    """
    # 생성여부가 1이 아니면 처리하지 않음
    if pd.isna(check_flag) or float(check_flag) != 1.0:
        return None, None, None
    
    if pd.isna(file_path) or not isinstance(file_path, str):
        return None, None, None
    
    # TEST → PROD 변환
    prod_path = file_path.replace('_TEST_SOURCE', '_PROD_SOURCE')
    
    # 파일 존재 여부 확인
    file_exists = "1" if os.path.exists(prod_path) else "0"
    
    # 디렉토리 파일 개수 확인
    dir_path = os.path.dirname(prod_path)
    if os.path.exists(dir_path):
        try:
            file_count = str(len(os.listdir(dir_path)))
        except:
            file_count = "X"
    else:
        file_count = "X"
    
    return prod_path, file_exists, file_count

def generate_excel_and_yaml(input_excel_path, output_excel_path, output_yaml_path):
    """
    입력 엑셀 파일을 읽어 TEST→PROD 변환 후 결과를 엑셀과 YAML로 출력
    
    Args:
        input_excel_path: 입력 엑셀 파일 경로
        output_excel_path: 출력 엑셀 파일 경로
        output_yaml_path: 출력 YAML 파일 경로
    """
    debug_print(f"입력 엑셀 파일 읽기: {input_excel_path}")
    
    # 엑셀 파일 읽기
    df = pd.read_excel(input_excel_path, engine='openpyxl')
    debug_print(f"총 {len(df)}개 행 읽기 완료")
    
    # 처리할 컬럼 정의 - 송신과 수신을 분리
    send_types = [
        ('송신파일경로', '송신파일생성여부'),
        ('송신스키마파일명', '송신스키마파일생성여부')
    ]
    recv_types = [
        ('수신파일경로', '수신파일생성여부'),
        ('수신스키마파일명', '수신스키마파일생성여부')
    ]
    
    # YAML 데이터 준비
    yaml_data = {'files': []}
    
    # 새로운 데이터프레임 준비
    result_data = []
    
    # 각 행 처리
    for idx, row in df.iterrows():
        # PROD_SOURCE가 포함된 행은 건너뛰기
        skip_row = False
        
        # 송신/수신 파일경로와 스키마파일명 확인
        for col in ['송신파일경로', '수신파일경로', '송신스키마파일명', '수신스키마파일명']:
            if col in df.columns:
                value = row.get(col)
                if not pd.isna(value) and isinstance(value, str) and 'PROD_SOURCE' in value:
                    skip_row = True
                    debug_print(f"Row {idx} skipped: PROD_SOURCE found in {col}")
                    break
        
        if skip_row:
            continue
        
        # 송신 데이터 처리
        send_data = {"구분": "송신"}
        
        for file_col, check_col in send_types:
            if file_col in df.columns and check_col in df.columns:
                file_path = row.get(file_col)
                check_flag = row.get(check_col)
                
                # 컬럼명에서 "송신" 제거
                clean_file_col = file_col.replace('송신', '')
                clean_check_col = check_col.replace('송신', '')
                
                # 원본 데이터 저장
                send_data[clean_file_col] = file_path if not pd.isna(file_path) else ""
                
                # 생성여부 처리 (1이 아니면 0으로)
                if pd.isna(check_flag):
                    send_data[clean_check_col] = "0"
                else:
                    send_data[clean_check_col] = "1" if float(check_flag) == 1.0 else "0"
                
                # 생성여부가 1인 경우만 PROD 변환 처리
                if not pd.isna(check_flag) and float(check_flag) == 1.0 and not pd.isna(file_path) and isinstance(file_path, str):
                    prod_path, file_exists, file_count = process_file_path(file_path, check_flag)
                    
                    # PROD 데이터 저장
                    send_data[f"{clean_file_col}PROD"] = prod_path
                    send_data[f"{clean_check_col}PROD"] = file_exists
                    
                    # DFPROD 데이터 저장
                    if '파일경로' in clean_file_col:
                        send_data["DFPROD"] = file_count
                    else:  # 스키마파일명
                        send_data["스키마DFPROD"] = file_count
                    
                    # YAML 데이터 추가
                    if prod_path and file_path:
                        yaml_data['files'].append({
                            'source': file_path,
                            'destination': prod_path
                        })
                else:
                    # 생성여부가 0인 경우 빈 값으로 설정
                    send_data[f"{clean_file_col}PROD"] = ""
                    send_data[f"{clean_check_col}PROD"] = "0"
                    if '파일경로' in clean_file_col:
                        send_data["DFPROD"] = ""
                    else:
                        send_data["스키마DFPROD"] = ""
        
        # 송신 데이터 추가
        result_data.append(send_data)
        
        # 수신 데이터 처리
        recv_data = {"구분": "수신"}
        
        for file_col, check_col in recv_types:
            if file_col in df.columns and check_col in df.columns:
                file_path = row.get(file_col)
                check_flag = row.get(check_col)
                
                # 컬럼명에서 "수신" 제거
                clean_file_col = file_col.replace('수신', '')
                clean_check_col = check_col.replace('수신', '')
                
                # 원본 데이터 저장
                recv_data[clean_file_col] = file_path if not pd.isna(file_path) else ""
                
                # 생성여부 처리 (1이 아니면 0으로)
                if pd.isna(check_flag):
                    recv_data[clean_check_col] = "0"
                else:
                    recv_data[clean_check_col] = "1" if float(check_flag) == 1.0 else "0"
                
                # 생성여부가 1인 경우만 PROD 변환 처리
                if not pd.isna(check_flag) and float(check_flag) == 1.0 and not pd.isna(file_path) and isinstance(file_path, str):
                    prod_path, file_exists, file_count = process_file_path(file_path, check_flag)
                    
                    # PROD 데이터 저장
                    recv_data[f"{clean_file_col}PROD"] = prod_path
                    recv_data[f"{clean_check_col}PROD"] = file_exists
                    
                    # DFPROD 데이터 저장
                    if '파일경로' in clean_file_col:
                        recv_data["DFPROD"] = file_count
                    else:  # 스키마파일명
                        recv_data["스키마DFPROD"] = file_count
                    
                    # YAML 데이터 추가
                    if prod_path and file_path:
                        yaml_data['files'].append({
                            'source': file_path,
                            'destination': prod_path
                        })
                else:
                    # 생성여부가 0인 경우 빈 값으로 설정
                    recv_data[f"{clean_file_col}PROD"] = ""
                    recv_data[f"{clean_check_col}PROD"] = "0"
                    if '파일경로' in clean_file_col:
                        recv_data["DFPROD"] = ""
                    else:
                        recv_data["스키마DFPROD"] = ""
        
        # 수신 데이터 추가
        result_data.append(recv_data)
    
    # 결과 데이터프레임 생성
    result_df = pd.DataFrame(result_data)
    
    # 컬럼 순서 정의 - 구분 컬럼을 맨 앞에, 송신/수신 제거된 컬럼명 사용
    all_columns = ['구분']
    # 파일경로 관련 컬럼
    all_columns.extend(['파일경로', '파일경로PROD', '파일생성여부', '파일생성여부PROD', 'DFPROD'])
    # 스키마파일명 관련 컬럼
    all_columns.extend(['스키마파일명', '스키마파일명PROD', '스키마파일생성여부', '스키마파일생성여부PROD', '스키마DFPROD'])
    
    # 모든 컬럼에 대해 빈 값으로 채우기
    for col in all_columns:
        if col not in result_df.columns:
            result_df[col] = ''
    
    # 컬럼 순서 재정렬
    result_df = result_df[all_columns]
    
    # 엑셀 파일 저장
    save_excel_with_style(result_df, output_excel_path)
    
    # YAML 파일 저장
    with open(output_yaml_path, 'w', encoding='utf-8') as yf:
        yaml.dump(yaml_data, yf, allow_unicode=True, sort_keys=False)
    
    print(f"\n엑셀 파일 생성 완료: {output_excel_path}")
    print(f"YAML 파일 생성 완료: {output_yaml_path}")
    print(f"총 {len(yaml_data['files'])}개 파일 매핑 생성")
    
    return len(yaml_data['files'])

def save_excel_with_style(df, excel_path):
    """데이터프레임을 스타일이 적용된 엑셀 파일로 저장"""
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    df.to_excel(writer, index=False)
    
    # 워크시트 가져오기
    worksheet = writer.sheets['Sheet1']
    
    # 폰트 크기를 10으로 설정
    font_10 = Font(size=10)
    
    # 헤더(첫 행) 스타일 설정
    light_blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    header_font = Font(size=10, bold=True)
    
    # 모든 셀에 대해 폰트 크기 10 적용
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = font_10
    
    # 헤더 행에 스타일 적용
    for cell in worksheet[1]:
        cell.fill = light_blue_fill
        cell.font = header_font
    
    # 컬럼 너비 자동 조절
    for column in worksheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        
        # 각 셀의 길이를 확인하여 최대 길이 계산
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # 컬럼 너비 설정 (최대 길이 + 여유 공간)
        adjusted_width = min(max_length + 2, 50)  # 최대 50으로 제한
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    writer.close()

def execute_file_copy(yaml_path, log_path):
    """
    YAML 파일에 정의된 파일 복사 작업을 실행
    
    Args:
        yaml_path: YAML 파일 경로
        log_path: 로그 파일 경로
    """
    try:
        with open(yaml_path, 'r', encoding='utf-8') as yf:
            data = yaml.safe_load(yf)
    except FileNotFoundError:
        print(f"YAML 파일을 찾을 수 없습니다: {yaml_path}")
        return
    
    files = data.get('files', [])
    if not files:
        print("복사할 파일이 없습니다.")
        return
    
    # 로그 파일 초기화
    with open(log_path, 'w', encoding='utf-8') as lf:
        lf.write(f"[{datetime.datetime.now()}] 파일 복사 시작\n")
    
    success_count = 0
    skip_count = 0
    error_count = 0
    
    for file_info in files:
        source = file_info.get('source')
        destination = file_info.get('destination')
        
        if not source or not destination:
            continue
        
        # 원본 파일 존재 확인
        if not os.path.exists(source):
            with open(log_path, 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.datetime.now()}] [ERROR] 원본 파일 없음: {source}\n")
            error_count += 1
            continue
        
        # 대상 파일이 이미 존재하는지 확인
        if os.path.exists(destination):
            with open(log_path, 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.datetime.now()}] [ERROR] 파일이 이미 존재: {destination}\n")
            skip_count += 1
            continue
        
        # 대상 디렉토리 생성
        dest_dir = os.path.dirname(destination)
        try:
            os.makedirs(dest_dir, exist_ok=True)
        except Exception as e:
            with open(log_path, 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.datetime.now()}] [ERROR] 디렉토리 생성 실패: {dest_dir} - {str(e)}\n")
            error_count += 1
            continue
        
        # 파일 복사
        try:
            shutil.copy2(source, destination)
            with open(log_path, 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.datetime.now()}] 복사 성공: {source} → {destination}\n")
            success_count += 1
        except Exception as e:
            with open(log_path, 'a', encoding='utf-8') as lf:
                lf.write(f"[{datetime.datetime.now()}] [ERROR] 복사 실패: {source} → {destination} - {str(e)}\n")
            error_count += 1
    
    # 최종 결과 로그
    with open(log_path, 'a', encoding='utf-8') as lf:
        lf.write(f"\n[{datetime.datetime.now()}] 파일 복사 완료\n")
        lf.write(f"성공: {success_count}개, 건너뜀: {skip_count}개, 오류: {error_count}개\n")
    
    print(f"\n파일 복사 완료")
    print(f"성공: {success_count}개")
    print(f"건너뜀: {skip_count}개 (이미 존재)")
    print(f"오류: {error_count}개")
    print(f"로그 파일: {log_path}")

def main():
    """메인 함수 - 메뉴 시스템"""
    while True:
        print("\n=== 파일 경로 변환 도구 ===")
        print("1. 엑셀 분석 및 YAML 생성")
        print("2. 파일 복사 실행")
        print("0. 종료")
        
        choice = input("\n원하는 작업을 선택하세요: ").strip()
        
        if choice == "1":
            input_excel = input("입력 엑셀 파일 경로를 입력하세요: ").strip()
            if not input_excel:
                print("파일 경로를 입력해주세요.")
                continue
            
            # 출력 파일 경로 설정
            output_excel = "iflist_to.xlsx"
            output_yaml = "iflist_to.yaml"
            
            try:
                count = generate_excel_and_yaml(input_excel, output_excel, output_yaml)
                print(f"\n작업 완료: {count}개 파일 매핑 생성")
            except Exception as e:
                print(f"오류 발생: {str(e)}")
        
        elif choice == "2":
            yaml_path = input("YAML 파일 경로를 입력하세요 (기본값: iflist_to.yaml): ").strip()
            if not yaml_path:
                yaml_path = "iflist_to.yaml"
            
            # 로그 파일 이름 생성
            log_path = f"file_copy_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            
            try:
                execute_file_copy(yaml_path, log_path)
            except Exception as e:
                print(f"오류 발생: {str(e)}")
        
        elif choice == "0":
            print("프로그램을 종료합니다.")
            break
        
        else:
            print("잘못된 선택입니다. 다시 시도하세요.")

if __name__ == "__main__":
    main()