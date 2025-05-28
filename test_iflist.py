"""
인터페이스 정보 엑셀 파일 리더 및 BW 수신파일 파서 모듈

이 모듈은 다음과 같은 기능을 제공합니다:
1. 특정 형식의 엑셀 파일에서 인터페이스 정보를 읽어 파이썬 자료구조로 변환
2. TIBCO BW .process 파일에서 수신용 INSERT 쿼리를 추출하고 파라미터 매핑 처리

주요 클래스:
- InterfaceExcelReader: 엑셀 파일에서 인터페이스 정보 추출
- BWProcessFileParser: BW .process 파일에서 INSERT 쿼리 추출
- ProcessFileMapper: 일련번호와 string_replacer용 엑셀을 매핑하는 클래스

주요 함수:
- parse_bw_receive_file: BW 수신파일 파싱을 위한 편의 함수
"""

import os
import ast
import re
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import datetime


class InterfaceExcelReader:
    """
    인터페이스 정보가 담긴 엑셀 파일을 읽어 파이썬 자료구조로 변환하는 클래스
    
    엑셀 파일 구조:
    - B열부터 3컬럼 단위로 하나의 인터페이스 블록
    - 1행: 인터페이스명
    - 2행: 인터페이스ID  
    - 3행: DB 연결 정보 (문자열로 저장된 딕셔너리)
    - 4행: 테이블 정보 (문자열로 저장된 딕셔너리)
    - 5행부터: 컬럼 매핑 정보
    """
    
    def __init__(self, replacer_excel_path: str = None):
        """
        InterfaceExcelReader 클래스 초기화
        
        Args:
            replacer_excel_path (str, optional): string_replacer용 엑셀 파일 경로
                                               None이면 기본값으로 'iflist03a_reordered_v8.3.xlsx' 사용
        """
        self.processed_count = 0
        self.error_count = 0
        self.last_error_messages = []
        
        # ProcessFileMapper 초기화 - 하드코딩된 기본 파일 경로 사용
        if replacer_excel_path is None:
            replacer_excel_path = "iflist03a_reordered_v8.3.xlsx"  # 기본 파일 경로
        
        self.process_mapper = None
        if os.path.exists(replacer_excel_path):
            try:
                self.process_mapper = ProcessFileMapper(replacer_excel_path)
                print(f"Info: ProcessFileMapper 초기화 완료 - 파일: {replacer_excel_path}")
            except Exception as e:
                print(f"Warning: ProcessFileMapper 초기화 실패: {str(e)}")
        else:
            print(f"Warning: ProcessFileMapper용 파일이 존재하지 않음: {replacer_excel_path}")
    
    def load_interfaces(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        엑셀 파일에서 모든 인터페이스 정보를 읽어 리스트로 반환
        
        Args:
            excel_path (str): 읽을 엑셀 파일의 경로
            
        Returns:
            List[Dict[str, Any]]: 인터페이스 정보 딕셔너리들의 리스트
            
        Raises:
            FileNotFoundError: 엑셀 파일이 존재하지 않는 경우
            PermissionError: 파일 접근 권한이 없는 경우
            ValueError: 엑셀 파일 형식이 올바르지 않은 경우
        """
        # 초기화
        self.processed_count = 0
        self.error_count = 0
        self.last_error_messages = []
        
        # 파일 존재 여부 확인
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")
        
        interfaces = []
        workbook = None
        
        try:
            # 엑셀 파일 열기
            workbook = openpyxl.load_workbook(excel_path, read_only=True)
            worksheet = workbook.active
            
            if worksheet is None:
                raise ValueError("활성 워크시트를 찾을 수 없습니다")
            
            # B열부터 시작하여 3컬럼 단위로 처리 (전체 인터페이스 처리)
            current_col = 2  # B열 = 2
            
            while current_col <= worksheet.max_column:
                try:
                    # 인터페이스 블록 읽기
                    interface_data = self._read_interface_block(worksheet, current_col)
                    
                    if interface_data is None:
                        # 빈 인터페이스 발견시 종료
                        break
                    
                    interfaces.append(interface_data)
                    self.processed_count += 1
                    print(f"인터페이스 {self.processed_count} 처리 완료: {interface_data.get('interface_name', 'Unknown')}")
                    
                except Exception as e:
                    self.error_count += 1
                    error_msg = f"컬럼 {current_col}에서 오류 발생: {str(e)}"
                    self.last_error_messages.append(error_msg)
                    print(f"Warning: {error_msg}")
                
                # 다음 인터페이스 블록으로 이동 (3컬럼씩)
                current_col += 3
            
            # 디버깅용 코드는 주석 처리
            """
            # [디버깅용] 첫 번째 인터페이스 블록만 처리 (B열부터 시작)
            current_col = 2  # B열 = 2
            
            try:
                print(f"=== 디버깅 모드: 첫 번째 인터페이스만 처리 (컬럼 {current_col}) ===")
                
                # 첫 번째 인터페이스 블록 읽기
                interface_data = self._read_interface_block(worksheet, current_col)
                
                if interface_data is not None:
                    interfaces.append(interface_data)
                    self.processed_count += 1
                    print(f"첫 번째 인터페이스 처리 완료: {interface_data.get('interface_name', 'Unknown')}")
                else:
                    print("첫 번째 인터페이스 블록이 비어있습니다.")
                
            except Exception as e:
                self.error_count += 1
                error_msg = f"첫 번째 인터페이스 블록(컬럼 {current_col})에서 오류 발생: {str(e)}"
                self.last_error_messages.append(error_msg)
                print(f"Warning: {error_msg}")
            """
                
        except Exception as e:
            raise ValueError(f"엑셀 파일 처리 중 오류 발생: {str(e)}")
        
        finally:
            # 리소스 정리
            if workbook:
                workbook.close()
        
        return interfaces
    
    def _read_interface_block(self, worksheet: Worksheet, start_col: int) -> Optional[Dict[str, Any]]:
        """
        단일 인터페이스 블록(3컬럼)에서 정보를 읽어 딕셔너리로 반환
        
        Args:
            worksheet: 엑셀 워크시트 객체
            start_col (int): 인터페이스 블록의 시작 컬럼 번호
            
        Returns:
            Optional[Dict[str, Any]]: 인터페이스 정보 딕셔너리, 빈 블록이면 None
        """
        # 기본 구조 생성
        interface_info = {
            'interface_name': '',
            'interface_id': '',
            'serial_number': '',
            'send_original': '',        # 송신 원본파일 경로
            'send_copy': '',            # 송신 복사파일 경로  
            'recv_original': '',        # 수신 원본파일 경로
            'recv_copy': '',            # 수신 복사파일 경로
            'send_schema': '',          # 송신 스키마파일
            'recv_schema': '',          # 수신 스키마파일
            'send': {
                'owner': None,
                'table_name': None,
                'columns': [],
                'db_info': {}
            },
            'recv': {
                'owner': None,
                'table_name': None,
                'columns': [],
                'db_info': {}
            }
        }
        
        # 1단계: 필수 정보만 먼저 체크 (interface_name, interface_id)
        try:
            # 1행: 인터페이스명 읽기
            interface_name_cell = worksheet.cell(row=1, column=start_col)
            interface_info['interface_name'] = interface_name_cell.value or ''
            
            # 1행: 일련번호 읽기 (인터페이스명 오른쪽으로 두 칸)
            serial_number_cell = worksheet.cell(row=1, column=start_col + 2)
            interface_info['serial_number'] = serial_number_cell.value or ''
            
            # 2행: 인터페이스ID 읽기 (필수값)
            interface_id_cell = worksheet.cell(row=2, column=start_col)
            interface_id = interface_id_cell.value
            
            if not interface_id:
                # 인터페이스 ID가 없으면 빈 블록으로 간주
                return None
            
            interface_info['interface_id'] = str(interface_id).strip()
            
        except Exception as e:
            print(f"Warning: 필수 정보 읽기 실패 (컬럼 {start_col}): {str(e)}")
            return None
        
        # 2단계: 선택적 정보들을 개별적으로 안전하게 처리
        # DB 연결 정보 읽기 (실패해도 계속)
        try:
            send_db_cell = worksheet.cell(row=3, column=start_col)
            recv_db_cell = worksheet.cell(row=3, column=start_col + 1)
            
            interface_info['send']['db_info'] = self._parse_cell_dict(send_db_cell.value)
            interface_info['recv']['db_info'] = self._parse_cell_dict(recv_db_cell.value)
            
        except Exception as e:
            print(f"Warning: DB 정보 읽기 실패 (컬럼 {start_col}): {str(e)}")
            # DB 정보 읽기 실패해도 빈 딕셔너리로 계속 진행
        
        # 테이블 정보 읽기 (실패해도 계속)
        try:
            # 송신 테이블 정보 읽기 (row=4, column=start_col)
            send_table_cell = worksheet.cell(row=4, column=start_col)
            send_table_dict = self._parse_cell_dict(send_table_cell.value)
            if send_table_dict:
                interface_info['send']['owner'] = send_table_dict.get('owner')
                interface_info['send']['table_name'] = send_table_dict.get('table_name')
            
            # 수신 테이블 정보 읽기 (row=4, column=start_col+1)
            recv_table_cell = worksheet.cell(row=4, column=start_col + 1)
            recv_table_dict = self._parse_cell_dict(recv_table_cell.value)
            if recv_table_dict:
                interface_info['recv']['owner'] = recv_table_dict.get('owner')
                interface_info['recv']['table_name'] = recv_table_dict.get('table_name')
            
        except Exception as e:
            print(f"Warning: 테이블 정보 읽기 실패 (컬럼 {start_col}): {str(e)}")
        
        # 컬럼 매핑 정보 읽기 (실패해도 계속)
        try:
            send_columns, recv_columns = self._read_column_mappings(worksheet, start_col, 5)
            interface_info['send']['columns'] = send_columns
            interface_info['recv']['columns'] = recv_columns
            
        except Exception as e:
            print(f"Warning: 컬럼 매핑 읽기 실패 (컬럼 {start_col}): {str(e)}")
            # 컬럼 매핑 읽기 실패해도 빈 리스트로 계속 진행
        
        # 3단계: ProcessFileMapper로 .process 파일 정보 추가
        print(f"\n=== ProcessFileMapper 처리 시작 ===")
        print(f"process_mapper 상태: {self.process_mapper is not None}")
        print(f"일련번호: '{interface_info['serial_number']}'")
        print(f"일련번호 존재 여부: {bool(interface_info['serial_number'])}")
        
        if self.process_mapper and interface_info['serial_number']:
            try:
                print(f"ProcessFileMapper에서 일련번호 {interface_info['serial_number']} 검색 중...")
                process_files = self.process_mapper.get_process_files_by_serial(interface_info['serial_number'])
                print(f"검색 결과: {process_files}")
                
                if process_files:
                    interface_info.update(process_files)
                    print(f"Info: 일련번호 {interface_info['serial_number']}의 process 파일 정보 추가됨")
                    print(f"추가된 정보: {process_files}")
                else:
                    print(f"Info: 일련번호 {interface_info['serial_number']}에 해당하는 process 파일 정보 없음")
                    
            except Exception as e:
                print(f"Warning: Process 파일 정보 가져오기 실패: {str(e)}")
        elif not self.process_mapper:
            print("Warning: ProcessFileMapper가 초기화되지 않음")
        elif not interface_info['serial_number']:
            print("Warning: 일련번호가 없어서 ProcessFileMapper 처리 건너뜀")
        
        print(f"=== ProcessFileMapper 처리 완료 ===\n")
        
        return interface_info
    
    def _parse_cell_dict(self, cell_value: Any) -> Dict[str, Any]:
        """
        셀 값을 딕셔너리로 안전하게 파싱
        
        Args:
            cell_value: 엑셀 셀의 값
            
        Returns:
            Dict[str, Any]: 파싱된 딕셔너리, 실패시 빈 딕셔너리
        """
        if cell_value is None:
            return {}
        
        try:
            # 문자열을 딕셔너리로 안전하게 변환
            if isinstance(cell_value, str) and cell_value.strip():
                return ast.literal_eval(cell_value.strip())
            else:
                return {}
        except (SyntaxError, ValueError, TypeError):
            # 파싱 실패시 빈 딕셔너리 반환
            return {}
    
    def _read_column_mappings(self, worksheet: Worksheet, start_col: int, start_row: int) -> tuple[List[str], List[str]]:
        """
        컬럼 매핑 정보를 읽어 송신/수신 컬럼 리스트로 반환
        
        Args:
            worksheet: 엑셀 워크시트 객체
            start_col (int): 시작 컬럼 번호
            start_row (int): 시작 행 번호
            
        Returns:
            tuple[List[str], List[str]]: (송신 컬럼 리스트, 수신 컬럼 리스트)
        """
        send_columns = []
        recv_columns = []
        
        current_row = start_row
        
        # 빈 행이 나올 때까지 계속 읽기
        while current_row <= worksheet.max_row:
            send_cell = worksheet.cell(row=current_row, column=start_col)
            recv_cell = worksheet.cell(row=current_row, column=start_col + 1)
            
            send_value = send_cell.value
            recv_value = recv_cell.value
            
            # 둘 다 비어있으면 종료
            if not send_value and not recv_value:
                break
            
            # 값이 있으면 문자열로 변환하고 앞뒤 공백 제거하여 추가
            send_columns.append(str(send_value).strip() if send_value else '')
            recv_columns.append(str(recv_value).strip() if recv_value else '')
            
            current_row += 1
        
        return send_columns, recv_columns
    
    def get_statistics(self) -> Dict[str, int]:
        """
        마지막 처리 결과의 통계 정보 반환
        
        Returns:
            Dict[str, int]: 처리 통계 정보
        """
        return {
            'processed_count': self.processed_count,
            'error_count': self.error_count,
            'total_attempts': self.processed_count + self.error_count
        }
    
    def get_last_errors(self) -> List[str]:
        """
        마지막 처리에서 발생한 오류 메시지들 반환
        
        Returns:
            List[str]: 오류 메시지 리스트
        """
        return self.last_error_messages.copy()

    def compare_column_mappings(self, interface_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        인터페이스 정보의 송신/수신 컬럼과 .process 파일의 컬럼 매핑을 비교
        
        Args:
            interface_data (Dict[str, Any]): 인터페이스 정보 딕셔너리
            
        Returns:
            Dict[str, Any]: 비교 결과
            {
                'send_comparison': {...},   # 송신 비교 결과 (엑셀 송신 vs .process SELECT)
                'recv_comparison': {...},   # 수신 비교 결과 (엑셀 수신 vs .process INSERT)
                'send_recv_comparison': {...}  # 송신과 수신 간 연결 비교
            }
        """
        print(f"\n=== 컬럼 매핑 비교 시작: {interface_data.get('interface_name', 'Unknown')} ===")
        
        comparison_result = {
            'send_comparison': {},
            'recv_comparison': {},
            'send_recv_comparison': {}
        }
        
        # 송신 파일 비교 (엑셀 송신 컬럼 vs .process SELECT 컬럼)
        if interface_data.get('send_copy'):
            print(f"\n--- 송신 파일 비교: {interface_data['send_copy']} ---")
            send_comparison = self._compare_send_mapping(
                interface_data['send']['columns'],
                interface_data['send_copy'],
                '송신'
            )
            comparison_result['send_comparison'] = send_comparison
        else:
            print("\n--- 송신 파일 경로 없음 ---")
        
        # 수신 파일 비교 (엑셀 수신 컬럼 vs .process INSERT 컬럼)
        if interface_data.get('recv_copy'):
            print(f"\n--- 수신 파일 비교: {interface_data['recv_copy']} ---")
            recv_comparison = self._compare_single_mapping(
                interface_data['recv']['columns'],
                interface_data['recv_copy'],
                '수신'
            )
            comparison_result['recv_comparison'] = recv_comparison
        else:
            print("\n--- 수신 파일 경로 없음 ---")
        
        # 송신-수신 연결 비교 (엑셀 송신-수신 매핑 쌍 vs .process 송신-수신 매핑 쌍)
        print(f"\n--- 송신-수신 연결 비교 ---")
        send_recv_comparison = self._compare_send_recv_connection(
            interface_data['send']['columns'],
            interface_data['recv']['columns'],
            interface_data.get('recv_copy', '')
        )
        comparison_result['send_recv_comparison'] = send_recv_comparison
        
        print(f"\n=== 컬럼 매핑 비교 완료 ===")
        return comparison_result
    
    def _compare_send_mapping(self, excel_send_columns: List[str], send_process_file_path: str, direction: str) -> Dict[str, Any]:
        """
        송신 컬럼 매핑 비교 (엑셀 송신 컬럼 vs .process SELECT 컬럼)
        
        Args:
            excel_send_columns (List[str]): 엑셀에서 읽은 송신 컬럼 리스트
            send_process_file_path (str): 송신 .process 파일 경로
            direction (str): 방향 ('송신')
            
        Returns:
            Dict[str, Any]: 송신 비교 결과
        """
        result = {
            'excel_columns': excel_send_columns,
            'process_select_columns': [],
            'table_info': {},
            'matches': [],
            'excel_only': [],
            'process_only': [],
            'match_count': 0,
            'total_excel': len(excel_send_columns),
            'total_process': 0,
            'match_percentage': 0.0,
            'file_exists': False,
            'error': None
        }
        
        try:
            # 파일 존재 여부 확인
            if not os.path.exists(send_process_file_path):
                result['error'] = f"송신 .process 파일 없음: {send_process_file_path}"
                print(f"Warning: {result['error']}")
                return result
            
            result['file_exists'] = True
            
            # BWProcessFileParser로 송신 컬럼 추출
            bw_parser = BWProcessFileParser()
            send_column_mappings = bw_parser.extract_send_column_mappings(send_process_file_path)
            
            process_send_columns = send_column_mappings.get('send_columns', [])
            table_info = {
                'table_name': send_column_mappings.get('table_name', ''),
                'where_condition': send_column_mappings.get('where_condition', ''),
                'order_by': send_column_mappings.get('order_by', '')
            }
            
            result['process_select_columns'] = process_send_columns
            result['table_info'] = table_info
            result['total_process'] = len(process_send_columns)
            
            print(f"\n=== {direction} 컬럼 비교 상세 ===")
            print(f"엑셀 송신 컬럼 ({len(excel_send_columns)}개): {excel_send_columns}")
            print(f"Process SELECT 컬럼 ({len(process_send_columns)}개): {process_send_columns}")
            print(f"테이블: {table_info.get('table_name', 'Unknown')}")
            
            # 대소문자 구분 없이 비교를 위한 매핑 생성
            excel_lower = [col.strip().lower() for col in excel_send_columns if col and col.strip()]
            process_lower = [col.strip().lower() for col in process_send_columns if col and col.strip()]
            
            # 매칭 찾기
            matches = []
            excel_only = []
            process_only = []
            
            # 엑셀 송신 컬럼 기준으로 매칭 찾기
            for excel_col in excel_send_columns:
                if not excel_col or not excel_col.strip():  # 빈 컬럼 제외
                    continue
                    
                excel_col_lower = excel_col.strip().lower()
                if excel_col_lower in process_lower:
                    # 매칭된 인덱스 찾기
                    process_idx = process_lower.index(excel_col_lower)
                    process_col = process_send_columns[process_idx]
                    
                    match_info = {
                        'excel_column': excel_col.strip(),
                        'process_column': process_col,
                        'match_type': 'direct'
                    }
                    matches.append(match_info)
                else:
                    excel_only.append(excel_col.strip())
            
            # Process SELECT에만 있는 컬럼 찾기
            for process_col in process_send_columns:
                if not process_col or not process_col.strip():  # 빈 컬럼 제외
                    continue
                    
                process_col_lower = process_col.strip().lower()
                if process_col_lower not in excel_lower:
                    process_only.append(process_col)
            
            result['matches'] = matches
            result['excel_only'] = excel_only
            result['process_only'] = process_only
            result['match_count'] = len(matches)
            
            # 매칭 비율 계산
            if result['total_excel'] > 0:
                result['match_percentage'] = (result['match_count'] / result['total_excel']) * 100
            
            # 결과 출력
            print(f"\n🔍 {direction} 매칭 결과:")
            print(f"✅ 매칭됨 ({len(matches)}개):")
            for match in matches:
                print(f"  - {match['excel_column']} = {match['process_column']}")
            
            print(f"\n❌ 엑셀에만 있음 ({len(excel_only)}개):")
            for col in excel_only:
                print(f"  - {col}")
            
            print(f"\n⚠️ Process SELECT에만 있음 ({len(process_only)}개):")
            for col in process_only:
                print(f"  - {col}")
            
            print(f"\n📊 매칭률: {result['match_percentage']:.1f}% ({result['match_count']}/{result['total_excel']})")
            
        except Exception as e:
            result['error'] = f"송신 비교 중 오류: {str(e)}"
            print(f"Error: {result['error']}")
            import traceback
            traceback.print_exc()
        
        return result
    
    def _compare_send_recv_connection(self, excel_send_columns: List[str], excel_recv_columns: List[str], recv_process_file_path: str) -> Dict[str, Any]:
        """
        엑셀에서 추출한 송신-수신 매핑 쌍과 수신 .process 파일의 송신-수신 매핑 쌍 비교
        엑셀의 순서대로 매핑된 송신-수신 쌍이 process에서 추출한 매핑 쌍에 포함되는지 확인
        
        Args:
            excel_send_columns (List[str]): 엑셀 송신 컬럼 리스트
            excel_recv_columns (List[str]): 엑셀 수신 컬럼 리스트
            recv_process_file_path (str): 수신 .process 파일 경로
            
        Returns:
            Dict[str, Any]: 송신-수신 연결 비교 결과
        """
        result = {
            'excel_send_recv_pairs': [],
            'process_send_recv_pairs': [],
            'matches': [],
            'excel_only': [],
            'process_only': [],
            'match_count': 0,
            'total_excel_pairs': 0,
            'total_process_pairs': 0,
            'match_percentage': 0.0,
            'recv_file_exists': False,
            'error': None
        }
        
        try:
            print(f"\n=== 송신-수신 연결 비교 (엑셀 vs Process 매핑 쌍) ===")
            
            # 1단계: 엑셀에서 송신-수신 매핑 쌍 생성 (순서대로 매핑)
            excel_send_recv_pairs = []
            min_length = min(len(excel_send_columns), len(excel_recv_columns))
            
            for i in range(min_length):
                send_col = excel_send_columns[i] if excel_send_columns[i] else ''
                recv_col = excel_recv_columns[i] if excel_recv_columns[i] else ''
                
                if send_col.strip() and recv_col.strip():
                    # 대소문자 무시하고 공백 제거한 쌍 생성
                    pair = (send_col.lower().strip(), recv_col.lower().strip())
                    excel_send_recv_pairs.append(pair)
            
            result['excel_send_recv_pairs'] = excel_send_recv_pairs
            result['total_excel_pairs'] = len(excel_send_recv_pairs)
            
            print(f"엑셀 송신-수신 매핑 쌍 ({len(excel_send_recv_pairs)}개):")
            for i, (send, recv) in enumerate(excel_send_recv_pairs, 1):
                print(f"  {i}. ({send}, {recv})")
            
            # 2단계: 수신 파일이 있는 경우에만 Process 매핑 쌍 추출
            if recv_process_file_path and os.path.exists(recv_process_file_path):
                result['recv_file_exists'] = True
                
                # BWProcessFileParser로 수신 파일에서 송신-수신 매핑 정보 추출
                bw_parser = BWProcessFileParser()
                recv_column_mappings = bw_parser.extract_column_mappings(recv_process_file_path)
                recv_detailed_mappings = recv_column_mappings.get('column_mappings', [])
                
                # Process에서 송신-수신 매핑 쌍 생성
                process_send_recv_pairs = []
                for mapping in recv_detailed_mappings:
                    send_col = mapping.get('send', '')
                    recv_col = mapping.get('recv', '')
                    
                    # 실제 송신 컬럼인지 확인 (literal, pattern 등 제외)
                    if (send_col and recv_col and 
                        not send_col.startswith("'") and 
                        not send_col.startswith('pattern_') and 
                        not send_col.startswith('conditional_') and
                        not send_col.startswith('unknown_')):
                        
                        # 대소문자 무시하고 공백 제거한 쌍 생성
                        pair = (send_col.lower().strip(), recv_col.lower().strip())
                        if pair not in process_send_recv_pairs:
                            process_send_recv_pairs.append(pair)
                
                result['process_send_recv_pairs'] = process_send_recv_pairs
                result['total_process_pairs'] = len(process_send_recv_pairs)
                
                print(f"Process 송신-수신 매핑 쌍 ({len(process_send_recv_pairs)}개):")
                for i, (send, recv) in enumerate(process_send_recv_pairs, 1):
                    print(f"  {i}. ({send}, {recv})")
                
                # 3단계: 매칭 비교 수행
                matches = []
                excel_only = []
                process_only = []
                
                # 엑셀 쌍이 process 쌍에 포함되는지 확인
                for excel_pair in excel_send_recv_pairs:
                    if excel_pair in process_send_recv_pairs:
                        matches.append({
                            'excel_pair': excel_pair,
                            'process_pair': excel_pair,
                            'match_type': 'exact'
                        })
                    else:
                        excel_only.append(excel_pair)
                
                # process에만 있는 쌍 찾기
                for process_pair in process_send_recv_pairs:
                    if process_pair not in excel_send_recv_pairs:
                        process_only.append(process_pair)
                
                # 결과 업데이트
                result['matches'] = matches
                result['excel_only'] = excel_only
                result['process_only'] = process_only
                result['match_count'] = len(matches)
                
                # 매칭률 계산 (엑셀 기준)
                if result['total_excel_pairs'] > 0:
                    result['match_percentage'] = (
                        result['match_count'] / result['total_excel_pairs']
                    ) * 100
                
                # 결과 출력
                print(f"\n🔗 송신-수신 매핑 쌍 비교 결과:")
                print(f"✅ 매칭됨 ({len(matches)}개):")
                for match in matches:
                    pair = match['excel_pair']
                    print(f"  - ({pair[0]}, {pair[1]})")
                
                print(f"\n❌ 엑셀에만 있음 ({len(excel_only)}개):")
                for pair in excel_only:
                    print(f"  - ({pair[0]}, {pair[1]})")
                
                print(f"\n⚠️ Process에만 있음 ({len(process_only)}개):")
                for pair in process_only:
                    print(f"  - ({pair[0]}, {pair[1]})")
                
                print(f"\n📊 매핑 쌍 매칭률: {result['match_percentage']:.1f}% ({result['match_count']}/{result['total_excel_pairs']})")
            
            else:
                # 수신 파일이 없어도 엑셀 쌍은 표시
                print(f"\n⚠️ 수신 .process 파일이 없어서 Process 매핑 쌍 추출 불가")
                print(f"수신 파일: {recv_process_file_path}")
                
                # 엑셀 쌍만 있는 상태로 결과 설정
                result['excel_only'] = excel_send_recv_pairs
                
                if result['total_excel_pairs'] > 0:
                    result['match_percentage'] = 0.0  # Process 정보가 없으므로 0%
                    print(f"\n📊 매핑 쌍 매칭률: 0.0% (수신 파일 없음)")
            
        except Exception as e:
            result['error'] = f"송신-수신 연결 비교 중 오류: {str(e)}"
            print(f"Error: {result['error']}")
            import traceback
            traceback.print_exc()
        
        return result
    
    def _compare_single_mapping(self, excel_columns: List[str], process_file_path: str, direction: str) -> Dict[str, Any]:
        """
        수신 컬럼 매핑 비교 (엑셀 수신 컬럼 vs .process INSERT 컬럼)
        
        Args:
            excel_columns (List[str]): 엑셀에서 읽은 수신 컬럼 리스트
            process_file_path (str): 수신 .process 파일 경로
            direction (str): 방향 ('수신')
            
        Returns:
            Dict[str, Any]: 수신 비교 결과
        """
        result = {
            'excel_columns': excel_columns,
            'process_recv_columns': [],
            'process_send_columns': [],
            'detailed_mappings': [],
            'matches': [],
            'excel_only': [],
            'process_only': [],
            'match_count': 0,
            'total_excel': len(excel_columns),
            'total_process': 0,
            'match_percentage': 0.0,
            'file_exists': False,
            'error': None
        }
        
        try:
            # 파일 존재 여부 확인
            if not os.path.exists(process_file_path):
                result['error'] = f"수신 .process 파일 없음: {process_file_path}"
                print(f"Warning: {result['error']}")
                return result
            
            result['file_exists'] = True
            
            # BWProcessFileParser로 컬럼 매핑 추출
            bw_parser = BWProcessFileParser()
            column_mappings = bw_parser.extract_column_mappings(process_file_path)
            
            recv_columns = column_mappings.get('recv_columns', [])
            send_columns = column_mappings.get('send_columns', [])
            detailed_mappings = column_mappings.get('column_mappings', [])
            
            result['process_recv_columns'] = recv_columns
            result['process_send_columns'] = send_columns
            result['detailed_mappings'] = detailed_mappings
            result['total_process'] = len(recv_columns)
            
            print(f"\n=== {direction} 컬럼 비교 상세 ===")
            print(f"엑셀 수신 컬럼 ({len(excel_columns)}개): {excel_columns}")
            print(f"Process 수신 컬럼 ({len(recv_columns)}개): {recv_columns}")
            print(f"Process 송신 컬럼 ({len(send_columns)}개): {send_columns}")
            
            # 수신 비교: 엑셀 수신 컬럼 vs Process 수신 컬럼
            process_compare_columns = recv_columns
            
            # 대소문자 구분 없이 비교를 위한 매핑 생성
            excel_lower = [col.strip().lower() for col in excel_columns if col and col.strip()]
            process_lower = [col.strip().lower() for col in process_compare_columns if col and col.strip()]
            
            # 매칭 찾기
            matches = []
            excel_only = []
            process_only = []
            
            # 엑셀 수신 컬럼 기준으로 매칭 찾기
            for excel_col in excel_columns:
                if not excel_col or not excel_col.strip():  # 빈 컬럼 제외
                    continue
                    
                excel_col_lower = excel_col.strip().lower()
                if excel_col_lower in process_lower:
                    # 매칭된 인덱스 찾기
                    process_idx = process_lower.index(excel_col_lower)
                    process_col = process_compare_columns[process_idx]
                    
                    # 상세 매핑 정보 찾기
                    detailed_info = None
                    for mapping in detailed_mappings:
                        if mapping['recv'].strip().lower() == excel_col_lower:
                            detailed_info = mapping
                            break
                    
                    match_info = {
                        'excel_column': excel_col.strip(),
                        'process_column': process_col,
                        'value_type': detailed_info['value_type'] if detailed_info else 'unknown',
                        'value_pattern': detailed_info.get('value_pattern', '') if detailed_info else ''
                    }
                    
                    if detailed_info:
                        match_info['mapped_send_column'] = detailed_info['send']
                    
                    matches.append(match_info)
                else:
                    excel_only.append(excel_col.strip())
            
            # Process 수신에만 있는 컬럼 찾기
            for process_col in process_compare_columns:
                if not process_col or not process_col.strip():  # 빈 컬럼 제외
                    continue
                    
                process_col_lower = process_col.strip().lower()
                if process_col_lower not in excel_lower:
                    process_only.append(process_col)
            
            result['matches'] = matches
            result['excel_only'] = excel_only
            result['process_only'] = process_only
            result['match_count'] = len(matches)
            
            # 매칭 비율 계산
            if result['total_excel'] > 0:
                result['match_percentage'] = (result['match_count'] / result['total_excel']) * 100
            
            # 결과 출력
            print(f"\n🔍 {direction} 매칭 결과:")
            print(f"✅ 매칭됨 ({len(matches)}개):")
            for match in matches:
                extra_info = ""
                if 'mapped_send_column' in match:
                    extra_info = f" -> 송신: {match['mapped_send_column']}"
                print(f"  - {match['excel_column']} = {match['process_column']} ({match['value_type']}){extra_info}")
            
            print(f"\n❌ 엑셀에만 있음 ({len(excel_only)}개):")
            for col in excel_only:
                print(f"  - {col}")
            
            print(f"\n⚠️ Process 수신에만 있음 ({len(process_only)}개):")
            for col in process_only:
                print(f"  - {col}")
            
            print(f"\n📊 매칭률: {result['match_percentage']:.1f}% ({result['match_count']}/{result['total_excel']})")
            
        except Exception as e:
            result['error'] = f"수신 비교 중 오류: {str(e)}"
            print(f"Error: {result['error']}")
            import traceback
            traceback.print_exc()
        
        return result

    def export_all_interfaces_to_log(self, interfaces: List[Dict[str, Any]], log_file_path: str = "test_iflist.log") -> None:
        """
        모든 인터페이스 정보를 로그 파일로 출력
        
        Args:
            interfaces (List[Dict[str, Any]]): 인터페이스 정보 리스트
            log_file_path (str): 로그 파일 경로 (기본값: "test_iflist.log")
        """
        try:
            with open(log_file_path, 'w', encoding='utf-8') as log_file:
                # 로그 헤더 작성
                current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                log_file.write(f"=== 인터페이스 정보 분석 결과 로그 ===\n")
                log_file.write(f"생성일시: {current_time}\n")
                log_file.write(f"총 인터페이스 수: {len(interfaces)}개\n")
                log_file.write("=" * 80 + "\n\n")
                
                # 각 인터페이스별로 상세 정보 출력
                for idx, interface in enumerate(interfaces, 1):
                    log_file.write(f"[{idx:03d}] =========================== 인터페이스 정보 ===========================\n")
                    log_file.write(f"인터페이스명: {interface['interface_name']}\n")
                    log_file.write(f"인터페이스ID: {interface['interface_id']}\n")
                    log_file.write(f"일련번호: {interface['serial_number']}\n")
                    log_file.write(f"송신 테이블: {interface['send']['table_name']}\n")
                    log_file.write(f"수신 테이블: {interface['recv']['table_name']}\n")
                    log_file.write(f"송신 컬럼 수: {len(interface['send']['columns'])}\n")
                    log_file.write(f"수신 컬럼 수: {len(interface['recv']['columns'])}\n")
                    log_file.write(f"송신 원본파일: {interface.get('send_original', 'N/A')}\n")
                    log_file.write(f"송신 복사파일: {interface.get('send_copy', 'N/A')}\n")
                    log_file.write(f"수신 원본파일: {interface.get('recv_original', 'N/A')}\n")
                    log_file.write(f"수신 복사파일: {interface.get('recv_copy', 'N/A')}\n")
                    log_file.write(f"송신 스키마파일: {interface.get('send_schema', 'N/A')}\n")
                    log_file.write(f"수신 스키마파일: {interface.get('recv_schema', 'N/A')}\n")
                    
                    # 송신 컬럼 상세 정보
                    log_file.write(f"\n--- 송신 컬럼 목록 ({len(interface['send']['columns'])}개) ---\n")
                    for i, col in enumerate(interface['send']['columns'], 1):
                        log_file.write(f"  {i:2d}. {col}\n")
                    
                    # 수신 컬럼 상세 정보
                    log_file.write(f"\n--- 수신 컬럼 목록 ({len(interface['recv']['columns'])}개) ---\n")
                    for i, col in enumerate(interface['recv']['columns'], 1):
                        log_file.write(f"  {i:2d}. {col}\n")
                    
                    # 컬럼 매핑 비교 수행
                    log_file.write(f"\n--- 컬럼 매핑 비교 결과 ---\n")
                    try:
                        comparison_result = self.compare_column_mappings(interface)
                        
                        # 송신 비교 결과
                        send_comp = comparison_result['send_comparison']
                        log_file.write(f"📤 송신 파일 비교: {interface.get('send_copy', 'N/A')}\n")
                        if send_comp.get('file_exists'):
                            log_file.write(f"   매칭률: {send_comp['match_percentage']:.1f}% ({send_comp['match_count']}/{send_comp['total_excel']})\n")
                            log_file.write(f"   테이블: {send_comp.get('table_info', {}).get('table_name', 'Unknown')}\n")
                            log_file.write(f"   WHERE: {send_comp.get('table_info', {}).get('where_condition', 'None')}\n")
                            
                            if send_comp['matches']:
                                log_file.write(f"   ✅ 매칭된 컬럼 ({len(send_comp['matches'])}개):\n")
                                for match in send_comp['matches']:
                                    log_file.write(f"      - {match['excel_column']} = {match['process_column']}\n")
                            
                            if send_comp['excel_only']:
                                log_file.write(f"   ❌ 엑셀에만 있는 컬럼 ({len(send_comp['excel_only'])}개):\n")
                                for col in send_comp['excel_only']:
                                    log_file.write(f"      - {col}\n")
                            
                            if send_comp['process_only']:
                                log_file.write(f"   ⚠️ Process SELECT에만 있는 컬럼 ({len(send_comp['process_only'])}개):\n")
                                for col in send_comp['process_only']:
                                    log_file.write(f"      - {col}\n")
                        else:
                            log_file.write(f"   오류: {send_comp.get('error', '파일 없음')}\n")
                        
                        # 수신 비교 결과
                        recv_comp = comparison_result['recv_comparison']
                        log_file.write(f"\n📥 수신 파일 비교: {interface.get('recv_copy', 'N/A')}\n")
                        if recv_comp.get('file_exists'):
                            log_file.write(f"   매칭률: {recv_comp['match_percentage']:.1f}% ({recv_comp['match_count']}/{recv_comp['total_excel']})\n")
                            
                            if recv_comp['matches']:
                                log_file.write(f"   ✅ 매칭된 컬럼 ({len(recv_comp['matches'])}개):\n")
                                for match in recv_comp['matches']:
                                    extra_info = ""
                                    if 'mapped_send_column' in match:
                                        extra_info = f" -> 송신: {match['mapped_send_column']}"
                                    log_file.write(f"      - {match['excel_column']} = {match['process_column']} ({match['value_type']}){extra_info}\n")
                            
                            if recv_comp['excel_only']:
                                log_file.write(f"   ❌ 엑셀에만 있는 컬럼 ({len(recv_comp['excel_only'])}개):\n")
                                for col in recv_comp['excel_only']:
                                    log_file.write(f"      - {col}\n")
                            
                            if recv_comp['process_only']:
                                log_file.write(f"   ⚠️ Process 수신에만 있는 컬럼 ({len(recv_comp['process_only'])}개):\n")
                                for col in recv_comp['process_only']:
                                    log_file.write(f"      - {col}\n")
                        else:
                            log_file.write(f"   오류: {recv_comp.get('error', '파일 없음')}\n")
                        
                        # 송신-수신 연결 비교 결과
                        conn_comp = comparison_result['send_recv_comparison']
                        log_file.write(f"\n🔗 송신-수신 연결 비교\n")
                        if conn_comp.get('recv_file_exists'):
                            log_file.write(f"   매핑 쌍 매칭률: {conn_comp['match_percentage']:.1f}% ({conn_comp['match_count']}/{conn_comp['total_excel_pairs']})\n")
                            log_file.write(f"   엑셀 매핑 쌍 수: {conn_comp['total_excel_pairs']}개\n")
                            log_file.write(f"   Process 매핑 쌍 수: {conn_comp['total_process_pairs']}개\n")
                            
                            if conn_comp['matches']:
                                log_file.write(f"   ✅ 매칭된 쌍 ({len(conn_comp['matches'])}개):\n")
                                for match in conn_comp['matches']:
                                    pair = match['excel_pair']
                                    log_file.write(f"      - ({pair[0]}, {pair[1]})\n")
                            
                            if conn_comp['excel_only']:
                                log_file.write(f"   ❌ 엑셀에만 있는 쌍 ({len(conn_comp['excel_only'])}개):\n")
                                for pair in conn_comp['excel_only']:
                                    log_file.write(f"      - ({pair[0]}, {pair[1]})\n")
                            
                            if conn_comp['process_only']:
                                log_file.write(f"   ⚠️ Process에만 있는 쌍 ({len(conn_comp['process_only'])}개):\n")
                                for pair in conn_comp['process_only']:
                                    log_file.write(f"      - ({pair[0]}, {pair[1]})\n")
                        else:
                            log_file.write(f"   오류: {conn_comp.get('error', '파일 없음')}\n")
                        
                        # 스키마 비교 결과 추가
                        log_file.write(f"\n--- 스키마 파일 비교 결과 ---\n")
                        try:
                            schema_comparison_result = self.compare_schema_mappings(interface)
                            
                            # 송신 스키마 비교 결과
                            send_schema_comp = schema_comparison_result['send_schema_comparison']
                            log_file.write(f"📋 송신 스키마 비교: {interface.get('send_schema', 'N/A')}\n")
                            if send_schema_comp.get('file_exists'):
                                log_file.write(f"   매칭률: {send_schema_comp['match_percentage']:.1f}% ({send_schema_comp['match_count']}/{send_schema_comp['total_process']})\n")
                                
                                if send_schema_comp['matches']:
                                    log_file.write(f"   ✅ 매칭된 컬럼 ({len(send_schema_comp['matches'])}개):\n")
                                    for match in send_schema_comp['matches']:
                                        log_file.write(f"      - {match['schema_column']} = {match['process_column']}\n")
                                
                                if send_schema_comp['schema_only']:
                                    log_file.write(f"   ❌ 송신 스키마에만 있는 컬럼 ({len(send_schema_comp['schema_only'])}개):\n")
                                    for col in send_schema_comp['schema_only']:
                                        log_file.write(f"      - {col}\n")
                                
                                if send_schema_comp['process_only']:
                                    log_file.write(f"   ⚠️ Process 송신에만 있는 컬럼 ({len(send_schema_comp['process_only'])}개):\n")
                                    for col in send_schema_comp['process_only']:
                                        log_file.write(f"      - {col}\n")
                            else:
                                log_file.write(f"   오류: {send_schema_comp.get('error', '파일 없음')}\n")
                            
                            # 수신 스키마 비교 결과
                            recv_schema_comp = schema_comparison_result['recv_schema_comparison']
                            log_file.write(f"\n📋 수신 스키마 비교: {interface.get('recv_schema', 'N/A')}\n")
                            if recv_schema_comp.get('file_exists'):
                                log_file.write(f"   매칭률: {recv_schema_comp['match_percentage']:.1f}% ({recv_schema_comp['match_count']}/{recv_schema_comp['total_process']})\n")
                                
                                if recv_schema_comp['matches']:
                                    log_file.write(f"   ✅ 매칭된 컬럼 ({len(recv_schema_comp['matches'])}개):\n")
                                    for match in recv_schema_comp['matches']:
                                        log_file.write(f"      - {match['schema_column']} = {match['process_column']}\n")
                                
                                if recv_schema_comp['schema_only']:
                                    log_file.write(f"   ❌ 수신 스키마에만 있는 컬럼 ({len(recv_schema_comp['schema_only'])}개):\n")
                                    for col in recv_schema_comp['schema_only']:
                                        log_file.write(f"      - {col}\n")
                                
                                if recv_schema_comp['process_only']:
                                    log_file.write(f"   ⚠️ Process 송신에만 있는 컬럼 ({len(recv_schema_comp['process_only'])}개):\n")
                                    for col in recv_schema_comp['process_only']:
                                        log_file.write(f"      - {col}\n")
                            else:
                                log_file.write(f"   오류: {recv_schema_comp.get('error', '파일 없음')}\n")
                        
                        except Exception as e:
                            log_file.write(f"   스키마 비교 중 오류: {str(e)}\n")
                    
                    except Exception as e:
                        log_file.write(f"   컬럼 매핑 비교 중 오류: {str(e)}\n")
                    
                    log_file.write("\n" + "=" * 80 + "\n\n")
                
                # 로그 푸터 작성
                log_file.write(f"=== 로그 작성 완료 ({current_time}) ===\n")
                
            print(f"✅ 전체 인터페이스 정보가 '{log_file_path}' 파일로 출력되었습니다.")
            print(f"   총 {len(interfaces)}개 인터페이스 처리 완료")
            
        except Exception as e:
            print(f"❌ 로그 파일 작성 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()

    def compare_schema_mappings(self, interface_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        스키마 파일의 컬럼과 .process 파일에서 추출한 송신 컬럼을 비교
        
        Args:
            interface_data (Dict[str, Any]): 인터페이스 정보 딕셔너리
            
        Returns:
            Dict[str, Any]: 스키마 비교 결과
            {
                'send_schema_comparison': {...},  # 송신 스키마 vs 수신 process의 송신 컬럼
                'recv_schema_comparison': {...}   # 수신 스키마 vs 수신 process의 송신 컬럼
            }
        """
        print(f"\n=== 스키마 매핑 비교 시작: {interface_data.get('interface_name', 'Unknown')} ===")
        
        comparison_result = {
            'send_schema_comparison': {},
            'recv_schema_comparison': {}
        }
        
        # BWProcessFileParser 인스턴스 생성
        bw_parser = BWProcessFileParser()
        
        # 수신 .process 파일에서 송신 컬럼 정보 추출 (비교 기준)
        recv_process_send_columns = []
        if interface_data.get('recv_copy'):
            try:
                print(f"\n--- 수신 .process에서 송신 컬럼 정보 추출 ---")
                column_mappings = bw_parser.extract_column_mappings(interface_data['recv_copy'])
                detailed_mappings = column_mappings.get('column_mappings', [])
                
                for mapping in detailed_mappings:
                    send_col = mapping.get('send', '')
                    # 실제 송신 컬럼인지 확인 (literal, pattern 등 제외)
                    if (send_col and not send_col.startswith("'") and 
                        not send_col.startswith('pattern_') and 
                        not send_col.startswith('conditional_') and
                        not send_col.startswith('unknown_')):
                        if send_col not in recv_process_send_columns:
                            recv_process_send_columns.append(send_col)
                
                print(f"수신 .process에서 추출한 송신 컬럼 ({len(recv_process_send_columns)}개): {recv_process_send_columns}")
                
            except Exception as e:
                print(f"Warning: 수신 .process에서 송신 컬럼 추출 실패: {str(e)}")
        
        # 송신 스키마 파일 비교
        if interface_data.get('send_schema'):
            print(f"\n--- 송신 스키마 파일 비교: {interface_data['send_schema']} ---")
            send_schema_comparison = self._compare_schema_with_columns(
                interface_data['send_schema'],
                recv_process_send_columns,
                '송신 스키마',
                '수신 process의 송신 컬럼'
            )
            comparison_result['send_schema_comparison'] = send_schema_comparison
        else:
            print("\n--- 송신 스키마 파일 경로 없음 ---")
        
        # 수신 스키마 파일 비교
        if interface_data.get('recv_schema'):
            print(f"\n--- 수신 스키마 파일 비교: {interface_data['recv_schema']} ---")
            recv_schema_comparison = self._compare_schema_with_columns(
                interface_data['recv_schema'],
                recv_process_send_columns,
                '수신 스키마',
                '수신 process의 송신 컬럼'
            )
            comparison_result['recv_schema_comparison'] = recv_schema_comparison
        else:
            print("\n--- 수신 스키마 파일 경로 없음 ---")
        
        print(f"\n=== 스키마 매핑 비교 완료 ===")
        return comparison_result
    
    def _compare_schema_with_columns(self, schema_file_path: str, process_send_columns: List[str], 
                                   schema_type: str, column_type: str) -> Dict[str, Any]:
        """
        스키마 파일과 송신 컬럼들을 비교
        
        Args:
            schema_file_path (str): 스키마 파일 경로
            process_send_columns (List[str]): 비교할 송신 컬럼 리스트
            schema_type (str): 스키마 타입 ('송신 스키마' 또는 '수신 스키마')
            column_type (str): 컬럼 타입 설명
            
        Returns:
            Dict[str, Any]: 스키마 비교 결과
        """
        result = {
            'schema_columns': [],
            'process_send_columns': process_send_columns,
            'matches': [],
            'schema_only': [],
            'process_only': [],
            'match_count': 0,
            'total_schema': 0,
            'total_process': len(process_send_columns),
            'match_percentage': 0.0,
            'file_exists': False,
            'error': None
        }
        
        try:
            # BWProcessFileParser로 스키마 컬럼 추출
            bw_parser = BWProcessFileParser()
            schema_result = bw_parser.extract_schema_columns(schema_file_path)
            
            if schema_result.get('error'):
                result['error'] = schema_result['error']
                print(f"Warning: {result['error']}")
                return result
            
            result['file_exists'] = schema_result.get('file_exists', False)
            schema_columns = schema_result.get('schema_columns', [])
            result['schema_columns'] = schema_columns
            result['total_schema'] = len(schema_columns)
            
            print(f"\n=== {schema_type} vs {column_type} 비교 상세 ===")
            print(f"{schema_type} 컬럼 ({len(schema_columns)}개): {schema_columns}")
            print(f"{column_type} ({len(process_send_columns)}개): {process_send_columns}")
            
            # 대소문자 구분 없이 비교를 위한 매핑 생성
            schema_lower = [col.strip().lower() for col in schema_columns if col and col.strip()]
            process_lower = [col.strip().lower() for col in process_send_columns if col and col.strip()]
            
            # 매칭 찾기
            matches = []
            schema_only = []
            process_only = []
            
            # 스키마 컬럼 기준으로 매칭 찾기
            for schema_col in schema_columns:
                if not schema_col or not schema_col.strip():
                    continue
                    
                schema_col_lower = schema_col.strip().lower()
                if schema_col_lower in process_lower:
                    # 매칭된 인덱스 찾기
                    process_idx = process_lower.index(schema_col_lower)
                    process_col = process_send_columns[process_idx]
                    
                    match_info = {
                        'schema_column': schema_col.strip(),
                        'process_column': process_col,
                        'match_type': 'direct'
                    }
                    matches.append(match_info)
                else:
                    schema_only.append(schema_col.strip())
            
            # Process 송신 컬럼에만 있는 컬럼 찾기
            for process_col in process_send_columns:
                if not process_col or not process_col.strip():
                    continue
                    
                process_col_lower = process_col.strip().lower()
                if process_col_lower not in schema_lower:
                    process_only.append(process_col)
            
            result['matches'] = matches
            result['schema_only'] = schema_only
            result['process_only'] = process_only
            result['match_count'] = len(matches)
            
            # 매칭 비율 계산 (process 컬럼 기준)
            if result['total_process'] > 0:
                result['match_percentage'] = (result['match_count'] / result['total_process']) * 100
            
            # 결과 출력
            print(f"\n🔍 {schema_type} 매칭 결과:")
            print(f"✅ 매칭됨 ({len(matches)}개):")
            for match in matches:
                print(f"  - {match['schema_column']} = {match['process_column']}")
            
            print(f"\n❌ {schema_type}에만 있음 ({len(schema_only)}개):")
            for col in schema_only:
                print(f"  - {col}")
            
            print(f"\n⚠️ {column_type}에만 있음 ({len(process_only)}개):")
            for col in process_only:
                print(f"  - {col}")
            
            print(f"\n📊 매칭률: {result['match_percentage']:.1f}% ({result['match_count']}/{result['total_process']})")
            
        except Exception as e:
            result['error'] = f"{schema_type} 비교 중 오류: {str(e)}"
            print(f"Error: {result['error']}")
            import traceback
            traceback.print_exc()
        
        return result

    def export_summary_to_excel(self, interfaces: List[Dict[str, Any]], excel_file_path: str = "test_iflist_result.xlsx") -> None:
        """
        모든 인터페이스의 비교 결과 요약을 엑셀 파일로 출력
        
        Args:
            interfaces (List[Dict[str, Any]]): 인터페이스 정보 리스트
            excel_file_path (str): 출력할 엑셀 파일 경로 (기본값: "test_iflist_result.xlsx")
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "인터페이스 비교 결과"
            
            # 헤더 정의
            headers = [
                "일련번호", "인터페이스명", "인터페이스ID", "송신DB", "수신DB", 
                "송신테이블", "수신테이블",
                "송신비교_매칭률", "송신비교_결과요약",
                "수신비교_매칭률", "수신비교_결과요약", 
                "연결비교_매칭률", "연결비교_결과요약",
                "송신스키마_매칭률", "송신스키마_결과요약",
                "수신스키마_매칭률", "수신스키마_결과요약"
            ]
            
            # 헤더 작성
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # 연한 파란색
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 각 인터페이스별 데이터 작성
            for row_idx, interface in enumerate(interfaces, 2):
                print(f"엑셀 요약 생성 중: {interface.get('interface_name', 'Unknown')} ({row_idx-1}/{len(interfaces)})")
                
                # 기본 정보 추출
                serial_number = interface.get('serial_number', '')
                interface_name = interface.get('interface_name', '')
                interface_id = interface.get('interface_id', '')
                
                # DB 정보 추출
                send_db = interface.get('send', {}).get('db_info', {})
                recv_db = interface.get('recv', {}).get('db_info', {})
                send_db_name = f"{send_db.get('owner', '')}.{send_db.get('database', '')}" if send_db else ''
                recv_db_name = f"{recv_db.get('owner', '')}.{recv_db.get('database', '')}" if recv_db else ''
                
                # 테이블 정보 추출
                send_table = f"{interface.get('send', {}).get('owner', '')}.{interface.get('send', {}).get('table_name', '')}"
                recv_table = f"{interface.get('recv', {}).get('owner', '')}.{interface.get('recv', {}).get('table_name', '')}"
                
                # 기본 정보 셀에 작성
                worksheet.cell(row=row_idx, column=1, value=serial_number)
                worksheet.cell(row=row_idx, column=2, value=interface_name)
                worksheet.cell(row=row_idx, column=3, value=interface_id)
                worksheet.cell(row=row_idx, column=4, value=send_db_name)
                worksheet.cell(row=row_idx, column=5, value=recv_db_name)
                worksheet.cell(row=row_idx, column=6, value=send_table)
                worksheet.cell(row=row_idx, column=7, value=recv_table)
                
                # 비교 결과 수행 및 데이터 추출
                try:
                    comparison_result = self.compare_column_mappings(interface)
                    schema_comparison = self.compare_schema_mappings(interface)
                    
                    # 1. 송신 비교 결과
                    send_comp = comparison_result.get('send_comparison', {})
                    send_match_rate = send_comp.get('match_percentage', 0)
                    send_summary = self._generate_comparison_summary(send_comp, '송신')
                    
                    worksheet.cell(row=row_idx, column=8, value=f"{send_match_rate:.1f}%")
                    worksheet.cell(row=row_idx, column=9, value=send_summary)
                    
                    if send_match_rate < 100:
                        worksheet.cell(row=row_idx, column=8).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                        worksheet.cell(row=row_idx, column=9).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    elif send_summary == "비교 미수행":
                        worksheet.cell(row=row_idx, column=8).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        worksheet.cell(row=row_idx, column=9).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    # 2. 수신 비교 결과
                    recv_comp = comparison_result.get('recv_comparison', {})
                    recv_match_rate = recv_comp.get('match_percentage', 0)
                    recv_summary = self._generate_comparison_summary(recv_comp, '수신')
                    
                    worksheet.cell(row=row_idx, column=10, value=f"{recv_match_rate:.1f}%")
                    worksheet.cell(row=row_idx, column=11, value=recv_summary)
                    
                    if recv_match_rate < 100:
                        worksheet.cell(row=row_idx, column=10).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                        worksheet.cell(row=row_idx, column=11).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    elif recv_summary == "비교 미수행":
                        worksheet.cell(row=row_idx, column=10).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        worksheet.cell(row=row_idx, column=11).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    # 3. 연결 비교 결과
                    conn_comp = comparison_result.get('send_recv_comparison', {})
                    conn_match_rate = conn_comp.get('match_percentage', 0)
                    conn_summary = self._generate_comparison_summary(conn_comp, '연결')
                    
                    worksheet.cell(row=row_idx, column=12, value=f"{conn_match_rate:.1f}%")
                    worksheet.cell(row=row_idx, column=13, value=conn_summary)
                    
                    if conn_match_rate < 100:
                        worksheet.cell(row=row_idx, column=12).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                        worksheet.cell(row=row_idx, column=13).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    elif conn_summary == "비교 미수행":
                        worksheet.cell(row=row_idx, column=12).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        worksheet.cell(row=row_idx, column=13).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    # 4. 송신 스키마 비교 결과
                    send_schema_comp = schema_comparison.get('send_schema_comparison', {})
                    send_schema_match_rate = send_schema_comp.get('match_percentage', 0)
                    send_schema_summary = self._generate_comparison_summary(send_schema_comp, '송신스키마')
                    
                    worksheet.cell(row=row_idx, column=14, value=f"{send_schema_match_rate:.1f}%")
                    worksheet.cell(row=row_idx, column=15, value=send_schema_summary)
                    
                    if send_schema_match_rate < 100:
                        worksheet.cell(row=row_idx, column=14).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                        worksheet.cell(row=row_idx, column=15).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    elif send_schema_summary == "비교 미수행":
                        worksheet.cell(row=row_idx, column=14).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        worksheet.cell(row=row_idx, column=15).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    # 5. 수신 스키마 비교 결과
                    recv_schema_comp = schema_comparison.get('recv_schema_comparison', {})
                    recv_schema_match_rate = recv_schema_comp.get('match_percentage', 0)
                    recv_schema_summary = self._generate_comparison_summary(recv_schema_comp, '수신스키마')
                    
                    worksheet.cell(row=row_idx, column=16, value=f"{recv_schema_match_rate:.1f}%")
                    worksheet.cell(row=row_idx, column=17, value=recv_schema_summary)
                    
                    if recv_schema_match_rate < 100:
                        worksheet.cell(row=row_idx, column=16).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                        worksheet.cell(row=row_idx, column=17).fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    elif recv_schema_summary == "비교 미수행":
                        worksheet.cell(row=row_idx, column=16).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        worksheet.cell(row=row_idx, column=17).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                
                except Exception as e:
                    print(f"Warning: 인터페이스 {interface_name} 비교 중 오류: {str(e)}")
                    # 오류 발생 시 기본값으로 채움
                    for col in range(8, 18):
                        if col % 2 == 0:  # 매칭률 컬럼
                            worksheet.cell(row=row_idx, column=col, value="오류")
                        else:  # 요약 컬럼
                            worksheet.cell(row=row_idx, column=col, value=f"처리 중 오류: {str(e)}")
                        worksheet.cell(row=row_idx, column=col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # 컬럼 너비 자동 조절
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 최대 50으로 제한
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 모든 셀 가운데 정렬
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 파일 저장
            workbook.save(excel_file_path)
            workbook.close()
            
            print(f"✅ 인터페이스 비교 결과 요약이 '{excel_file_path}' 파일로 저장되었습니다.")
            print(f"   총 {len(interfaces)}개 인터페이스 처리 완료")
            
        except ImportError:
            print(f"❌ openpyxl 라이브러리가 필요합니다. 'pip install openpyxl'로 설치해주세요.")
        except Exception as e:
            print(f"❌ 엑셀 요약 파일 생성 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _generate_comparison_summary(self, comparison_result: Dict[str, Any], comparison_type: str) -> str:
        """
        비교 결과에서 간단한 요약 문자열 생성
        
        Args:
            comparison_result (Dict[str, Any]): 비교 결과 딕셔너리
            comparison_type (str): 비교 타입 ('송신', '수신', '연결', '송신스키마', '수신스키마')
            
        Returns:
            str: 요약 문자열
        """
        if not comparison_result:
            return "비교 미수행"
        
        # 연결 비교는 특별 처리 (엑셀 데이터가 있으면 비교 가능)
        if comparison_type == '연결':
            total_excel_pairs = comparison_result.get('total_excel_pairs', 0)
            if total_excel_pairs == 0:
                return "데이터 없음"
            
            match_count = comparison_result.get('match_count', 0)
            recv_file_exists = comparison_result.get('recv_file_exists', False)
            
            if not recv_file_exists:
                return f"수신파일 없음 (엑셀 쌍 {total_excel_pairs}개)"
            
            if match_count == total_excel_pairs:
                return "완전일치"
            else:
                not_matched_count = total_excel_pairs - match_count
                return f"불일치 {not_matched_count}개"
        
        # 다른 비교 타입들의 파일 존재 여부 확인
        file_exists_key = {
            '송신': 'file_exists',
            '수신': 'file_exists',
            '송신스키마': 'file_exists',
            '수신스키마': 'file_exists'
        }.get(comparison_type, 'file_exists')
        
        if not comparison_result.get(file_exists_key, False):
            error = comparison_result.get('error', '파일 없음')
            return f"오류: {error}"
        
        match_count = comparison_result.get('match_count', 0)
        
        if comparison_type in ['송신스키마', '수신스키마']:
            total_count = comparison_result.get('total_process', 0)
            not_matched = comparison_result.get('process_only', [])
        else:
            total_count = comparison_result.get('total_excel', 0)
            not_matched = comparison_result.get('excel_only', [])
        
        if total_count == 0:
            return "데이터 없음"
        
        if match_count == total_count:
            return "완전일치"
        else:
            not_matched_count = len(not_matched) if not_matched else (total_count - match_count)
            return f"불일치 {not_matched_count}개"


class BWProcessFileParser:
    """
    TIBCO BW .process 파일에서 수신용 INSERT 쿼리를 추출하는 클래스
    """
    
    def __init__(self):
        """BWProcessFileParser 초기화"""
        self.ns = {
            'pd': 'http://xmlns.tibco.com/bw/process/2003',
            'xsl': 'http://www.w3.org/1999/XSL/Transform',
            'xsi': 'http://www.w3.org/2001/XMLSchema-instance'
        }
        self.processed_count = 0
        self.error_count = 0
        self.last_error_messages = []
    
    def parse_bw_process_file(self, process_file_path: str) -> List[str]:
        """
        BW .process 파일에서 수신용 INSERT 쿼리를 추출
        
        Args:
            process_file_path (str): .process 파일의 경로
            
        Returns:
            List[str]: 추출된 INSERT 쿼리 목록
            
        Raises:
            FileNotFoundError: 파일이 존재하지 않는 경우
            ValueError: 파일 형식이 올바르지 않은 경우
        """
        # 초기화
        self.processed_count = 0
        self.error_count = 0
        self.last_error_messages = []
        
        # 파일 존재 여부 확인
        if not os.path.exists(process_file_path):
            raise FileNotFoundError(f"BW process 파일을 찾을 수 없습니다: {process_file_path}")
        
        insert_queries = []
        
        try:
            # XML 파일 파싱
            tree = ET.parse(process_file_path)
            root = tree.getroot()
            
            print(f"\n=== BW Process 파일 처리 시작: {process_file_path} ===")
            
            # JDBC 액티비티 찾기
            activities = root.findall('.//pd:activity', self.ns)
            
            for activity in activities:
                try:
                    # JDBC 액티비티 타입 확인
                    activity_type = activity.find('./pd:type', self.ns)
                    if activity_type is None or 'jdbc' not in activity_type.text.lower():
                        continue
                    
                    activity_name = activity.get('name', 'Unknown')
                    print(f"\nJDBC 액티비티 발견: {activity_name}")
                    
                    # statement 추출
                    statement = activity.find('.//config/statement')
                    if statement is not None and statement.text:
                        query = statement.text.strip()
                        print(f"\n발견된 쿼리:\n{query}")
                        
                        # INSERT 쿼리인 경우만 처리
                        if query.lower().startswith('insert'):
                            # Oracle 힌트 제거
                            cleaned_query = self._remove_oracle_hints(query)
                            
                            # 파라미터 매핑 처리
                            mapped_query = self._process_query_parameters(activity, cleaned_query)
                            
                            insert_queries.append(mapped_query)
                            self.processed_count += 1
                            print(f"=> 최종 처리된 INSERT 쿼리:\n{mapped_query}")
                        else:
                            print(f"=> INSERT 쿼리가 아니므로 제외: {query[:50]}...")
                
                except Exception as e:
                    self.error_count += 1
                    error_msg = f"액티비티 '{activity.get('name', 'Unknown')}' 처리 중 오류: {str(e)}"
                    self.last_error_messages.append(error_msg)
                    print(f"Warning: {error_msg}")
            
            print(f"\n=== 처리된 INSERT 쿼리 수: {len(insert_queries)} ===")
            
        except ET.ParseError as e:
            raise ValueError(f"XML 파싱 오류: {str(e)}")
        except Exception as e:
            raise ValueError(f"BW process 파일 처리 중 오류 발생: {str(e)}")
        
        return insert_queries
    
    def _remove_oracle_hints(self, query: str) -> str:
        """
        SQL 쿼리에서 Oracle 힌트(/*+ ... */) 제거
        
        Args:
            query (str): 원본 SQL 쿼리
            
        Returns:
            str: 힌트가 제거된 SQL 쿼리
        """
        # /*+ ... */ 패턴의 힌트 제거
        cleaned_query = re.sub(r'/\*\+[^*]*\*/', '', query)
        # 불필요한 공백 정리 (여러 개의 공백을 하나로)
        cleaned_query = re.sub(r'\s+', ' ', cleaned_query).strip()
        
        if cleaned_query != query:
            print("\n=== Oracle 힌트 제거 ===")
            print(f"원본 쿼리: {query}")
            print(f"정리된 쿼리: {cleaned_query}")
        
        return cleaned_query
    
    def _process_query_parameters(self, activity, query: str) -> str:
        """
        쿼리의 파라미터를 실제 값으로 매핑
        
        Args:
            activity: JDBC 액티비티 XML 요소
            query (str): SQL 쿼리
            
        Returns:
            str: 파라미터가 매핑된 SQL 쿼리
        """
        try:
            # 1단계: prepared_Param_DataType의 파라미터 이름으로 매핑
            param_names = self._get_parameter_names(activity)
            stage1_query = self._replace_with_param_names(query, param_names)
            
            # 2단계: Record의 실제 값으로 매핑
            mappings = self._get_record_mappings(activity, param_names)
            stage2_query = self._replace_with_actual_values(stage1_query, mappings)
            
            return stage2_query
            
        except Exception as e:
            print(f"파라미터 매핑 중 오류 발생: {str(e)}")
            return query  # 오류 발생시 원본 쿼리 반환
    
    def _get_parameter_names(self, activity) -> List[str]:
        """
        Prepared_Param_DataType에서 파라미터 이름 목록 추출
        
        Args:
            activity: JDBC 액티비티 XML 요소
            
        Returns:
            List[str]: 파라미터 이름 목록
        """
        param_names = []
        
        prepared_params = activity.find('.//Prepared_Param_DataType', self.ns)
        if prepared_params is not None:
            for param in prepared_params.findall('./parameter', self.ns):
                param_name = param.find('./parameterName', self.ns)
                if param_name is not None and param_name.text:
                    name = param_name.text.strip()
                    param_names.append(name)
                    print(f"파라미터 이름 추출: {name}")
        
        return param_names
    
    def _replace_with_param_names(self, query: str, param_names: List[str]) -> str:
        """
        SQL 쿼리의 ? 플레이스홀더를 파라미터 이름으로 대체
        
        Args:
            query (str): 원본 SQL 쿼리
            param_names (List[str]): 파라미터 이름 목록
            
        Returns:
            str: 파라미터 이름이 대체된 SQL 쿼리
        """
        parts = query.split('?')
        if len(parts) == 1:  # 플레이스홀더가 없는 경우
            return query
        
        result = parts[0]
        for i, param_name in enumerate(param_names):
            if i < len(parts) - 1:
                result += f":{param_name}" + parts[i+1]
        
        print(f"\n1단계 매핑 결과: {result}")
        return result
    
    def _get_record_mappings(self, activity, param_names: List[str]) -> Dict[str, str]:
        """
        Record 태그에서 실제 값 매핑 정보 추출
        
        Args:
            activity: JDBC 액티비티 XML 요소
            param_names: 파라미터 이름 목록
            
        Returns:
            Dict[str, str]: 파라미터 이름과 매핑된 실제 값의 딕셔너리
        """
        mappings = {}
        
        input_bindings = activity.find('.//pd:inputBindings', self.ns)
        if input_bindings is None:
            return mappings
        
        # jdbcUpdateActivityInput/Record 찾기
        jdbc_input = input_bindings.find('.//jdbcUpdateActivityInput')
        if jdbc_input is None:
            return mappings
        
        # for-each/Record 찾기
        for_each = jdbc_input.find('.//xsl:for-each', self.ns)
        record = for_each.find('./Record') if for_each is not None else jdbc_input
        
        if record is not None:
            # 각 파라미터 이름에 대해 매핑 찾기
            for param_name in param_names:
                param_element = record.find(f'.//{param_name}')
                if param_element is not None:
                    # value-of 체크
                    value_of = param_element.find('.//xsl:value-of', self.ns)
                    if value_of is not None:
                        select_attr = value_of.get('select', '')
                        if select_attr:
                            value = select_attr.split('/')[-1]
                            mappings[param_name] = value
                            print(f"매핑 발견: {param_name} -> {value}")
                    
                    # choose/when 체크
                    else:
                        choose = param_element.find('.//xsl:choose', self.ns)
                        if choose is not None:
                            when = choose.find('.//xsl:when', self.ns)
                            if when is not None:
                                test_attr = when.get('test', '')
                                if 'exists(' in test_attr:
                                    value = test_attr[test_attr.find('(')+1:test_attr.find(')')]
                                    mappings[param_name] = value
                                    print(f"매핑 발견: {param_name} -> {value}")
        
        return mappings
    
    def _replace_with_actual_values(self, query: str, mappings: Dict[str, str]) -> str:
        """
        파라미터 이름을 실제 값으로 대체
        
        Args:
            query (str): 파라미터 이름이 대체된 쿼리
            mappings (Dict[str, str]): 파라미터 이름과 실제 값의 매핑
            
        Returns:
            str: 실제 값이 대체된 SQL 쿼리
        """
        result = query
        
        for param_name, actual_value in mappings.items():
            # 정확한 파라미터 이름만 대체
            result = re.sub(f":{param_name}\\b", f":{actual_value}", result)
        
        print(f"\n2단계 매핑 결과: {result}")
        return result
    
    def get_statistics(self) -> Dict[str, int]:
        """
        마지막 처리 결과의 통계 정보 반환
        
        Returns:
            Dict[str, int]: 처리 통계 정보
        """
        return {
            'processed_count': self.processed_count,
            'error_count': self.error_count,
            'total_attempts': self.processed_count + self.error_count
        }
    
    def get_last_errors(self) -> List[str]:
        """
        마지막 처리에서 발생한 오류 메시지들 반환
        
        Returns:
            List[str]: 오류 메시지 리스트
        """
        return self.last_error_messages.copy()

    def extract_column_mappings(self, process_file_path: str) -> Dict[str, List[str]]:
        """
        BW .process 파일에서 컬럼 매핑 정보를 추출
        
        Args:
            process_file_path (str): .process 파일의 경로
            
        Returns:
            Dict[str, List[str]]: {
                'recv_columns': ['수신컬럼1', '수신컬럼2', ...],     # INSERT 쿼리의 수신 컬럼들
                'send_columns': ['송신컬럼1', '송신컬럼2', ...],     # 매핑된 송신 컬럼들
                'column_mappings': [{'recv': '수신컬럼', 'send': '송신컬럼', 'value_type': 'direct|function|literal'}, ...]
            }
            
        Raises:
            FileNotFoundError: 파일이 존재하지 않는 경우
            ValueError: 파일 형식이 올바르지 않은 경우
        """
        # 파일 존재 여부 확인
        if not os.path.exists(process_file_path):
            raise FileNotFoundError(f"BW process 파일을 찾을 수 없습니다: {process_file_path}")
        
        column_mappings = {
            'recv_columns': [],
            'send_columns': [],
            'column_mappings': []
        }
        
        try:
            # XML 파일 파싱
            tree = ET.parse(process_file_path)
            root = tree.getroot()
            
            print(f"\n=== BW Process 파일 컬럼 매핑 추출 시작: {process_file_path} ===")
            
            # JDBC 액티비티 찾기 (특히 'InsertAll' 타입)
            activities = root.findall('.//pd:activity', self.ns)
            
            for activity in activities:
                try:
                    # JDBC 액티비티 타입 확인
                    activity_type = activity.find('./pd:type', self.ns)
                    if activity_type is None or 'jdbc' not in activity_type.text.lower():
                        continue
                    
                    activity_name = activity.get('name', 'Unknown')
                    print(f"\nJDBC 액티비티 발견: {activity_name}")
                    
                    # 'InsertAll' 액티비티인지 확인
                    if 'InsertAll' not in activity_name and 'insertall' not in activity_name.lower():
                        print(f"'{activity_name}'는 InsertAll이 아니므로 건너뜀")
                        continue
                    
                    print(f"InsertAll 액티비티 발견: {activity_name}")
                    
                    # statement 추출
                    statement = activity.find('.//config/statement')
                    if statement is not None and statement.text:
                        query = statement.text.strip()
                        print(f"\n발견된 INSERT 쿼리:\n{query}")
                        
                        # INSERT 쿼리인 경우만 처리
                        if query.lower().startswith('insert'):
                            # 상세한 컬럼과 값 매핑 추출
                            recv_columns, send_columns, detailed_mappings = self._extract_detailed_column_mapping(activity, query)
                            
                            if recv_columns and send_columns:
                                column_mappings['recv_columns'] = recv_columns
                                column_mappings['send_columns'] = send_columns
                                column_mappings['column_mappings'] = detailed_mappings
                                
                                print(f"\n✅ 추출된 컬럼 매핑 ({len(recv_columns)}개):")
                                for mapping in detailed_mappings:
                                    print(f"  🔸 {mapping['recv']} <- {mapping['send']} ({mapping['value_type']})")
                                break  # 첫 번째 InsertAll 액티비티만 처리
                        
                except Exception as e:
                    print(f"Warning: 액티비티 '{activity.get('name', 'Unknown')}' 처리 중 오류: {str(e)}")
            
            print(f"\n=== 컬럼 매핑 추출 완료 ===")
            
        except ET.ParseError as e:
            raise ValueError(f"XML 파싱 오류: {str(e)}")
        except Exception as e:
            raise ValueError(f"BW process 파일 처리 중 오류 발생: {str(e)}")
        
        return column_mappings
    
    def _extract_detailed_column_mapping(self, activity, query: str) -> tuple[List[str], List[str], List[Dict[str, str]]]:
        """
        INSERT 쿼리와 액티비티에서 상세한 컬럼 매핑을 추출
        
        구조: <pd:inputBindings> -> <jdbcUpdateActivityInput> -> <xsl:for-each> -> <Record> -> <COL1> -> <xsl:value-of select="SEND_COL1"/>
        
        Args:
            activity: JDBC 액티비티 XML 요소
            query (str): INSERT SQL 쿼리
            
        Returns:
            tuple[List[str], List[str], List[Dict[str, str]]]: (수신 컬럼 리스트, 송신 컬럼 리스트, 상세한 컬럼 매핑 리스트)
        """
        recv_columns = []
        send_columns = []
        detailed_mappings = []
        
        try:
            print(f"\n=== 상세한 컬럼 매핑 추출 시작 ===")
            
            # 1단계: INSERT 쿼리에서 수신 컬럼명과 VALUES 구조 추출
            insert_pattern = r'INSERT\s+INTO\s+[\w.]+\s*\(\s*([^)]+)\s*\)\s*VALUES\s*\(\s*([^)]+)\s*\)'
            match = re.search(insert_pattern, query, re.IGNORECASE | re.DOTALL)
            
            if not match:
                print("Warning: INSERT 쿼리 패턴을 찾을 수 없습니다")
                return recv_columns, send_columns, detailed_mappings
            
            columns_part = match.group(1).strip()
            values_part = match.group(2).strip()
            
            # 수신 컬럼명 분리 (공백 제거)
            column_names = [col.strip() for col in columns_part.split(',')]
            # VALUES 부분을 괄호를 고려하여 분리 (함수 처리)
            value_patterns_raw = self._smart_column_split(values_part)
            
            # 함수 패턴을 '?'로 단순화
            value_patterns = []
            for pattern in value_patterns_raw:
                pattern = pattern.strip()
                # 함수 패턴 감지 (TO_DATE, TRIM, NVL 등)
                if self._is_function_pattern(pattern):
                    value_patterns.append('?')  # 함수는 모두 '?'로 단순화
                    print(f"함수 패턴 감지하여 '?'로 변환: {pattern} -> ?")
                else:
                    value_patterns.append(pattern)
            
            print(f"수신 컬럼들: {column_names}")
            print(f"VALUES 패턴들 (원본): {value_patterns_raw}")
            print(f"VALUES 패턴들 (처리후): {value_patterns}")
            
            # 2단계: XML에서 실제 매핑 정보 추출
            # <pd:inputBindings> -> <jdbcUpdateActivityInput> -> <xsl:for-each> -> <Record>
            input_bindings = activity.find('.//pd:inputBindings', self.ns)
            if input_bindings is None:
                print("Warning: pd:inputBindings를 찾을 수 없습니다")
                return recv_columns, send_columns, detailed_mappings
            
            jdbc_input = input_bindings.find('.//jdbcUpdateActivityInput')
            if jdbc_input is None:
                print("Warning: jdbcUpdateActivityInput을 찾을 수 없습니다")
                return recv_columns, send_columns, detailed_mappings
            
            # <xsl:for-each select="$DATA/data/pfx3:TEST_TABLE"> 찾기
            for_each = jdbc_input.find('.//xsl:for-each', self.ns)
            if for_each is None:
                print("Warning: xsl:for-each를 찾을 수 없습니다")
                return recv_columns, send_columns, detailed_mappings
            
            for_each_select = for_each.get('select', '')
            print(f"for-each select: {for_each_select}")
            
            # <Record> 태그 찾기
            record = for_each.find('./Record')
            if record is None:
                print("Warning: Record 태그를 찾을 수 없습니다")
                return recv_columns, send_columns, detailed_mappings
            
            print(f"Record 태그 발견, 하위 요소 개수: {len(list(record))}")
            
            # 3단계: Record 하위의 각 컬럼 매핑 분석
            xml_column_mappings = {}
            
            for child in record:
                if child.tag and child.tag.strip():
                    recv_col = child.tag.strip()
                    
                    # <xsl:value-of select="SEND_COL1"/> 찾기
                    value_of = child.find('.//xsl:value-of', self.ns)
                    if value_of is not None:
                        select_attr = value_of.get('select', '')
                        if select_attr:
                            # select="SEND_COL1" 또는 select="$some/path/SEND_COL1"에서 마지막 부분 추출
                            send_col = select_attr.split('/')[-1].strip()
                            xml_column_mappings[recv_col] = send_col
                            print(f"  XML 매핑: {recv_col} <- {send_col}")
                    
                    # <xsl:choose> 등 다른 구조도 확인
                    elif child.find('.//xsl:choose', self.ns) is not None:
                        xml_column_mappings[recv_col] = f"conditional_{recv_col}"
                        print(f"  XML 매핑: {recv_col} <- conditional (조건부)")
                    
                    # 직접 텍스트 값
                    elif child.text and child.text.strip():
                        xml_column_mappings[recv_col] = f"literal_{child.text.strip()}"
                        print(f"  XML 매핑: {recv_col} <- literal '{child.text.strip()}'")
            
            # 4단계: INSERT 쿼리의 컬럼과 XML 매핑 결합
            for i, recv_col in enumerate(column_names):
                recv_columns.append(recv_col)
                
                # VALUES 패턴 분석
                value_pattern = value_patterns[i] if i < len(value_patterns) else '?'
                
                # 송신 컬럼 결정 로직
                send_col = None
                
                # 1) 먼저 VALUES 패턴이 리터럴 값인지 확인
                if value_pattern.startswith("'") and value_pattern.endswith("'"):
                    # 리터럴 값인 경우: 'N' -> N
                    literal_value = value_pattern[1:-1]  # 따옴표 제거
                    send_col = f"'{literal_value}'"
                    print(f"  리터럴 값 발견: {recv_col} <- {send_col}")
                
                # 2) 리터럴이 아니면 XML 매핑에서 찾기
                else:
                    send_col = xml_column_mappings.get(recv_col)
                    if send_col:
                        print(f"  XML 매핑 사용: {recv_col} <- {send_col}")
                    else:
                        # 3) XML 매핑도 없고 리터럴도 아니면 패턴 분석
                        if value_pattern != '?':
                            # 함수나 다른 패턴이 있는 경우
                            send_col = f"pattern_{value_pattern}"
                            print(f"  패턴 매핑: {recv_col} <- {send_col} (패턴: {value_pattern})")
                        else:
                            # 완전히 알 수 없는 경우
                            send_col = f"unknown_{recv_col}"
                            print(f"  알 수 없는 매핑: {recv_col} <- {send_col}")
                
                send_columns.append(send_col)
                
                # 값 타입 결정
                value_type = self._determine_value_type(value_pattern, send_col)
                
                detailed_mappings.append({
                    'recv': recv_col,
                    'send': send_col,
                    'value_type': value_type,
                    'value_pattern': value_pattern
                })
            
            print(f"\n=== 최종 매핑 결과 ===")
            print(f"수신 컬럼 ({len(recv_columns)}개): {recv_columns}")
            print(f"송신 컬럼 ({len(send_columns)}개): {send_columns}")
            
        except Exception as e:
            print(f"Warning: 상세한 컬럼-값 매핑 추출 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
        
        return recv_columns, send_columns, detailed_mappings
    
    def _determine_value_type(self, value_pattern: str, send_col: str) -> str:
        """
        VALUES 패턴과 송신 컬럼을 분석하여 값 타입을 결정
        
        Args:
            value_pattern (str): VALUES에서의 패턴 (?, TRIM(?), 'N' 등)
            send_col (str): 송신 컬럼명
            
        Returns:
            str: 'direct', 'function', 'literal', 'conditional' 중 하나
        """
        value_pattern = value_pattern.strip()
        
        # 1) 리터럴 값 확인 (가장 우선)
        if value_pattern.startswith("'") and value_pattern.endswith("'"):
            return 'literal'
        
        # 2) 직접 매핑 확인
        elif value_pattern == '?':
            return 'direct'
        
        # 3) 함수 적용 확인
        elif ('TRIM(' in value_pattern.upper() or 
              'UPPER(' in value_pattern.upper() or 
              'LOWER(' in value_pattern.upper() or
              'SUBSTR(' in value_pattern.upper() or
              'NVL(' in value_pattern.upper() or
              'TO_DATE(' in value_pattern.upper() or
              'TO_CHAR(' in value_pattern.upper()):
            return 'function'
        
        # 4) 조건부 확인 (send_col에서 판별)
        elif 'conditional' in send_col:
            return 'conditional'
        
        # 5) 패턴 매핑 확인
        elif send_col.startswith('pattern_'):
            return 'function'
        
        # 6) 알 수 없는 경우
        else:
            return 'unknown'

    def extract_send_column_mappings(self, process_file_path: str) -> Dict[str, List[str]]:
        """
        송신 BW .process 파일에서 SELECT 쿼리의 컬럼 정보를 추출
        
        Args:
            process_file_path (str): .process 파일의 경로
            
        Returns:
            Dict[str, List[str]]: {
                'send_columns': ['SEND_COL1', 'SEND_COL2', ...],  # SELECT 쿼리의 송신 컬럼들
                'table_name': 'AAA_MGR.TABLE_XXX',               # 테이블명
                'where_condition': "TRANSFER_FLAG='P'",          # WHERE 조건
                'order_by': 'SEND_COL1'                         # ORDER BY 절
            }
            
        Raises:
            FileNotFoundError: 파일이 존재하지 않는 경우
            ValueError: 파일 형식이 올바르지 않은 경우
        """
        # 파일 존재 여부 확인
        if not os.path.exists(process_file_path):
            raise FileNotFoundError(f"BW process 파일을 찾을 수 없습니다: {process_file_path}")
        
        column_mappings = {
            'send_columns': [],
            'table_name': '',
            'where_condition': '',
            'order_by': ''
        }
        
        try:
            # XML 파일 파싱
            tree = ET.parse(process_file_path)
            root = tree.getroot()
            
            print(f"\n=== 송신 BW Process 파일 컬럼 추출 시작: {process_file_path} ===")
            
            # SelectP 액티비티 찾기
            activities = root.findall('.//pd:activity', self.ns)
            
            for activity in activities:
                try:
                    activity_name = activity.get('name', 'Unknown')
                    
                    # 'SelectP' 액티비티인지 확인
                    if 'SelectP' not in activity_name and 'selectp' not in activity_name.lower():
                        continue
                    
                    print(f"\nSelectP 액티비티 발견: {activity_name}")
                    
                    # config/statement 추출
                    statement = activity.find('.//config/statement')
                    if statement is not None and statement.text:
                        query = statement.text.strip()
                        print(f"\n발견된 SELECT 쿼리:\n{query}")
                        
                        # SELECT 쿼리인 경우만 처리
                        if query.lower().startswith('select'):
                            # SELECT 쿼리에서 컬럼과 테이블 정보 추출
                            send_columns, table_info = self._parse_select_query(query)
                            
                            if send_columns:
                                column_mappings['send_columns'] = send_columns
                                column_mappings.update(table_info)
                                
                                print(f"\n✅ 추출된 송신 컬럼 ({len(send_columns)}개):")
                                for i, col in enumerate(send_columns, 1):
                                    print(f"  {i}. {col}")
                                print(f"테이블: {table_info.get('table_name', 'Unknown')}")
                                print(f"WHERE: {table_info.get('where_condition', 'None')}")
                                print(f"ORDER BY: {table_info.get('order_by', 'None')}")
                                break  # 첫 번째 SelectP 액티비티만 처리
                        
                except Exception as e:
                    print(f"Warning: 액티비티 '{activity.get('name', 'Unknown')}' 처리 중 오류: {str(e)}")
            
            print(f"\n=== 송신 컬럼 추출 완료 ===")
            
        except ET.ParseError as e:
            raise ValueError(f"XML 파싱 오류: {str(e)}")
        except Exception as e:
            raise ValueError(f"BW process 파일 처리 중 오류 발생: {str(e)}")
        
        return column_mappings
    
    def _parse_select_query(self, query: str) -> tuple[List[str], Dict[str, str]]:
        """
        SELECT 쿼리를 파싱하여 컬럼명과 테이블 정보를 추출
        
        Args:
            query (str): SELECT SQL 쿼리
            
        Returns:
            tuple[List[str], Dict[str, str]]: (컬럼 리스트, 테이블 정보 딕셔너리)
        """
        send_columns = []
        table_info = {
            'table_name': '',
            'where_condition': '',
            'order_by': ''
        }
        
        try:
            print(f"\n=== SELECT 쿼리 파싱 시작 ===")
            
            # Oracle 힌트 제거
            cleaned_query = self._remove_oracle_hints(query)
            print(f"힌트 제거된 쿼리:\n{cleaned_query}")
            
            # 1단계: SELECT 컬럼 부분 추출
            # SELECT ... FROM 사이의 컬럼들 추출
            select_pattern = r'SELECT\s+(.*?)\s+FROM'
            select_match = re.search(select_pattern, cleaned_query, re.IGNORECASE | re.DOTALL)
            
            if select_match:
                columns_part = select_match.group(1).strip()
                print(f"컬럼 부분: {columns_part}")
                
                # 괄호를 고려한 정교한 컬럼명 분리
                column_lines = self._smart_column_split(columns_part)
                for col_line in column_lines:
                    # 각 라인에서 실제 컬럼명 추출 (AS 별칭 등 제거)
                    col_name = self._extract_column_name(col_line)
                    if col_name:
                        send_columns.append(col_name)
                        print(f"  추출된 컬럼: {col_name} (원본: {col_line.strip()})")
            
            # 2단계: FROM 절에서 테이블명 추출
            from_pattern = r'FROM\s+([\w.]+)'
            from_match = re.search(from_pattern, cleaned_query, re.IGNORECASE)
            if from_match:
                table_info['table_name'] = from_match.group(1).strip()
                print(f"테이블명: {table_info['table_name']}")
            
            # 3단계: WHERE 절 추출
            where_pattern = r'WHERE\s+(.*?)(?:\s+ORDER\s+BY|$)'
            where_match = re.search(where_pattern, cleaned_query, re.IGNORECASE | re.DOTALL)
            if where_match:
                table_info['where_condition'] = where_match.group(1).strip()
                print(f"WHERE 조건: {table_info['where_condition']}")
            
            # 4단계: ORDER BY 절 추출
            order_pattern = r'ORDER\s+BY\s+(.*?)$'
            order_match = re.search(order_pattern, cleaned_query, re.IGNORECASE | re.DOTALL)
            if order_match:
                table_info['order_by'] = order_match.group(1).strip()
                print(f"ORDER BY: {table_info['order_by']}")
            
            print(f"\n=== SELECT 쿼리 파싱 완료 ===")
            
        except Exception as e:
            print(f"Warning: SELECT 쿼리 파싱 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
        
        return send_columns, table_info
    
    def _extract_column_name(self, column_expression: str) -> str:
        """
        컬럼 표현식에서 실제 컬럼명을 추출
        
        Args:
            column_expression (str): 컬럼 표현식 (예: "SEND_COL1", "TO_CHAR(CREATION_DATE, 'YYYYMMDDHH24MISS') CREATION_DATE")
            
        Returns:
            str: 추출된 컬럼명
        """
        column_expression = column_expression.strip()
        
        # 1단계: AS 키워드가 있는 경우 별칭 추출
        if ' AS ' in column_expression.upper():
            alias_part = column_expression.upper().split(' AS ')
            if len(alias_part) >= 2:
                alias = alias_part[1].strip()
                print(f"    AS 별칭 발견: {alias}")
                return alias
        
        # 2단계: AS 없이 공백으로 구분된 별칭 확인
        # 함수나 연산자가 포함된 경우 마지막 단어가 별칭일 가능성 높음
        if any(char in column_expression for char in ['(', ')', '+', '-', '*', '/', '||']):
            # 공백으로 분리하여 마지막 부분이 별칭인지 확인
            parts = column_expression.split()
            if len(parts) >= 2:
                # 마지막 부분이 괄호나 연산자를 포함하지 않으면 별칭으로 간주
                last_part = parts[-1].strip()
                if not any(char in last_part for char in ['(', ')', '+', '-', '*', '/', '||', "'", '"']):
                    print(f"    공백 별칭 발견: {last_part}")
                    return last_part
        
        # 3단계: 단순한 함수 패턴 처리 (예: UPPER(COLUMN_NAME))
        simple_func_pattern = r'^[A-Z_]+\s*\(\s*([\w.]+)\s*\)$'
        simple_func_match = re.search(simple_func_pattern, column_expression, re.IGNORECASE)
        if simple_func_match:
            inner_column = simple_func_match.group(1).strip()
            print(f"    단순 함수 패턴: {inner_column}")
            return inner_column
        
        # 4단계: 복잡한 함수 패턴에서 첫 번째 컬럼명 추출
        # TO_CHAR(CREATION_DATE, 'FORMAT') 같은 경우 CREATION_DATE 추출
        complex_func_pattern = r'[A-Z_]+\s*\(\s*([\w.]+)\s*,'
        complex_func_match = re.search(complex_func_pattern, column_expression, re.IGNORECASE)
        if complex_func_match:
            inner_column = complex_func_match.group(1).strip()
            print(f"    복잡 함수 첫 번째 인자: {inner_column}")
            return inner_column
        
        # 5단계: 일반적인 컬럼명 (스키마.테이블.컬럼 또는 테이블.컬럼 또는 컬럼)
        # 공백이 없고 단순한 컬럼명인 경우
        if ' ' not in column_expression and not any(char in column_expression for char in ['(', ')', "'", '"']):
            parts = column_expression.split('.')
            final_column = parts[-1].strip()
            print(f"    단순 컬럼명: {final_column}")
            return final_column
        
        # 6단계: 위의 모든 패턴에 해당하지 않는 경우
        # 첫 번째 단어를 반환 (최후의 수단)
        first_word = column_expression.split()[0] if column_expression.split() else column_expression
        # 괄호나 연산자 제거
        first_word = re.sub(r'[()\'"+\-*/]', '', first_word)
        print(f"    기본 추출: {first_word}")
        return first_word.strip()
    
    def _smart_column_split(self, columns_part: str) -> List[str]:
        """
        괄호를 고려하여 컬럼들을 올바르게 분리
        
        Args:
            columns_part (str): SELECT절의 컬럼 부분
            
        Returns:
            List[str]: 올바르게 분리된 컬럼 리스트
        """
        columns = []
        current_column = ""
        paren_depth = 0
        quote_char = None
        
        i = 0
        while i < len(columns_part):
            char = columns_part[i]
            
            # 따옴표 처리 (작은따옴표, 큰따옴표)
            if char in ["'", '"'] and quote_char is None:
                quote_char = char
                current_column += char
            elif char == quote_char:
                quote_char = None
                current_column += char
            elif quote_char is not None:
                # 따옴표 안에서는 모든 문자를 그대로 추가
                current_column += char
            elif char == '(':
                # 괄호 깊이 증가
                paren_depth += 1
                current_column += char
            elif char == ')':
                # 괄호 깊이 감소
                paren_depth -= 1
                current_column += char
            elif char == ',' and paren_depth == 0:
                # 괄호 밖의 콤마만 구분자로 인식
                if current_column.strip():
                    columns.append(current_column.strip())
                current_column = ""
            else:
                current_column += char
            
            i += 1
        
        # 마지막 컬럼 추가
        if current_column.strip():
            columns.append(current_column.strip())
        
        print(f"스마트 컬럼 분리 결과: {len(columns)}개")
        for i, col in enumerate(columns, 1):
            print(f"  {i}. {col}")
        
        return columns

    def extract_schema_columns(self, schema_file_path: str) -> Dict[str, List[str]]:
        """
        XSD 스키마 파일에서 컬럼 정보를 추출
        
        Args:
            schema_file_path (str): .xsd 스키마 파일의 경로
            
        Returns:
            Dict[str, List[str]]: {
                'schema_columns': ['SEND_COL1', 'SEND_COL2', ...],  # xs:element의 name 속성들
                'file_exists': True/False,
                'error': None 또는 오류 메시지
            }
            
        Raises:
            FileNotFoundError: 파일이 존재하지 않는 경우
            ValueError: 파일 형식이 올바르지 않은 경우
        """
        result = {
            'schema_columns': [],
            'file_exists': False,
            'error': None
        }
        
        # 파일 존재 여부 확인
        if not schema_file_path or not schema_file_path.strip():
            result['error'] = "스키마 파일 경로가 비어있음"
            return result
            
        if not os.path.exists(schema_file_path):
            result['error'] = f"스키마 파일 없음: {schema_file_path}"
            print(f"Warning: {result['error']}")
            return result
        
        result['file_exists'] = True
        
        try:
            print(f"\n=== 스키마 파일 컬럼 추출 시작: {schema_file_path} ===")
            
            # XML 파일 파싱
            tree = ET.parse(schema_file_path)
            root = tree.getroot()
            
            # XML 네임스페이스 처리
            namespaces = {}
            # 기본 XML 스키마 네임스페이스
            namespaces['xs'] = 'http://www.w3.org/2001/XMLSchema'
            
            # 루트에서 네임스페이스 정보 추출
            for prefix, uri in root.nsmap.items() if hasattr(root, 'nsmap') else {}:
                if prefix:
                    namespaces[prefix] = uri
                else:
                    # 기본 네임스페이스
                    namespaces['default'] = uri
            
            print(f"네임스페이스: {namespaces}")
            
            # xs:sequence 하위의 xs:element 찾기
            schema_columns = []
            
            # 다양한 패턴으로 xs:element 검색
            element_patterns = [
                './/xs:element[@name]',  # xs: 접두사 사용
                './/element[@name]',    # 접두사 없음
                './/*[local-name()="element"][@name]'  # local-name 사용
            ]
            
            for pattern in element_patterns:
                try:
                    elements = root.findall(pattern, namespaces)
                    if elements:
                        print(f"패턴 '{pattern}'로 {len(elements)}개 요소 발견")
                        for element in elements:
                            name_attr = element.get('name')
                            if name_attr and name_attr.strip():
                                column_name = name_attr.strip()
                                if column_name not in schema_columns:
                                    schema_columns.append(column_name)
                                    print(f"  스키마 컬럼: {column_name}")
                        break  # 첫 번째로 성공한 패턴 사용
                except Exception as e:
                    print(f"패턴 '{pattern}' 검색 실패: {str(e)}")
                    continue
            
            # 네임스페이스 없이도 시도
            if not schema_columns:
                print("네임스페이스 검색 실패, 직접 검색 시도")
                for element in root.iter():
                    tag_name = element.tag.split('}')[-1] if '}' in element.tag else element.tag
                    if tag_name == 'element' and element.get('name'):
                        name_attr = element.get('name')
                        if name_attr and name_attr.strip():
                            column_name = name_attr.strip()
                            if column_name not in schema_columns:
                                schema_columns.append(column_name)
                                print(f"  스키마 컬럼 (직접검색): {column_name}")
            
            result['schema_columns'] = schema_columns
            
            print(f"\n✅ 스키마 파일에서 {len(schema_columns)}개 컬럼 추출 완료")
            print(f"추출된 컬럼: {schema_columns}")
            print(f"=== 스키마 컬럼 추출 완료 ===")
            
        except ET.ParseError as e:
            result['error'] = f"XML 파싱 오류: {str(e)}"
            print(f"Error: {result['error']}")
        except Exception as e:
            result['error'] = f"스키마 파일 처리 중 오류 발생: {str(e)}"
            print(f"Error: {result['error']}")
            import traceback
            traceback.print_exc()
        
        return result


class ProcessFileMapper:
    """
    test_iflist.py의 일련번호와 string_replacer용 엑셀을 매핑하는 클래스
    """
    
    def __init__(self, replacer_excel_path: str):
        """ProcessFileMapper 초기화
        
        Args:
            replacer_excel_path (str): string_replacer.py에서 사용하는 엑셀 파일 경로
        """
        self.replacer_excel_path = replacer_excel_path
        self.df = None
        if os.path.exists(replacer_excel_path):
            try:
                self.df = pd.read_excel(replacer_excel_path, engine='openpyxl')
            except Exception as e:
                print(f"Warning: ProcessFileMapper - 엑셀 파일 로드 실패: {str(e)}")
    
    def get_process_files_by_serial(self, serial_number: str) -> Dict[str, str]:
        """
        일련번호(serial_number)로 .process 파일 경로들을 가져옴
        
        Args:
            serial_number (str): 인터페이스 일련번호
            
        Returns:
            Dict[str, str]: 프로세스 파일 정보
            {
                'send_original': '송신 원본파일 경로',
                'send_copy': '송신 복사파일 경로', 
                'recv_original': '수신 원본파일 경로',
                'recv_copy': '수신 복사파일 경로',
                'send_schema': '송신 스키마파일',
                'recv_schema': '수신 스키마파일'
            }
        """
        print(f"ProcessFileMapper.get_process_files_by_serial 호출됨")
        print(f"입력 일련번호: '{serial_number}'")
        print(f"DataFrame 상태: {self.df is not None}")
        
        if self.df is None or not serial_number:
            print(f"조기 반환: DataFrame={self.df is not None}, serial_number='{serial_number}'")
            return {}
        
        try:
            # N번째 행 = serial_number 매핑 (1-based to 0-based)
            row_index = int(serial_number) - 1
            print(f"계산된 row_index: {row_index}")
            print(f"DataFrame 크기: {len(self.df)}")
            
            if row_index * 2 + 1 >= len(self.df):
                print(f"행 인덱스 초과: {row_index * 2 + 1} >= {len(self.df)}")
                return {}
            
            normal_row = self.df.iloc[row_index * 2]     # 기본행
            match_row = self.df.iloc[row_index * 2 + 1]  # 매칭행
            
            print(f"기본행 인덱스: {row_index * 2}")
            print(f"매칭행 인덱스: {row_index * 2 + 1}")
            print(f"기본행 데이터: {normal_row.to_dict()}")
            print(f"매칭행 데이터: {match_row.to_dict()}")
            
            result = {}
            
            # 송신 파일 생성 여부 확인
            send_create_flag = normal_row.get('송신파일생성여부')
            print(f"송신파일생성여부: {send_create_flag} (타입: {type(send_create_flag)})")
            if (pd.notna(send_create_flag) and float(send_create_flag) == 1.0):
                result['send_original'] = str(match_row.get('송신파일경로', ''))
                send_copy_path = str(normal_row.get('송신파일경로', ''))
                
                # ===== 테스트용 임시 경로 변환 로직 (향후 주석 처리) =====
                # 'C:\BwProject\' -> 'C:\tBwProject\'로 변경
                if send_copy_path.startswith('C:\\BwProject\\'):
                    send_copy_path = send_copy_path.replace('C:\\BwProject\\', 'C:\\tBwProject\\', 1)
                    print(f"테스트용 송신 경로 변환: {normal_row.get('송신파일경로', '')} -> {send_copy_path}")
                # ===========================================================
                
                result['send_copy'] = send_copy_path
                print(f"송신 파일 정보 추가됨")
            
            # 수신 파일 생성 여부 확인  
            recv_create_flag = normal_row.get('수신파일생성여부')
            print(f"수신파일생성여부: {recv_create_flag} (타입: {type(recv_create_flag)})")
            if (pd.notna(recv_create_flag) and float(recv_create_flag) == 1.0):
                result['recv_original'] = str(match_row.get('수신파일경로', ''))
                recv_copy_path = str(normal_row.get('수신파일경로', ''))
                
                # ===== 테스트용 임시 경로 변환 로직 (향후 주석 처리) =====
                # 'C:\BwProject\' -> 'C:\tBwProject\'로 변경
                if recv_copy_path.startswith('C:\\BwProject\\'):
                    recv_copy_path = recv_copy_path.replace('C:\\BwProject\\', 'C:\\tBwProject\\', 1)
                    print(f"테스트용 수신 경로 변환: {normal_row.get('수신파일경로', '')} -> {recv_copy_path}")
                # ===========================================================
                
                result['recv_copy'] = recv_copy_path
                print(f"수신 파일 정보 추가됨")
            
            # 송신 스키마 파일 생성 여부 확인
            send_schema_flag = normal_row.get('송신스키마파일생성여부')
            print(f"송신스키마파일생성여부: {send_schema_flag} (타입: {type(send_schema_flag)})")
            if (pd.notna(send_schema_flag) and float(send_schema_flag) == 1.0):
                send_schema_path = str(normal_row.get('송신스키마파일명', ''))
                
                # ===== 테스트용 임시 경로 변환 로직 (향후 주석 처리) =====
                # 'C:\BwProject\' -> 'C:\tBwProject\'로 변경
                if send_schema_path.startswith('C:\\BwProject\\'):
                    send_schema_path = send_schema_path.replace('C:\\BwProject\\', 'C:\\tBwProject\\', 1)
                    print(f"테스트용 송신 스키마 경로 변환: {normal_row.get('송신스키마파일명', '')} -> {send_schema_path}")
                # ===========================================================
                
                result['send_schema'] = send_schema_path
                print(f"송신 스키마 파일 정보 추가됨")
            
            # 수신 스키마 파일 생성 여부 확인
            recv_schema_flag = normal_row.get('수신스키마파일생성여부')
            print(f"수신스키마파일생성여부: {recv_schema_flag} (타입: {type(recv_schema_flag)})")
            if (pd.notna(recv_schema_flag) and float(recv_schema_flag) == 1.0):
                recv_schema_path = str(normal_row.get('수신스키마파일명', ''))
                
                # ===== 테스트용 임시 경로 변환 로직 (향후 주석 처리) =====
                # 'C:\BwProject\' -> 'C:\tBwProject\'로 변경
                if recv_schema_path.startswith('C:\\BwProject\\'):
                    recv_schema_path = recv_schema_path.replace('C:\\BwProject\\', 'C:\\tBwProject\\', 1)
                    print(f"테스트용 수신 스키마 경로 변환: {normal_row.get('수신스키마파일명', '')} -> {recv_schema_path}")
                # ===========================================================
                
                result['recv_schema'] = recv_schema_path
                print(f"수신 스키마 파일 정보 추가됨")
            
            print(f"최종 결과: {result}")
            return result
            
        except Exception as e:
            print(f"Warning: ProcessFileMapper - 일련번호 {serial_number} 처리 실패: {str(e)}")
            return {}


def parse_bw_receive_file(process_file_path: str) -> List[str]:
    """
    BW의 수신파일(.process)을 파싱하여 INSERT 쿼리를 추출하는 함수
    
    Args:
        process_file_path (str): BW .process 파일의 경로
        
    Returns:
        List[str]: 추출된 INSERT 쿼리 목록
        
    Raises:
        FileNotFoundError: 파일이 존재하지 않는 경우
        ValueError: 파일 형식이 올바르지 않은 경우
    """
    parser = BWProcessFileParser()
    return parser.parse_bw_process_file(process_file_path)


# 사용 예시 및 테스트
if __name__ == "__main__":
    # 테스트용 샘플 코드
    def test_interface_reader():
        """InterfaceExcelReader 테스트 함수"""
        # ProcessFileMapper용 엑셀 파일 경로 (None이면 기본값 사용)
        # replacer_excel_path = None  # 기본값 'iflist03a_reordered_v8.3.xlsx' 사용
        reader = InterfaceExcelReader()  # 기본값으로 'iflist03a_reordered_v8.3.xlsx' 사용
        
        # 테스트할 인터페이스 엑셀 파일 경로
        test_excel_path = "iflist_in.xlsx"  # 인터페이스 정보가 담긴 파일
        
        try:
            print("=== 인터페이스 엑셀 리더 테스트 시작 ===")
            print(f"인터페이스 정보 파일: {test_excel_path}")
            print(f"ProcessFileMapper 파일: iflist03a_reordered_v8.3.xlsx")
            
            # 파일 존재 여부 확인
            if not os.path.exists(test_excel_path):
                print(f"테스트 파일이 없습니다: {test_excel_path}")
                print("테스트를 위해 실제 엑셀 파일 경로를 지정해주세요.")
                return
            
            # 인터페이스 정보 로드
            interfaces = reader.load_interfaces(test_excel_path)
            
            # 결과 출력
            print(f"\n=== 처리 결과 ===")
            print(f"총 {len(interfaces)}개의 인터페이스를 읽었습니다.")
            
            # 통계 정보 출력
            stats = reader.get_statistics()
            print(f"처리 성공: {stats['processed_count']}개")
            print(f"처리 실패: {stats['error_count']}개")
            
            # 오류가 있었다면 출력
            errors = reader.get_last_errors()
            if errors:
                print(f"\n=== 발생한 오류들 ===")
                for error in errors:
                    print(f"- {error}")
            
            # 첫 번째 인터페이스 정보 샘플 출력
            if interfaces:
                print(f"\n=== 첫 번째 인터페이스 샘플 ===")
                first_interface = interfaces[0]
                print(f"인터페이스명: {first_interface['interface_name']}")
                print(f"인터페이스ID: {first_interface['interface_id']}")
                print(f"일련번호: {first_interface['serial_number']}")
                print(f"송신 테이블: {first_interface['send']['table_name']}")
                print(f"수신 테이블: {first_interface['recv']['table_name']}")
                print(f"송신 컬럼 수: {len(first_interface['send']['columns'])}")
                print(f"수신 컬럼 수: {len(first_interface['recv']['columns'])}")
                print(f"송신 원본파일: {first_interface.get('send_original', 'N/A')}")
                print(f"송신 복사파일: {first_interface.get('send_copy', 'N/A')}")
                print(f"수신 원본파일: {first_interface.get('recv_original', 'N/A')}")
                print(f"수신 복사파일: {first_interface.get('recv_copy', 'N/A')}")
                print(f"송신 스키마파일: {first_interface.get('send_schema', 'N/A')}")
                print(f"수신 스키마파일: {first_interface.get('recv_schema', 'N/A')}")
                
                # 컬럼 매핑 비교 수행
                print(f"\n=== 컬럼 매핑 비교 수행 ===")
                try:
                    comparison_result = reader.compare_column_mappings(first_interface)
                    
                    # 비교 결과 요약 출력
                    send_comp = comparison_result['send_comparison']
                    if send_comp.get('file_exists'):
                        print(f"송신 매칭률: {send_comp['match_percentage']:.1f}%")
                        print(f"매칭된 컬럼: {send_comp['matches']}")
                        print(f"엑셀에만 있는 컬럼: {send_comp['excel_only']}")
                    
                    recv_comp = comparison_result['recv_comparison']
                    if recv_comp.get('file_exists'):
                        print(f"수신 매칭률: {recv_comp['match_percentage']:.1f}%")
                    
                except Exception as e:
                    print(f"컬럼 매핑 비교 중 오류: {str(e)}")
            
            # 전체 인터페이스 정보를 로그 파일로 출력
            print(f"\n=== 전체 인터페이스 정보 로그 출력 ===")
            reader.export_all_interfaces_to_log(interfaces)
            
            # 비교 결과 요약을 엑셀 파일로 출력
            print(f"\n=== 비교 결과 요약 엑셀 출력 ===")
            reader.export_summary_to_excel(interfaces, "test_iflist_result.xlsx")
            
            print("\n=== 테스트 완료 ===")
        
        except FileNotFoundError as e:
            print(f"파일 오류: {e}")
        except ValueError as e:
            print(f"데이터 오류: {e}")
        except Exception as e:
            print(f"예상치 못한 오류: {e}")
    
    # 간단한 사용법 예시
    def usage_example():
        """모듈 사용법 예시"""
        print("\n=== 사용법 예시 ===")
        print("""
# 1. InterfaceExcelReader 인스턴스 생성
# 기본값으로 'iflist03a_reordered_v8.3.xlsx' 파일을 ProcessFileMapper로 사용
reader = InterfaceExcelReader()

# 또는 특정 ProcessFileMapper 파일 지정
# reader = InterfaceExcelReader('custom_replacer_file.xlsx')

# 2. 인터페이스 정보 엑셀 파일에서 정보 로드
# 'iflist_in.xlsx'는 인터페이스 정보가 담긴 파일
interfaces = reader.load_interfaces('iflist_in.xlsx')

# 3. 결과 활용
for interface in interfaces:
    print(f"인터페이스: {interface['interface_name']}")
    print(f"ID: {interface['interface_id']}")
    print(f"일련번호: {interface['serial_number']}")
    print(f"송신 테이블: {interface['send']['table_name']}")
    print(f"수신 테이블: {interface['recv']['table_name']}")
    print(f"송신 원본파일: {interface.get('send_original', 'N/A')}")
    print(f"수신 복사파일: {interface.get('recv_copy', 'N/A')}")

# 4. 전체 결과 로그 출력
reader.export_all_interfaces_to_log(interfaces)

# 5. 비교 결과 요약을 엑셀 파일로 출력 (새로운 기능!)
reader.export_summary_to_excel(interfaces, "test_iflist_result.xlsx")
# 출력 컬럼: 일련번호, 인터페이스명, ID, 송신DB, 수신DB, 송신테이블, 수신테이블
# 5가지 비교결과 (송신, 수신, 연결, 송신스키마, 수신스키마) - 각각 매칭률과 결과요약
# 매칭률 100% 미만인 셀은 주황색으로 표시

# 6. 처리 통계 확인
stats = reader.get_statistics()
print(f"처리된 인터페이스 수: {stats['processed_count']}")

# 7. BW 수신파일(.process) 파싱
insert_queries = parse_bw_receive_file('your_bw_file.process')
for query in insert_queries:
    print(f"추출된 INSERT 쿼리: {query}")

# 8. BWProcessFileParser 클래스 직접 사용
bw_parser = BWProcessFileParser()
queries = bw_parser.parse_bw_process_file('your_bw_file.process')
bw_stats = bw_parser.get_statistics()
print(f"BW 파싱 통계: {bw_stats}")

# 9. 컬럼 매핑 비교 (새로운 기능!)
# 엑셀의 송신/수신 컬럼과 .process 파일의 컬럼 매핑을 비교
for interface in interfaces:
    comparison_result = reader.compare_column_mappings(interface)
    
    # 송신 비교 결과
    send_comp = comparison_result['send_comparison']
    if send_comp.get('file_exists'):
        print(f"송신 매칭률: {send_comp['match_percentage']:.1f}%")
        print(f"매칭된 컬럼: {send_comp['matches']}")
        print(f"엑셀에만 있는 컬럼: {send_comp['excel_only']}")
    
    # 수신 비교 결과
    recv_comp = comparison_result['recv_comparison']
    if recv_comp.get('file_exists'):
        print(f"수신 매칭률: {recv_comp['match_percentage']:.1f}%")

# 10. .process 파일에서 직접 컬럼 매핑 추출 (개선된 기능!)
bw_parser = BWProcessFileParser()
column_mappings = bw_parser.extract_column_mappings('path/to/your.process')
print(f"수신 컬럼: {column_mappings['recv_columns']}")
print(f"송신 컬럼: {column_mappings['send_columns']}")
print(f"상세 매핑: {column_mappings['column_mappings']}")

# 상세 매핑 정보 활용
for mapping in column_mappings['column_mappings']:
    print(f"  {mapping['recv']} <- {mapping['send']} ({mapping['value_type']})")
        
# 11. 송신 .process 파일에서 SELECT 컬럼 추출 (새로운 기능!)
# 송신 .process 파일에서 SelectP 액티비티의 SELECT 쿼리 컬럼들을 추출
send_column_mappings = bw_parser.extract_send_column_mappings('path/to/send.process')
print(f"송신 SELECT 컬럼: {send_column_mappings['send_columns']}")
print(f"테이블명: {send_column_mappings['table_name']}")
print(f"WHERE 조건: {send_column_mappings['where_condition']}")
print(f"ORDER BY: {send_column_mappings['order_by']}")

# 12. 개선된 컬럼 매핑 비교 (3단계 비교!)
# - 송신: 엑셀 송신 컬럼 vs .process SELECT 컬럼
# - 수신: 엑셀 수신 컬럼 vs .process INSERT 컬럼  
# - 연결: 엑셀 송신-수신 매핑 쌍 vs .process 송신-수신 매핑 쌍
for interface in interfaces:
    comparison_result = reader.compare_column_mappings(interface)
    
    # 송신 비교 결과 (새로 추가!)
    send_comp = comparison_result['send_comparison']
    if send_comp.get('file_exists'):
        print(f"송신 매칭률: {send_comp['match_percentage']:.1f}%")
        print(f"송신 테이블: {send_comp.get('table_info', {}).get('table_name', 'Unknown')}")
    
    # 수신 비교 결과 (기존)
    recv_comp = comparison_result['recv_comparison']
    if recv_comp.get('file_exists'):
        print(f"수신 매칭률: {recv_comp['match_percentage']:.1f}%")
    
    # 송신-수신 연결 비교 결과 (개선됨!)
    conn_comp = comparison_result['send_recv_comparison']
    if conn_comp.get('recv_file_exists'):
        print(f"송신-수신 매핑 쌍 매칭률: {conn_comp['match_percentage']:.1f}%")
        print(f"엑셀 매핑 쌍 수: {conn_comp['total_excel_pairs']}개")
        print(f"Process 매핑 쌍 수: {conn_comp['total_process_pairs']}개")
        print(f"매칭된 쌍: {len(conn_comp['matches'])}개")

# 파일 구조:
# - iflist_in.xlsx: 인터페이스 정보 엑셀 (B열부터 3컬럼 단위)
# - iflist03a_reordered_v8.3.xlsx: ProcessFileMapper용 파일 (원본파일, 복사파일 정보)
# - 송신 .process: SelectP 액티비티에 SELECT 쿼리 포함
# - 수신 .process: InsertAll 액티비티에 INSERT 쿼리 및 컬럼 매핑 포함

# 13. 스키마 파일과 .process 파일 비교 (새로운 기능!)
# XSD 스키마 파일의 xs:element name 속성과 수신 .process의 송신 컬럼을 비교
for interface in interfaces:
    schema_comparison_result = reader.compare_schema_mappings(interface)
    
    # 송신 스키마 비교 결과
    send_schema_comp = schema_comparison_result['send_schema_comparison']
    if send_schema_comp.get('file_exists'):
        print(f"송신 스키마 매칭률: {send_schema_comp['match_percentage']:.1f}%")
        print(f"스키마 컬럼 수: {send_schema_comp['total_schema']}개")
    
    # 수신 스키마 비교 결과
    recv_schema_comp = schema_comparison_result['recv_schema_comparison']
    if recv_schema_comp.get('file_exists'):
        print(f"수신 스키마 매칭률: {recv_schema_comp['match_percentage']:.1f}%")

# 14. 스키마 파일에서 직접 컬럼 추출 (새로운 기능!)
# XSD 스키마 파일에서 xs:element의 name 속성들을 추출
bw_parser = BWProcessFileParser()
schema_result = bw_parser.extract_schema_columns('path/to/schema.xsd')
if schema_result.get('file_exists'):
    print(f"스키마 컬럼: {schema_result['schema_columns']}")
else:
    print(f"스키마 파일 오류: {schema_result.get('error', '알 수 없음')}")

# 파일 구조:
# - iflist_in.xlsx: 인터페이스 정보 엑셀 (B열부터 3컬럼 단위)
# - iflist03a_reordered_v8.3.xlsx: ProcessFileMapper용 파일 (원본파일, 복사파일 정보)
# - 송신 .process: SelectP 액티비티에 SELECT 쿼리 포함
# - 수신 .process: InsertAll 액티비티에 INSERT 쿼리 및 컬럼 매핑 포함
# - 송신/수신 .xsd: XML 스키마 파일 (xs:element name 속성에 컬럼명 포함)
        """)
    
    # BW Process 파일 파싱 테스트 함수 추가
    def test_bw_process_parser():
        """BWProcessFileParser 테스트 함수"""
        print("\n=== BW Process 파일 파서 테스트 시작 ===")
        
        # 테스트할 .process 파일 경로 (실제 환경에 맞게 수정 필요)
        test_process_path = "sample.process"
        
        try:
            if not os.path.exists(test_process_path):
                print(f"테스트 파일이 없습니다: {test_process_path}")
                print("테스트를 위해 실제 .process 파일 경로를 지정해주세요.")
                return
            
            # BW 수신파일 파싱
            insert_queries = parse_bw_receive_file(test_process_path)
            
            # 결과 출력
            print(f"\n=== 처리 결과 ===")
            print(f"총 {len(insert_queries)}개의 INSERT 쿼리를 추출했습니다.")
            
            # 추출된 쿼리들 출력
            for i, query in enumerate(insert_queries, 1):
                print(f"\n=== INSERT 쿼리 {i} ===")
                print(query)
            
            print("\n=== BW Process 파일 파싱 테스트 완료 ===")
            
        except FileNotFoundError as e:
            print(f"파일 오류: {e}")
        except ValueError as e:
            print(f"데이터 오류: {e}")
        except Exception as e:
            print(f"예상치 못한 오류: {e}")
    
    # 테스트 실행
    test_interface_reader()
    usage_example()
    
    # 새로운 BW Process 파서 테스트 실행
    test_bw_process_parser()